"""Trusted inspector registry, CLI, and Artifact MCP tests."""
from __future__ import annotations

import base64
import concurrent.futures
import errno
import hashlib
import json
import multiprocessing
import os
import shutil
import signal
import struct
import subprocess
import sys
import tempfile
import time
import types
import zipfile
import zlib
from pathlib import Path

import pytest
from mcp import ClientSession, StdioServerParameters
from mcp.client.stdio import stdio_client
from mcp.shared.memory import create_connected_server_and_client_session
from mcp.types import ImageContent, TextContent
from PIL import Image as PillowImage

from skillopt.envs.skilleval import artifact_mcp
from skillopt.envs.skilleval import inspectors as inspectors_mod
from skillopt.envs.skilleval.inspectors import _process as process_mod
from skillopt.envs.skilleval.inspectors import _scratch as scratch_mod
from skillopt.envs.skilleval.inspectors import _scratch_root as scratch_root_mod
from skillopt.envs.skilleval.inspectors import base as inspector_base
from skillopt.envs.skilleval.inspectors import (
    InspectionError,
    extract_artifact,
    inspect_artifact,
    inventory_artifacts,
    render_artifact,
)
from skillopt.envs.skilleval.inspectors.__main__ import main as artifactctl_main
from skillopt.envs.skilleval.inspectors._scratch import scratch_transaction
from skillopt.envs.skilleval.inspectors._secure_files import open_evidence_file
from skillopt.envs.skilleval.inspectors.base import (
    MAX_COMMAND_OUTPUT_CHARS,
    MAX_LOGICAL_COMPONENTS,
    MAX_LOGICAL_COMPONENT_BYTES,
    MAX_LOGICAL_PATH_BYTES,
    MAX_LOGICAL_PATH_CHARS,
    MAX_RENDER_PIXELS,
    MAX_RESPONSE_BYTES,
    MIN_RESPONSE_BYTES,
    RenderBudget,
    ResponseBudget,
    normalize_selectors,
    safe_run,
    validate_logical_path,
    validate_roots,
)


@pytest.fixture
def anyio_backend():
    return "asyncio"


def _roots(tmp_path: Path) -> tuple[Path, Path]:
    evidence = tmp_path / "evidence"
    scratch = tmp_path / "scratch"
    evidence.mkdir()
    scratch.mkdir()
    return evidence, scratch


def test_package_metadata_targets_stable_mcp_and_python_310() -> None:
    root = Path(__file__).resolve().parents[1]
    pyproject = (root / "pyproject.toml").read_text(encoding="utf-8")
    assert 'requires-python = ">=3.10"' in pyproject
    assert '"mcp>=1.26,<2"' in pyproject
    requirements = (root / "requirements.txt").read_text(encoding="utf-8")
    assert "\nmcp>=1.26,<2\n" in f"\n{requirements}"


def _write_pdf(path: Path, body: bytes = b"content") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"%PDF-1.4\n" + body)


def _scratch_commit_worker(
    scratch: str,
    barrier,
    entered,
    results,
) -> None:
    try:
        barrier.wait(timeout=5)
        with scratch_transaction(
            scratch,
            max_bytes=100,
            max_entries=20,
            max_depth=4,
        ) as transaction:
            with entered.get_lock():
                entered.value += 1
            deadline = time.monotonic() + 0.5
            while entered.value < 2 and time.monotonic() < deadline:
                time.sleep(0.01)
            output = Path(transaction.proc_path) / "render.bin"
            output.write_bytes(b"x" * 80)
            transaction.commit_outputs([str(output)])
        results.put("ok")
    except BaseException as exc:
        results.put(f"error:{type(exc).__name__}:{exc}")


def _nested_transaction_worker(scratch: str, results) -> None:
    try:
        with scratch_transaction(
            scratch,
            max_bytes=1_000,
            max_entries=20,
            max_depth=4,
        ):
            with scratch_transaction(
                scratch,
                max_bytes=1_000,
                max_entries=20,
                max_depth=4,
            ):
                pass
    except BaseException as exc:
        results.put(f"error:{type(exc).__name__}:{exc}")
    else:
        results.put("ok")


def _single_transaction_worker(scratch: str, results) -> None:
    try:
        with scratch_transaction(
            scratch,
            max_bytes=1_000,
            max_entries=20,
            max_depth=4,
        ):
            pass
    except BaseException as exc:
        results.put(f"error:{type(exc).__name__}:{exc}")
    else:
        results.put("ok")


def _wait_for_pid_exit(pid: int, timeout: float = 5.0) -> bool:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        if not Path(f"/proc/{pid}").exists():
            return True
        time.sleep(0.02)
    return not Path(f"/proc/{pid}").exists()


class _FakePdfInspector:
    calls: list[tuple] = []
    inspect_value = "UNTRUSTED_META"
    extract_value = "UNTRUSTED_TEXT"
    render_escape: str | None = None
    fail_message: str | None = None
    unreturned_scratch_bytes = 0
    render_bytes: bytes | None = None

    def _write_unreturned(self, scratch_dir):
        if self.unreturned_scratch_bytes:
            (Path(scratch_dir) / "unreturned.bin").write_bytes(
                b"x" * self.unreturned_scratch_bytes
            )

    def inspect(self, path, scratch_dir, *, response_budget):
        self.calls.append(("inspect", path, scratch_dir, response_budget))
        self._write_unreturned(scratch_dir)
        if self.fail_message is not None:
            raise InspectionError(self.fail_message)
        return {
            "filename": Path(path).name,
            "metadata": self.inspect_value,
        }

    def render(self, path, scratch_dir, selectors, budget):
        self.calls.append(("render", path, scratch_dir, selectors, budget))
        self._write_unreturned(scratch_dir)
        if self.fail_message is not None:
            raise InspectionError(self.fail_message)
        if self.render_escape is not None:
            return [self.render_escape]
        output = Path(scratch_dir) / "render-UNTRUSTED_META.png"
        if self.render_bytes is None:
            PillowImage.new("RGB", (2, 3), color=(25, 50, 75)).save(output)
        else:
            output.write_bytes(self.render_bytes)
        return [str(output)]

    def extract(self, path, scratch_dir, selectors, *, response_budget):
        self.calls.append(("extract", path, scratch_dir, selectors, response_budget))
        self._write_unreturned(scratch_dir)
        if self.fail_message is not None:
            raise InspectionError(self.fail_message)
        return {
            "filename": Path(path).name,
            "text": self.extract_value,
            "selectors": selectors,
        }


@pytest.fixture
def fake_pdf_inspector(monkeypatch):
    _FakePdfInspector.calls = []
    _FakePdfInspector.inspect_value = "UNTRUSTED_META"
    _FakePdfInspector.extract_value = "UNTRUSTED_TEXT"
    _FakePdfInspector.render_escape = None
    _FakePdfInspector.fail_message = None
    _FakePdfInspector.unreturned_scratch_bytes = 0
    _FakePdfInspector.render_bytes = None
    module = types.ModuleType("skillopt.envs.skilleval.inspectors.pdf_image")
    module.PdfInspector = _FakePdfInspector
    module.ImageInspector = _FakePdfInspector
    monkeypatch.setitem(sys.modules, module.__name__, module)
    return _FakePdfInspector


class TestBudgetsAndSelectors:
    def test_budget_defaults_are_frozen_and_bounded(self) -> None:
        render = RenderBudget()
        response = ResponseBudget()

        assert render.max_pixels == 500_000_000
        assert response.max_bytes <= MAX_RESPONSE_BYTES
        assert response.max_extract_chars > 0
        with pytest.raises(Exception):
            render.max_pixels = 1  # type: ignore[misc]

    def test_response_budget_enforces_documented_minimum(self) -> None:
        assert MIN_RESPONSE_BYTES >= len(
            b'{"status":"error","error":"InspectionError: bounded failure"}'
        )
        assert ResponseBudget(max_bytes=MIN_RESPONSE_BYTES).max_bytes == (
            MIN_RESPONSE_BYTES
        )
        with pytest.raises(InspectionError, match="at least"):
            ResponseBudget(max_bytes=MIN_RESPONSE_BYTES - 1)

    def test_scratch_budget_is_part_of_inspector_contract(self) -> None:
        default = inspector_base.DEFAULT_SCRATCH_BYTES
        maximum = inspector_base.MAX_SCRATCH_BYTES

        assert default == 1024 * 1024 * 1024
        assert default <= maximum
        assert RenderBudget(max_scratch_bytes=default).max_scratch_bytes == default
        response = ResponseBudget(
            max_scratch_bytes=default,
            max_scratch_entries=10,
            max_scratch_depth=3,
        )
        assert response.max_scratch_bytes == default
        assert response.max_scratch_entries == 10
        assert response.max_scratch_depth == 3
        with pytest.raises(InspectionError, match="scratch"):
            RenderBudget(max_scratch_bytes=maximum + 1)
        with pytest.raises(InspectionError, match="entr"):
            ResponseBudget(max_scratch_entries=0)
        with pytest.raises(InspectionError, match="depth"):
            ResponseBudget(max_scratch_depth=0)

    @pytest.mark.parametrize(
        "selectors",
        [
            "page:1",
            [""],
            ["\x00"],
            [1],
            ["x" * 257],
            ["x"] * 257,
        ],
    )
    def test_invalid_selectors_are_rejected(self, selectors) -> None:
        with pytest.raises(InspectionError, match="selector"):
            normalize_selectors(selectors)

    def test_valid_selectors_are_copied(self) -> None:
        selectors = ["page:1", "sheet:Summary"]
        normalized = normalize_selectors(selectors)
        selectors.append("page:2")
        assert normalized == ["page:1", "sheet:Summary"]


class TestSafeRun:
    def test_uses_minimal_environment_and_explicit_cwd(
        self, tmp_path, monkeypatch
    ) -> None:
        monkeypatch.setenv("HTTPS_PROXY", "http://proxy.invalid")
        monkeypatch.setenv("AWS_SECRET_ACCESS_KEY", "secret")
        home = tmp_path / "home"
        cwd = tmp_path / "cwd"
        home.mkdir()
        cwd.mkdir()
        script = (
            "import json, os; "
            "print(json.dumps({'keys': sorted(os.environ), "
            "'cwd': os.getcwd(), 'home': os.environ['HOME']}))"
        )

        proc = safe_run(
            [sys.executable, "-c", script],
            timeout=5,
            cwd=str(cwd),
            home=str(home),
        )
        payload = json.loads(proc.stdout)

        assert payload == {
            "keys": ["HOME", "LANG", "PATH"],
            "cwd": str(cwd),
            "home": str(home),
        }

    def test_never_invokes_a_shell(self, tmp_path) -> None:
        cwd = tmp_path / "cwd"
        home = tmp_path / "home"
        cwd.mkdir()
        home.mkdir()
        marker = tmp_path / "shell-ran"
        argument = f"; touch {marker}"

        proc = safe_run(
            [sys.executable, "-c", "import sys; print(sys.argv[1])", argument],
            timeout=5,
            cwd=str(cwd),
            home=str(home),
        )

        assert argument in proc.stdout
        assert not marker.exists()

    @pytest.mark.parametrize("command", [[], [""], ["bad\x00arg"], [1]])
    def test_rejects_invalid_command_arguments(self, tmp_path, command) -> None:
        with pytest.raises(InspectionError, match="invalid inspector command"):
            safe_run(
                command,
                timeout=5,
                cwd=str(tmp_path),
                home=str(tmp_path),
            )

    def test_timeout_is_wrapped_and_bounded(self, tmp_path) -> None:
        with pytest.raises(InspectionError, match="timed out") as excinfo:
            safe_run(
                [sys.executable, "-c", "import time; time.sleep(2)"],
                timeout=0.05,
                cwd=str(tmp_path),
                home=str(tmp_path),
            )
        assert len(str(excinfo.value)) < 1_000

    def test_oserror_is_wrapped(self, tmp_path) -> None:
        with pytest.raises(InspectionError, match="could not start"):
            safe_run(
                [str(tmp_path / "missing-command")],
                timeout=1,
                cwd=str(tmp_path),
                home=str(tmp_path),
            )

    def test_nonzero_diagnostic_is_bounded(self, tmp_path) -> None:
        script = "import sys; sys.stderr.write('S' * 200000); raise SystemExit(7)"
        with pytest.raises(InspectionError, match="exited 7") as excinfo:
            safe_run(
                [sys.executable, "-c", script],
                timeout=5,
                cwd=str(tmp_path),
                home=str(tmp_path),
            )
        assert len(str(excinfo.value)) <= MAX_COMMAND_OUTPUT_CHARS + 200

    def test_success_output_is_bounded(self, tmp_path) -> None:
        script = (
            "import sys; "
            "sys.stdout.write('O' * 200000); "
            "sys.stderr.write('E' * 200000)"
        )
        proc = safe_run(
            [sys.executable, "-c", script],
            timeout=5,
            cwd=str(tmp_path),
            home=str(tmp_path),
        )

        assert len(proc.stdout) <= MAX_COMMAND_OUTPUT_CHARS + 100
        assert len(proc.stderr) <= MAX_COMMAND_OUTPUT_CHARS + 100
        assert "truncated" in proc.stdout
        assert "truncated" in proc.stderr

    def test_supervisor_spawn_does_not_use_preexec_fn(
        self, tmp_path, monkeypatch
    ) -> None:
        real_popen = process_mod.subprocess.Popen
        missing = object()
        preexec_values = []

        def recording_popen(*args, **kwargs):
            preexec_values.append(kwargs.get("preexec_fn", missing))
            return real_popen(*args, **kwargs)

        monkeypatch.setattr(process_mod.subprocess, "Popen", recording_popen)

        process = safe_run(
            [sys.executable, "-c", "print('ok')"],
            timeout=5,
            cwd=str(tmp_path),
            home=str(tmp_path),
        )

        assert process.stdout.strip() == "ok"
        assert preexec_values == [missing]

    def test_safe_run_is_stable_across_concurrent_threads(
        self, tmp_path
    ) -> None:
        def run(index: int) -> str:
            process = safe_run(
                [sys.executable, "-c", f"print({index})"],
                timeout=5,
                cwd=str(tmp_path),
                home=str(tmp_path),
            )
            return process.stdout.strip()

        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as pool:
            outputs = list(pool.map(run, range(16)))

        assert outputs == [str(index) for index in range(16)]

    def test_normal_parent_exit_terminates_lingering_process_group(
        self, tmp_path
    ) -> None:
        pid_file = tmp_path / "child.pid"
        script = (
            "import pathlib, subprocess, sys; "
            "child = subprocess.Popen("
            "[sys.executable, '-c', 'import time; time.sleep(60)']); "
            "pathlib.Path(sys.argv[1]).write_text(str(child.pid))"
        )
        child_pid = None
        started = time.monotonic()
        try:
            safe_run(
                [sys.executable, "-c", script, str(pid_file)],
                timeout=5,
                cwd=str(tmp_path),
                home=str(tmp_path),
            )
            child_pid = int(pid_file.read_text(encoding="utf-8"))
            deadline = time.monotonic() + 2
            while time.monotonic() < deadline:
                try:
                    state = Path(f"/proc/{child_pid}/stat").read_text().split()[2]
                except (FileNotFoundError, ProcessLookupError):
                    break
                if state == "Z":
                    break
                time.sleep(0.02)
            else:
                pytest.fail("lingering child process remained alive")
            assert time.monotonic() - started < 3
        finally:
            if child_pid is not None:
                try:
                    os.kill(child_pid, signal.SIGKILL)
                except ProcessLookupError:
                    pass

    def test_setsid_descendant_is_reaped_after_parent_exits(
        self, tmp_path
    ) -> None:
        pid_file = tmp_path / "setsid-child.pid"
        child_code = (
            "import os, pathlib, sys, time; "
            "os.setsid(); "
            "pathlib.Path(sys.argv[1]).write_text(str(os.getpid())); "
            "time.sleep(60)"
        )
        parent_code = (
            "import pathlib, subprocess, sys, time; "
            "subprocess.Popen("
            "[sys.executable, '-c', sys.argv[1], sys.argv[2]], "
            "stdin=subprocess.DEVNULL, stdout=subprocess.DEVNULL, "
            "stderr=subprocess.DEVNULL); "
            "pid_file = pathlib.Path(sys.argv[2]); "
            "[(time.sleep(0.01)) for _ in range(200) "
            "if not pid_file.exists()]"
        )
        child_pid = None
        try:
            safe_run(
                [
                    sys.executable,
                    "-c",
                    parent_code,
                    child_code,
                    str(pid_file),
                ],
                timeout=5,
                cwd=str(tmp_path),
                home=str(tmp_path),
            )
            child_pid = int(pid_file.read_text(encoding="utf-8"))
            deadline = time.monotonic() + 2
            while time.monotonic() < deadline and Path(
                f"/proc/{child_pid}"
            ).exists():
                time.sleep(0.02)
            assert not Path(f"/proc/{child_pid}").exists()
        finally:
            if child_pid is not None:
                try:
                    os.kill(child_pid, signal.SIGKILL)
                except ProcessLookupError:
                    pass

    @pytest.mark.parametrize(
        "termination_signal",
        [signal.SIGTERM, signal.SIGINT],
    )
    def test_caller_termination_removes_supervisor_and_payload(
        self, tmp_path, termination_signal
    ) -> None:
        payload_pid_file = tmp_path / "payload.pid"
        payload_code = (
            "import os, pathlib, sys, time; "
            "pathlib.Path(sys.argv[1]).write_text(str(os.getpid())); "
            "time.sleep(60)"
        )
        caller_code = (
            "import sys; "
            "from skillopt.envs.skilleval.inspectors.base import safe_run; "
            "safe_run([sys.executable, '-c', sys.argv[1], sys.argv[2]], "
            "timeout=120, cwd=sys.argv[3], home=sys.argv[3])"
        )
        caller = subprocess.Popen(
            [
                sys.executable,
                "-c",
                caller_code,
                payload_code,
                str(payload_pid_file),
                str(tmp_path),
            ],
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True,
        )
        supervisor_pid = None
        payload_pid = None
        try:
            deadline = time.monotonic() + 5
            while time.monotonic() < deadline:
                children_path = Path(
                    f"/proc/{caller.pid}/task/{caller.pid}/children"
                )
                if children_path.exists():
                    children = children_path.read_text(
                        encoding="ascii"
                    ).split()
                    if children:
                        supervisor_pid = int(children[0])
                if payload_pid_file.exists():
                    payload_pid = int(
                        payload_pid_file.read_text(encoding="ascii")
                    )
                if supervisor_pid is not None and payload_pid is not None:
                    break
                time.sleep(0.02)
            assert supervisor_pid is not None
            assert payload_pid is not None

            os.kill(caller.pid, termination_signal)
            caller.wait(timeout=5)

            assert _wait_for_pid_exit(supervisor_pid)
            assert _wait_for_pid_exit(payload_pid)
        finally:
            if caller.poll() is None:
                caller.kill()
                caller.wait()
            for pid in (payload_pid, supervisor_pid):
                if pid is not None:
                    try:
                        os.kill(pid, signal.SIGKILL)
                    except ProcessLookupError:
                        pass


class TestSecurePaths:
    def test_logical_path_boundaries_are_accepted(self) -> None:
        max_total = "/".join(["a" * 240] * 17)
        assert len(max_total) == MAX_LOGICAL_PATH_CHARS
        assert len(max_total.encode("utf-8")) == MAX_LOGICAL_PATH_BYTES
        assert validate_logical_path(max_total) == tuple(
            ["a" * 240] * 17
        )
        assert len(validate_logical_path("a/" * 63 + "a")) == (
            MAX_LOGICAL_COMPONENTS
        )
        assert validate_logical_path("a" * MAX_LOGICAL_COMPONENT_BYTES) == (
            "a" * MAX_LOGICAL_COMPONENT_BYTES,
        )

    @pytest.mark.parametrize(
        ("logical_path", "message"),
        [
            ("/".join(["a" * 240] * 16 + ["a" * 241]), "character"),
            ("/".join(["界" * 85] * 17), "UTF-8"),
            ("/".join(["a"] * 65), "component count"),
            ("界" * 86, "component.*UTF-8"),
            ("\ud800", "valid UTF-8"),
        ],
    )
    def test_logical_path_resource_limits_are_rejected(
        self, logical_path, message
    ) -> None:
        with pytest.raises(InspectionError, match=message):
            validate_logical_path(logical_path)

    @pytest.mark.parametrize(
        "logical_path",
        [
            "../outside.pdf",
            "nested/../outside.pdf",
            "/absolute.pdf",
            r"nested\report.pdf",
            "nested//report.pdf",
            "./report.pdf",
            "report.pdf\x00suffix",
            "",
        ],
    )
    def test_rejects_unsafe_logical_paths(self, tmp_path, logical_path) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")

        with pytest.raises(InspectionError, match="logical|relative"):
            inspect_artifact(
                logical_path,
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_rejects_overlapping_roots(self, tmp_path) -> None:
        evidence = tmp_path / "evidence"
        scratch = evidence / "scratch"
        scratch.mkdir(parents=True)

        with pytest.raises(InspectionError, match="overlap"):
            validate_roots(str(evidence), str(scratch))

    def test_rejects_symlinked_root_component(self, tmp_path) -> None:
        real = tmp_path / "real"
        real.mkdir()
        evidence = tmp_path / "evidence"
        evidence.symlink_to(real, target_is_directory=True)
        scratch = tmp_path / "scratch"
        scratch.mkdir()

        with pytest.raises(InspectionError, match="symlink"):
            validate_roots(str(evidence), str(scratch))

    def test_rejects_symlinked_artifact_component(self, tmp_path) -> None:
        evidence, scratch = _roots(tmp_path)
        outside = tmp_path / "outside"
        outside.mkdir()
        _write_pdf(outside / "report.pdf")
        (evidence / "linked").symlink_to(outside, target_is_directory=True)

        with pytest.raises(InspectionError, match="symlink"):
            with open_evidence_file(str(evidence), "linked/report.pdf"):
                pass

    @pytest.mark.skipif(not hasattr(os, "mkfifo"), reason="FIFO unsupported")
    def test_rejects_nonregular_artifact(self, tmp_path) -> None:
        evidence, _scratch = _roots(tmp_path)
        os.mkfifo(evidence / "pipe.pdf")

        with pytest.raises(InspectionError, match="regular"):
            with open_evidence_file(str(evidence), "pipe.pdf"):
                pass

    def test_unstable_path_resolvers_are_not_public(self) -> None:
        assert not hasattr(inspector_base, "resolve_evidence_path")
        assert not hasattr(inspector_base, "resolve_scratch_path")

    def test_rejects_render_output_escape(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        escaped = tmp_path / "escaped.png"
        PillowImage.new("RGB", (1, 1)).save(escaped)
        fake_pdf_inspector.render_escape = str(escaped)

        with pytest.raises(InspectionError, match="scratch"):
            render_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_rejects_symlinked_render_output(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        outside = tmp_path / "outside.png"
        PillowImage.new("RGB", (1, 1)).save(outside)
        linked = scratch / "linked.png"
        linked.symlink_to(outside)
        fake_pdf_inspector.render_escape = str(linked)

        with pytest.raises(InspectionError, match="symlink"):
            render_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_png_read_uses_open_directory_descriptors_during_replacement(
        self, tmp_path, monkeypatch
    ) -> None:
        _evidence, scratch = _roots(tmp_path)
        nested = scratch / "nested"
        nested.mkdir()
        original = nested / "render.png"
        PillowImage.new("RGB", (2, 3), color=(1, 2, 3)).save(original)
        outside = tmp_path / "outside"
        outside.mkdir()
        PillowImage.new("RGB", (9, 9), color=(4, 5, 6)).save(
            outside / "render.png"
        )
        held = scratch / "held"
        real_open = os.open
        replaced = False

        def replacing_open(path, flags, *args, **kwargs):
            nonlocal replaced
            dir_fd = kwargs.get("dir_fd")
            if not replaced and os.fspath(path) == "nested" and dir_fd is not None:
                descriptor = real_open(path, flags, *args, **kwargs)
                nested.rename(held)
                nested.symlink_to(outside, target_is_directory=True)
                replaced = True
                return descriptor
            return real_open(path, flags, *args, **kwargs)

        monkeypatch.setattr(os, "open", replacing_open)
        _data, details = artifact_mcp._read_png(
            str(original),
            str(scratch),
            remaining_pixels=100,
            remaining_bytes=10_000,
        )

        assert replaced is True
        assert details["width"] == 2
        assert details["height"] == 3

    @pytest.mark.parametrize("raw_suffix", ["nested/../render.png", "/render.png"])
    def test_rejects_non_normalized_absolute_render_output(
        self, tmp_path, fake_pdf_inspector, raw_suffix
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        output = scratch / "render.png"
        PillowImage.new("RGB", (1, 1)).save(output)
        if raw_suffix.startswith("/"):
            returned = f"{scratch}/{raw_suffix}"
        else:
            returned = f"{scratch}/{raw_suffix}"
        fake_pdf_inspector.render_escape = returned

        with pytest.raises(InspectionError, match="normalized|component"):
            render_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )


class TestRegistry:
    def test_unknown_format_is_rejected(self, tmp_path) -> None:
        evidence, scratch = _roots(tmp_path)
        (evidence / "data.bin").write_bytes(b"\x00\x01")

        with pytest.raises(InspectionError, match="unsupported"):
            inspect_artifact(
                "data.bin",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_conflicting_content_and_suffix_are_rejected(self, tmp_path) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "workbook.xlsx")

        with pytest.raises(InspectionError, match="conflict"):
            inspect_artifact(
                "workbook.xlsx",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_concrete_modules_are_loaded_only_when_used(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        sys.modules.pop("skillopt.envs.skilleval.inspectors.spreadsheet", None)
        sys.modules.pop("skillopt.envs.skilleval.inspectors.office", None)

        inventory_artifacts(str(evidence), str(scratch))
        assert "skillopt.envs.skilleval.inspectors.spreadsheet" not in sys.modules
        assert "skillopt.envs.skilleval.inspectors.office" not in sys.modules

        result = inspect_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )
        assert result["metadata"] == "UNTRUSTED_META"
        assert fake_pdf_inspector.calls[0][0] == "inspect"

    def test_missing_concrete_module_fails_only_for_its_kind(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        real_import = inspectors_mod.importlib.import_module

        def missing_pdf(name):
            if name.endswith(".pdf_image"):
                raise ModuleNotFoundError("optional parser missing")
            return real_import(name)

        monkeypatch.setattr(inspectors_mod.importlib, "import_module", missing_pdf)
        assert inventory_artifacts(str(evidence), str(scratch))[0]["kind"] == "pdf"
        with pytest.raises(InspectionError, match="unavailable"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_inventory_is_deterministic_and_complete(self, tmp_path) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "z-report.pdf", b"z")
        _write_pdf(evidence / "nested" / "a-report.pdf", b"a")
        (evidence / "notes.unknown").write_text("notes", encoding="utf-8")

        first = inventory_artifacts(str(evidence), str(scratch))
        second = inventory_artifacts(str(evidence), str(scratch))

        assert first == second
        assert [row["path"] for row in first] == [
            "nested/a-report.pdf",
            "notes.unknown",
            "z-report.pdf",
        ]
        assert first[0]["sha256"] == hashlib.sha256(
            (evidence / "nested" / "a-report.pdf").read_bytes()
        ).hexdigest()
        for row in first:
            assert set(row) == {
                "path",
                "size",
                "sha256",
                "mime",
                "kind",
                "unit_summary",
            }
            assert row["unit_summary"] == {
                "status": "not_inspected",
                "units": [],
            }

    def test_inventory_rejects_paths_deeper_than_tool_path_contract(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        too_deep = evidence.joinpath(*(["a"] * (MAX_LOGICAL_COMPONENTS + 1)))
        too_deep.mkdir(parents=True)
        _write_pdf(too_deep / "report.pdf")

        with pytest.raises(InspectionError, match="component count"):
            inventory_artifacts(str(evidence), str(scratch))

    def test_inspection_uses_stable_copy_when_evidence_directory_is_replaced(
        self, tmp_path, fake_pdf_inspector, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        nested = evidence / "nested"
        _write_pdf(nested / "report.pdf", b"ORIGINAL")
        outside = tmp_path / "outside"
        _write_pdf(outside / "report.pdf", b"ESCAPED")
        held = evidence / "held"
        real_detect = inspectors_mod.detect_artifact_kind
        replaced = False

        def replacing_detect(path):
            nonlocal replaced
            if not replaced:
                nested.rename(held)
                nested.symlink_to(outside, target_is_directory=True)
                replaced = True
            return real_detect(path)

        def reading_inspect(self, path, scratch_dir, *, response_budget):
            return {
                "payload": Path(path).read_bytes().decode(
                    "ascii",
                    errors="replace",
                )
            }

        monkeypatch.setattr(
            inspectors_mod,
            "detect_artifact_kind",
            replacing_detect,
        )
        monkeypatch.setattr(_FakePdfInspector, "inspect", reading_inspect)

        result = inspect_artifact(
            "nested/report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert replaced is True
        assert "ORIGINAL" in result["payload"]
        assert "ESCAPED" not in result["payload"]

    def test_staged_file_replacement_cannot_change_inspector_input(
        self, tmp_path, fake_pdf_inspector, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf", b"ORIGINAL")
        outside = tmp_path / "outside.pdf"
        _write_pdf(outside, b"ESCAPED")
        real_detect = inspectors_mod.detect_artifact_kind
        observed = {}

        def replace_after_detection(path):
            kind = real_detect(path)
            if path.startswith("/proc/self/fd/"):
                proc_dir, basename = os.path.split(path)
                staged = Path(os.readlink(proc_dir)) / basename
            else:
                staged = Path(path)
            staged.unlink()
            staged.symlink_to(outside)
            return kind

        def reading_inspect(self, path, scratch_dir, *, response_budget):
            payload = Path(path).read_bytes().decode(
                "ascii",
                errors="replace",
            )
            observed["payload"] = payload
            return {"payload": payload}

        monkeypatch.setattr(
            inspectors_mod,
            "detect_artifact_kind",
            replace_after_detection,
        )
        monkeypatch.setattr(_FakePdfInspector, "inspect", reading_inspect)

        with pytest.raises(InspectionError, match="scratch"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

        assert "ORIGINAL" in observed["payload"]
        assert "ESCAPED" not in observed["payload"]

    def test_legacy_detection_fallback_never_reopens_staged_name(
        self, tmp_path, fake_pdf_inspector, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        ole = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
        (evidence / "report.doc").write_bytes(ole + b"ORIGINAL")
        outside = tmp_path / "outside.doc"
        outside.write_bytes(ole + b"ESCAPED")
        office = types.ModuleType(
            "skillopt.envs.skilleval.inspectors.office"
        )
        office.OfficeInspector = _FakePdfInspector
        monkeypatch.setitem(sys.modules, office.__name__, office)
        real_open = os.open
        reopened = False
        observed = {}

        def tracking_open(path, flags, *args, **kwargs):
            nonlocal reopened
            if (
                isinstance(path, str)
                and path.startswith("/proc/self/fd/")
                and path.endswith("/report.doc")
            ):
                reopened = True
            return real_open(path, flags, *args, **kwargs)

        def replacing_detect(path, mime=None):
            if mime is not None:
                return "doc"
            proc_dir, basename = os.path.split(path)
            staged = Path(os.readlink(proc_dir)) / basename
            staged.unlink()
            staged.symlink_to(outside)
            return None

        def reading_inspect(self, path, scratch_dir, *, response_budget):
            observed["payload"] = Path(path).read_bytes()
            return {"ok": True}

        monkeypatch.setattr(inspectors_mod.os, "open", tracking_open)
        monkeypatch.setattr(
            inspectors_mod,
            "detect_artifact_kind",
            replacing_detect,
        )
        monkeypatch.setattr(_FakePdfInspector, "inspect", reading_inspect)

        with pytest.raises(InspectionError, match="scratch"):
            inspect_artifact(
                "report.doc",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

        assert reopened is False
        assert b"ORIGINAL" in observed["payload"]
        assert b"ESCAPED" not in observed["payload"]

    def test_inventory_response_limit_fails_closed(self, tmp_path) -> None:
        evidence, scratch = _roots(tmp_path)
        for index in range(10):
            _write_pdf(evidence / f"report-{index}.pdf")

        with pytest.raises(InspectionError, match="response"):
            inventory_artifacts(
                str(evidence),
                str(scratch),
                max_response_bytes=100,
            )

    def test_registry_validates_json_and_response_bounds(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        fake_pdf_inspector.inspect_value = object()

        with pytest.raises(InspectionError, match="JSON-compatible"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

        fake_pdf_inspector.inspect_value = "x" * 2_000
        with pytest.raises(InspectionError, match="response"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                max_response_bytes=100,
            )

    def test_render_and_extract_validate_selectors_and_budgets(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")

        paths = render_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["page:1"],
            max_pixels=1_000,
        )
        extracted = extract_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["page:1"],
            max_extract_chars=1_000,
        )

        assert Path(paths[0]).is_relative_to(scratch)
        assert extracted["selectors"] == ["page:1"]
        assert fake_pdf_inspector.calls[-2][3] == ["page:1"]
        assert fake_pdf_inspector.calls[-2][4] == RenderBudget(1_000)
        assert fake_pdf_inspector.calls[-1][3] == ["page:1"]

    @pytest.mark.parametrize(
        ("function", "kwargs"),
        [
            (render_artifact, {"max_pixels": 0}),
            (render_artifact, {"max_pixels": MAX_RENDER_PIXELS + 1}),
            (inspect_artifact, {"max_response_bytes": 0}),
            (inspect_artifact, {"max_response_bytes": MAX_RESPONSE_BYTES + 1}),
            (extract_artifact, {"max_extract_chars": 0}),
        ],
    )
    def test_registry_rejects_invalid_budgets(
        self, tmp_path, fake_pdf_inspector, function, kwargs
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")

        with pytest.raises(InspectionError, match="budget|positive|maximum"):
            function(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                **kwargs,
            )

    def test_scratch_budget_counts_preexisting_and_unreturned_files(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        (scratch / "preexisting.bin").write_bytes(b"p" * 65)

        with pytest.raises(InspectionError, match="scratch"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                max_scratch_bytes=64,
            )
        assert fake_pdf_inspector.calls == []

        (scratch / "preexisting.bin").unlink()
        fake_pdf_inspector.unreturned_scratch_bytes = 128
        with pytest.raises(InspectionError, match="scratch"):
            render_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                max_scratch_bytes=100,
        )
        assert not (scratch / "unreturned.bin").exists()

    def test_scratch_budget_counts_staged_input_and_outputs_at_peak(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf", b"e" * 60)
        fake_pdf_inspector.unreturned_scratch_bytes = 50

        with pytest.raises(InspectionError, match="scratch"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                max_scratch_bytes=100,
            )

        assert not (scratch / "unreturned.bin").exists()

    def test_scratch_transaction_limits_external_write_and_rolls_back(
        self, tmp_path, fake_pdf_inspector, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        escaped_marker = tmp_path / "writer-completed"
        script = (
            "from pathlib import Path; "
            "Path('oversized.bin').write_bytes(b'x' * (5 * 1024 * 1024)); "
            f"Path({str(escaped_marker)!r}).write_text('completed')"
        )

        def process_inspect(self, path, scratch_dir, *, response_budget):
            safe_run(
                [sys.executable, "-c", script],
                timeout=5,
                cwd=scratch_dir,
                home=scratch_dir,
            )
            return {"ok": True}

        monkeypatch.setattr(_FakePdfInspector, "inspect", process_inspect)

        with pytest.raises(InspectionError, match="scratch|exited"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                max_scratch_bytes=1024,
            )

        assert not escaped_marker.exists()
        assert list(scratch.iterdir()) == []

    def test_scratch_root_escape_is_budgeted_and_rolled_back(
        self, tmp_path, fake_pdf_inspector, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf", b"")
        script = (
            "from pathlib import Path; "
            "Path('../escape-a.bin').write_bytes(b'a' * 40); "
            "Path('../escape-b.bin').write_bytes(b'b' * 40)"
        )

        def process_inspect(self, path, scratch_dir, *, response_budget):
            safe_run(
                [sys.executable, "-c", script],
                timeout=5,
                cwd=scratch_dir,
                home=scratch_dir,
            )
            return {"ok": True}

        monkeypatch.setattr(_FakePdfInspector, "inspect", process_inspect)

        with pytest.raises(InspectionError, match="scratch"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                max_scratch_bytes=64,
            )

        assert list(scratch.iterdir()) == []

    def test_zero_mode_subtree_is_removed_after_safe_run_failure(
        self, tmp_path, fake_pdf_inspector, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        temporary_root = Path(tempfile.gettempdir())
        before = set(temporary_root.glob(".skillopt-artifact-*"))
        script = (
            "from pathlib import Path; import os, time; "
            "nested = Path('sealed/inner'); "
            "nested.mkdir(parents=True); "
            "(nested / 'payload.bin').write_bytes(b'x'); "
            "os.chmod('sealed', 0); "
            "time.sleep(2)"
        )

        def process_inspect(self, path, scratch_dir, *, response_budget):
            safe_run(
                [sys.executable, "-c", script],
                timeout=5,
                cwd=scratch_dir,
                home=scratch_dir,
            )
            return {"ok": True}

        monkeypatch.setattr(_FakePdfInspector, "inspect", process_inspect)

        with pytest.raises(InspectionError, match="scratch"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

        assert set(temporary_root.glob(".skillopt-artifact-*")) == before
        assert list(scratch.iterdir()) == []

    @pytest.mark.parametrize(
        "operation",
        ["delete", "overwrite", "truncate", "rename"],
    )
    def test_transaction_cannot_modify_preexisting_scratch_content(
        self,
        tmp_path,
        fake_pdf_inspector,
        monkeypatch,
        operation,
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        existing = scratch / "existing.bin"
        moved = scratch / "moved.bin"
        existing.write_bytes(b"ORIGINAL")

        def destructive_inspect(
            self, path, scratch_dir, *, response_budget
        ):
            target = Path(scratch_dir) / ".." / "existing.bin"
            destination = Path(scratch_dir) / ".." / "moved.bin"
            target.write_bytes(b"OUTER")
            if operation == "delete":
                target.unlink()
                happened = not target.exists()
            elif operation == "overwrite":
                target.write_bytes(b"CHANGED")
                happened = target.read_bytes() == b"CHANGED"
            elif operation == "truncate":
                target.open("wb").close()
                happened = target.stat().st_size == 0
            else:
                target.rename(destination)
                happened = (
                    not target.exists()
                    and destination.read_bytes() == b"OUTER"
                )
            return {"ok": True, "branch_happened": happened}

        monkeypatch.setattr(
            _FakePdfInspector,
            "inspect",
            destructive_inspect,
        )

        result = inspect_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert result == {"ok": True, "branch_happened": True}
        assert existing.read_bytes() == b"ORIGINAL"
        assert not moved.exists()

    def test_same_thread_nested_transaction_fails_without_hanging(
        self, tmp_path
    ) -> None:
        _evidence, scratch = _roots(tmp_path)
        context = multiprocessing.get_context("fork")
        results = context.Queue()
        process = context.Process(
            target=_nested_transaction_worker,
            args=(str(scratch), results),
        )
        process.start()
        try:
            process.join(timeout=1)
            assert not process.is_alive()
            outcome = results.get(timeout=1)
        finally:
            if process.is_alive():
                process.kill()
                process.join()

        assert outcome.startswith("error:InspectionError:")
        assert "nested" in outcome

    def test_fork_after_lock_rebuilds_child_lock_state(
        self, tmp_path
    ) -> None:
        _evidence, scratch = _roots(tmp_path)
        context = multiprocessing.get_context("fork")
        results = context.Queue()
        process = context.Process(
            target=_single_transaction_worker,
            args=(str(scratch), results),
        )
        try:
            with scratch_transaction(
                str(scratch),
                max_bytes=1_000,
                max_entries=20,
                max_depth=4,
            ):
                process.start()
                time.sleep(0.1)
                assert process.is_alive()

            process.join(timeout=3)
            assert not process.is_alive()
            assert results.get(timeout=1) == "ok"
        finally:
            if process.is_alive():
                process.kill()
                process.join()

    def test_interrupted_lock_acquisition_does_not_leak_root_fd(
        self, tmp_path, monkeypatch
    ) -> None:
        _evidence, scratch = _roots(tmp_path)

        class InterruptingLock:
            def acquire(self):
                raise KeyboardInterrupt

            def release(self):
                pytest.fail("unacquired lock must not be released")

        monkeypatch.setattr(
            scratch_root_mod,
            "_thread_lock",
            lambda _path: InterruptingLock(),
        )
        before = set(os.listdir("/proc/self/fd"))

        with pytest.raises(KeyboardInterrupt):
            with scratch_transaction(
                str(scratch),
                max_bytes=1_000,
                max_entries=20,
                max_depth=4,
            ):
                pass

        assert set(os.listdir("/proc/self/fd")) == before

    def test_stable_evidence_fd_is_inherited_by_real_child(
        self, tmp_path, fake_pdf_inspector, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf", b"ORIGINAL")

        def process_inspect(self, path, scratch_dir, *, response_budget):
            process = safe_run(
                [
                    sys.executable,
                    "-c",
                    (
                        "from pathlib import Path; import sys; "
                        "sys.stdout.buffer.write(Path(sys.argv[1]).read_bytes())"
                    ),
                    path,
                ],
                timeout=5,
                cwd=scratch_dir,
                home=scratch_dir,
            )
            return {"payload": process.stdout}

        monkeypatch.setattr(_FakePdfInspector, "inspect", process_inspect)

        result = inspect_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert "ORIGINAL" in result["payload"]
        assert list(scratch.iterdir()) == []

    def test_scratch_root_budget_serializes_process_transactions(
        self, tmp_path
    ) -> None:
        _evidence, scratch = _roots(tmp_path)
        context = multiprocessing.get_context("fork")
        barrier = context.Barrier(2)
        entered = context.Value("i", 0)
        results = context.Queue()
        processes = [
            context.Process(
                target=_scratch_commit_worker,
                args=(str(scratch), barrier, entered, results),
            )
            for _ in range(2)
        ]
        for process in processes:
            process.start()
        try:
            for process in processes:
                process.join(timeout=10)
                assert not process.is_alive()
                assert process.exitcode == 0
            outcomes = sorted(
                [results.get(timeout=2), results.get(timeout=2)]
            )
        finally:
            for process in processes:
                if process.is_alive():
                    process.kill()
                    process.join()

        assert outcomes[0].startswith("error:InspectionError:scratch")
        assert outcomes[1] == "ok"
        files = [path for path in scratch.rglob("*") if path.is_file()]
        assert len(files) == 1
        assert files[0].stat().st_size == 80

    def test_cross_filesystem_commit_reserves_copy_peak(
        self, tmp_path, monkeypatch
    ) -> None:
        _evidence, scratch = _roots(tmp_path)
        copy_called = False
        real_copy = scratch_mod._copy_stable_file

        def cross_device(*args, **kwargs):
            raise OSError(errno.EXDEV, "cross-device link")

        def recording_copy(*args, **kwargs):
            nonlocal copy_called
            copy_called = True
            return real_copy(*args, **kwargs)

        monkeypatch.setattr(scratch_mod.os, "rename", cross_device)
        monkeypatch.setattr(
            scratch_mod,
            "_copy_stable_file",
            recording_copy,
        )

        with pytest.raises(InspectionError, match="scratch byte"):
            with scratch_transaction(
                str(scratch),
                max_bytes=100,
                max_entries=20,
                max_depth=4,
            ) as transaction:
                output = Path(transaction.proc_path) / "render.bin"
                output.write_bytes(b"x" * 80)
                transaction.commit_outputs([str(output)])

        assert copy_called is False
        assert list(scratch.iterdir()) == []

    @pytest.mark.parametrize(
        ("limit_name", "limit_value", "script"),
        [
            (
                "max_scratch_entries",
                3,
                (
                    "from pathlib import Path; import time; "
                    "[(Path(f'entry-{i}').touch(), time.sleep(0.005)) "
                    "for i in range(100)]; "
                ),
            ),
            (
                "max_scratch_depth",
                2,
                (
                    "from pathlib import Path; import time; "
                    "Path('a/b/c').mkdir(parents=True); time.sleep(1); "
                ),
            ),
        ],
    )
    def test_external_watchdog_limits_entries_and_depth(
        self,
        tmp_path,
        fake_pdf_inspector,
        monkeypatch,
        limit_name,
        limit_value,
        script,
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        completed = tmp_path / f"{limit_name}-completed"
        command = script + f"Path({str(completed)!r}).write_text('completed')"

        def process_inspect(self, path, scratch_dir, *, response_budget):
            safe_run(
                [sys.executable, "-c", command],
                timeout=5,
                cwd=scratch_dir,
                home=scratch_dir,
            )
            return {"ok": True}

        monkeypatch.setattr(_FakePdfInspector, "inspect", process_inspect)

        with pytest.raises(InspectionError, match="scratch"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                **{limit_name: limit_value},
            )

        assert not completed.exists()
        assert list(scratch.iterdir()) == []

    @pytest.mark.parametrize(
        ("limit_name", "limit_value", "writer"),
        [
            (
                "max_scratch_entries",
                3,
                lambda root: [
                    (root / f"entry-{index}").write_bytes(b"")
                    for index in range(4)
                ],
            ),
            (
                "max_scratch_depth",
                2,
                lambda root: (root / "a" / "b" / "c").mkdir(parents=True),
            ),
        ],
    )
    def test_scratch_transaction_limits_entries_and_depth(
        self,
        tmp_path,
        fake_pdf_inspector,
        monkeypatch,
        limit_name,
        limit_value,
        writer,
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")

        def writing_inspect(self, path, scratch_dir, *, response_budget):
            writer(Path(scratch_dir))
            return {"ok": True}

        monkeypatch.setattr(_FakePdfInspector, "inspect", writing_inspect)

        with pytest.raises(InspectionError, match="scratch"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                **{limit_name: limit_value},
            )

        assert list(scratch.iterdir()) == []

    def test_success_commits_only_returned_render_outputs(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        fake_pdf_inspector.unreturned_scratch_bytes = 10

        outputs = render_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        files = sorted(path for path in scratch.rglob("*") if path.is_file())
        assert files == [Path(outputs[0])]
        assert files[0].exists()

    def test_render_commit_respects_root_depth_budget(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")

        with pytest.raises(InspectionError, match="scratch depth"):
            render_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                max_scratch_depth=1,
            )

        assert list(scratch.iterdir()) == []

    def test_failed_response_budget_rolls_back_committed_outputs(
        self, tmp_path, fake_pdf_inspector, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")

        def render_many(self, path, scratch_dir, selectors, budget):
            outputs = []
            for index in range(16):
                output = Path(scratch_dir) / (
                    f"{index:02d}-" + "x" * 80 + ".png"
                )
                PillowImage.new("RGB", (1, 1)).save(output)
                outputs.append(str(output))
            return outputs

        monkeypatch.setattr(_FakePdfInspector, "render", render_many)

        with pytest.raises(InspectionError, match="response"):
            render_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                max_response_bytes=MIN_RESPONSE_BYTES,
            )

        assert list(scratch.iterdir()) == []

    def test_failed_transaction_entry_does_not_leak_root_descriptor(
        self, tmp_path
    ) -> None:
        _evidence, scratch = _roots(tmp_path)
        (scratch / "existing.bin").write_bytes(b"x")
        before = set(os.listdir("/proc/self/fd"))

        for _ in range(3):
            with pytest.raises(InspectionError, match="scratch"):
                with scratch_transaction(
                    str(scratch),
                    max_bytes=0,
                    max_entries=10,
                    max_depth=2,
                ):
                    pytest.fail("over-budget transaction entered")

        assert set(os.listdir("/proc/self/fd")) == before


class TestSpreadsheetInspector:
    @staticmethod
    def _save_xlsx(path: Path, *, formula: bool = True) -> None:
        from openpyxl import Workbook

        book = Workbook()
        sheet = book.active
        sheet.title = "Summary"
        sheet["A1"] = "Revenue"
        sheet["B1"] = 10
        sheet["B2"] = 20
        if formula:
            sheet["B3"] = "=SUM(B1:B2)"
        sheet.merge_cells("A5:B5")
        book.save(path)
        book.close()

    @classmethod
    def _add_shared_strings_part(cls, path: Path) -> None:
        shared_strings = (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<sst xmlns="http://schemas.openxmlformats.org/'
            b'spreadsheetml/2006/main" count="1" uniqueCount="1">'
            b"<si><t>Revenue</t></si></sst>"
        )
        with zipfile.ZipFile(path, "a") as archive:
            archive.writestr("xl/sharedStrings.xml", shared_strings)
        override = (
            b'<Override PartName="/xl/sharedStrings.xml" '
            b'ContentType="application/vnd.openxmlformats-'
            b'officedocument.spreadsheetml.sharedStrings+xml"/>'
        )
        relationship = (
            b'<Relationship Type="http://schemas.openxmlformats.org/'
            b'officeDocument/2006/relationships/sharedStrings" '
            b'Target="/xl/sharedStrings.xml" Id="rIdShared"/>'
        )
        cls._rewrite_xlsx_member(
            path,
            "[Content_Types].xml",
            lambda payload: payload.replace(
                b"</Types>",
                override + b"</Types>",
                1,
            ),
        )
        cls._rewrite_xlsx_member(
            path,
            "xl/_rels/workbook.xml.rels",
            lambda payload: payload.replace(
                b"</Relationships>",
                relationship + b"</Relationships>",
                1,
            ),
        )

    @staticmethod
    def _rewrite_xlsx_member(path: Path, member_name: str, transform) -> None:
        rewritten = path.with_name(f"{path.stem}-rewritten.xlsx")
        with zipfile.ZipFile(path) as source:
            members = [
                (
                    info,
                    transform(source.read(info))
                    if info.filename == member_name
                    else source.read(info),
                )
                for info in source.infolist()
            ]
        with zipfile.ZipFile(rewritten, "w") as destination:
            for info, payload in members:
                destination.writestr(info, payload)
        os.replace(rewritten, path)

    @staticmethod
    def _rewrite_xlsx_package(path: Path, transform) -> None:
        rewritten = path.with_name(f"{path.stem}-package-rewritten.xlsx")
        with (
            zipfile.ZipFile(path) as source,
            zipfile.ZipFile(rewritten, "w") as destination,
        ):
            for info in source.infolist():
                name, payload = transform(
                    info.filename,
                    source.read(info),
                )
                destination.writestr(name, payload)
        os.replace(rewritten, path)

    @staticmethod
    def _assert_inspect_rejected_before_openpyxl(
        evidence: Path,
        scratch: Path,
        filename: str,
        monkeypatch,
        match: str,
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        load_calls = []

        def forbidden_load(*args, **kwargs):
            load_calls.append((args, kwargs))
            raise AssertionError("load_workbook must not be called")

        monkeypatch.setattr(
            spreadsheet.openpyxl,
            "load_workbook",
            forbidden_load,
        )
        with pytest.raises(InspectionError, match=match):
            inspect_artifact(
                filename,
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )
        assert load_calls == []

    @staticmethod
    def _mutate_local_header(
        path: Path,
        member_name: str,
        relative_offset: int,
        replacement: bytes,
    ) -> None:
        with zipfile.ZipFile(path) as archive:
            header_offset = archive.getinfo(member_name).header_offset
        with path.open("r+b") as target:
            target.seek(header_offset + relative_offset)
            target.write(replacement)

    @staticmethod
    def _rewrite_xlsx_with_data_descriptors(
        path: Path,
        *,
        zip64: bool,
        signature: bool,
        descriptor_transform=None,
    ) -> None:
        with zipfile.ZipFile(path) as source:
            members = [
                (info, source.read(info))
                for info in source.infolist()
            ]

        output = bytearray()
        central_records = []
        for info, payload in members:
            name = info.filename.encode("utf-8")
            compressor = zlib.compressobj(level=6, wbits=-15)
            compressed = compressor.compress(payload) + compressor.flush()
            crc = zlib.crc32(payload) & 0xFFFFFFFF
            version = 45 if zip64 else 20
            flags = 0x808
            local_offset = len(output)
            if zip64:
                local_compressed = 0xFFFFFFFF
                local_uncompressed = 0xFFFFFFFF
                local_extra = struct.pack("<HHQQ", 0x0001, 16, 0, 0)
                descriptor = struct.pack(
                    "<LQQ",
                    crc,
                    len(compressed),
                    len(payload),
                )
            else:
                local_compressed = 0
                local_uncompressed = 0
                local_extra = b""
                descriptor = struct.pack(
                    "<3L",
                    crc,
                    len(compressed),
                    len(payload),
                )
            if signature:
                descriptor = b"PK\x07\x08" + descriptor
            if descriptor_transform is not None:
                descriptor = descriptor_transform(info.filename, descriptor)

            output.extend(
                struct.pack(
                    "<4s5H3L2H",
                    b"PK\x03\x04",
                    version,
                    flags,
                    zipfile.ZIP_DEFLATED,
                    0,
                    0,
                    0,
                    local_compressed,
                    local_uncompressed,
                    len(name),
                    len(local_extra),
                )
            )
            output.extend(name)
            output.extend(local_extra)
            output.extend(compressed)
            output.extend(descriptor)

            central_records.append(
                struct.pack(
                    "<4s6H3L5H2L",
                    b"PK\x01\x02",
                    (3 << 8) | version,
                    version,
                    flags,
                    zipfile.ZIP_DEFLATED,
                    0,
                    0,
                    crc,
                    len(compressed),
                    len(payload),
                    len(name),
                    0,
                    0,
                    0,
                    0,
                    info.external_attr,
                    local_offset,
                )
                + name
            )

        central_offset = len(output)
        for record in central_records:
            output.extend(record)
        central_size = len(output) - central_offset
        output.extend(
            struct.pack(
                "<4s4H2LH",
                b"PK\x05\x06",
                0,
                0,
                len(members),
                len(members),
                central_size,
                central_offset,
                0,
            )
        )
        path.write_bytes(output)

    @staticmethod
    def _data_descriptor_offset(path: Path, member_name: str) -> int:
        with zipfile.ZipFile(path) as archive:
            info = archive.getinfo(member_name)
            with path.open("rb") as source:
                source.seek(info.header_offset)
                fields = struct.unpack("<4s5H3L2H", source.read(30))
            return (
                info.header_offset
                + 30
                + fields[-2]
                + fields[-1]
                + info.compress_size
            )

    def test_xlsx_inspection_reports_values_formulas_and_layout(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "report.xlsx"
        self._save_xlsx(path)

        result = inspect_artifact(
            "report.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert result["kind"] == "spreadsheet"
        assert result["opens"] is True
        summary = result["sheets"][0]
        assert summary["name"] == "Summary"
        assert summary["used_range"] == "A1:B5"
        assert summary["cells"]["B3"] == {
            "coordinate": "B3",
            "value": None,
            "cached_value": None,
            "formula": "=SUM(B1:B2)",
            "number_format": "General",
            "style_id": 0,
        }
        assert "A5:B5" in summary["merged_ranges"]
        assert isinstance(summary["charts"], list)
        assert isinstance(summary["drawings"], list)

    def test_xlsx_loads_formula_and_cached_views(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        self._save_xlsx(evidence / "report.xlsx")
        from skillopt.envs.skilleval.inspectors import spreadsheet

        calls = []
        real_load = spreadsheet.openpyxl.load_workbook

        def recording_load(*args, **kwargs):
            calls.append(kwargs.get("data_only"))
            return real_load(*args, **kwargs)

        monkeypatch.setattr(spreadsheet.openpyxl, "load_workbook", recording_load)

        inspect_artifact(
            "report.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert calls == [False, True]

    def test_xlsx_check_evaluator_is_deterministic(self, tmp_path) -> None:
        from openpyxl import Workbook
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            evaluate_xlsx_check,
        )

        path = tmp_path / "report.xlsx"
        book = Workbook()
        book.active.title = "Summary"
        book.active["B12"] = "=SUM(B2:B11)"
        book.active["C12"] = 1
        book.save(path)
        book.close()

        formula_check = {
            "id": "formula",
            "type": "xlsx_formula",
            "path": "report.xlsx",
            "required": True,
            "weight": 1.0,
            "spec": {
                "sheet": "Summary",
                "cell": "B12",
                "formula": "SUM(B2:B11)",
            },
        }
        cell_check = {
            "id": "cell",
            "type": "xlsx_cell",
            "path": "report.xlsx",
            "required": True,
            "weight": 1.0,
            "spec": {"sheet": "Summary", "cell": "C12", "value": 1},
        }

        assert evaluate_xlsx_check(str(path), formula_check) == {
            "id": "formula",
            "passed": True,
            "score": 1.0,
            "reason": "formula matches exactly",
            "evidence": [
                {
                    "path": "report.xlsx",
                    "locator": "sheet=Summary,cell=B12",
                    "source": "structure",
                }
            ],
        }
        assert evaluate_xlsx_check(str(path), cell_check)["passed"] is True
        cell_check["spec"]["value"] = True
        assert evaluate_xlsx_check(str(path), cell_check)["passed"] is False

    def test_xlsx_array_formula_is_json_safe_and_checkable(
        self, tmp_path
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.worksheet.formula import ArrayFormula
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            evaluate_xlsx_check,
        )

        evidence, scratch = _roots(tmp_path)
        path = evidence / "array-formula.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "Summary"
        sheet["A1"] = ArrayFormula(
            ref="A1:A2",
            text="=ROW(A1:A2)",
        )
        book.save(path)
        book.close()

        inspected = inspect_artifact(
            "array-formula.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )
        detail = inspected["sheets"][0]["cells"]["A1"]
        assert detail["formula"] == "=ROW(A1:A2)"
        assert detail["formula_metadata"] == {
            "kind": "array",
            "ref": "A1:A2",
        }
        json.dumps(inspected, allow_nan=False)

        check = {
            "id": "array",
            "type": "xlsx_formula",
            "path": "array-formula.xlsx",
            "spec": {
                "sheet": "Summary",
                "cell": "A1",
                "formula": "ROW(A1:A2)",
            },
        }
        assert evaluate_xlsx_check(str(path), check)["passed"] is True

    def test_xlsx_unknown_formula_object_is_controlled(self) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        cached_sheet = types.SimpleNamespace(_cells={})
        values = [
            object(),
            types.SimpleNamespace(
                text="=SUM(A1:A2)",
                ref="A1:A2",
                t="array",
            ),
        ]

        for value in values:
            formula_cell = types.SimpleNamespace(
                row=1,
                column=1,
                coordinate="A1",
                value=value,
                data_type="f",
                number_format="General",
                style_id=0,
            )
            detail = spreadsheet._cell_detail(formula_cell, cached_sheet)

            assert detail["formula"] is None
            assert detail["formula_metadata"] == {"kind": "unsupported"}
            json.dumps(detail, allow_nan=False)

    @pytest.mark.parametrize(
        "value",
        [float("nan"), float("inf"), float("-inf")],
        ids=["nan", "positive-infinity", "negative-infinity"],
    )
    def test_xlsx_cell_equality_rejects_nonfinite_numbers(
        self, value
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        assert spreadsheet._exact_cell_equal(value, value) is False
        assert spreadsheet._exact_cell_equal(1, value) is False
        assert spreadsheet._exact_cell_equal(value, 1) is False

    @pytest.mark.parametrize(
        ("sheet_name", "cell_ref", "reason"),
        [
            ("Missing", "A1", "worksheet"),
            ("Summary", "Z99", "cell"),
        ],
    )
    def test_xlsx_check_missing_sheet_or_cell_fails_deterministically(
        self, tmp_path, sheet_name, cell_ref, reason
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            evaluate_xlsx_check,
        )

        path = tmp_path / "report.xlsx"
        self._save_xlsx(path)
        check = {
            "id": "missing",
            "type": "xlsx_cell",
            "path": "report.xlsx",
            "spec": {
                "sheet": sheet_name,
                "cell": cell_ref,
                "value": "expected",
            },
        }

        result = evaluate_xlsx_check(str(path), check)

        assert result["passed"] is False
        assert result["score"] == 0.0
        assert reason in result["reason"]

    def test_xlsx_forged_dimension_is_ignored_for_cell_pagination(
        self, tmp_path
    ) -> None:
        from openpyxl import Workbook

        evidence, scratch = _roots(tmp_path)
        path = evidence / "forged-dimension.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "Summary"
        for row in range(1, 131):
            sheet.cell(row=row, column=1, value=row)
        book.save(path)
        book.close()

        def forge_dimension(payload: bytes) -> bytes:
            original = b'<dimension ref="A1:A130"/>'
            assert original in payload
            return payload.replace(
                original,
                b'<dimension ref="A1:XFD1048576"/>',
                1,
            )

        self._rewrite_xlsx_member(
            path,
            "xl/worksheets/sheet1.xml",
            forge_dimension,
        )

        inspected = inspect_artifact(
            "forged-dimension.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )
        summary = inspected["sheets"][0]

        assert summary["max_row"] == 130
        assert summary["max_column"] == 1
        assert summary["used_range"] == "A1:A130"
        assert summary["cell_page"] == {
            "page": 1,
            "page_size": 128,
            "total": 130,
            "returned": 128,
            "omitted": 2,
            "omitted_due_to_budget": 0,
        }

        extracted = extract_artifact(
            "forged-dimension.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["sheet:Summary:page:2"],
        )
        assert list(extracted["sheets"][0]["cells"]) == ["A129", "A130"]

    def test_xlsx_real_far_cell_is_paginated_without_grid_walk(
        self, tmp_path
    ) -> None:
        from openpyxl import Workbook

        evidence, scratch = _roots(tmp_path)
        path = evidence / "sparse.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "Summary"
        for row in range(1, 131):
            sheet.cell(row=row, column=1, value=row)
        sheet["XFD1048576"] = "edge"
        book.save(path)
        book.close()

        inspected = inspect_artifact(
            "sparse.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )
        summary = inspected["sheets"][0]

        assert summary["max_row"] == 1_048_576
        assert summary["max_column"] == 16_384
        assert summary["cell_page"] == {
            "page": 1,
            "page_size": 128,
            "total": 131,
            "returned": 128,
            "omitted": 3,
            "omitted_due_to_budget": 0,
        }
        assert "XFD1048576" not in summary["cells"]

        extracted = extract_artifact(
            "sparse.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["sheet:Summary:page:2"],
        )
        page = extracted["sheets"][0]
        assert list(page["cells"]) == ["A129", "A130", "XFD1048576"]
        assert page["cell_page"]["omitted"] == 128

    def test_xlsx_inventory_includes_chart_sheets(self, tmp_path) -> None:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference

        evidence, scratch = _roots(tmp_path)
        path = evidence / "chartsheet.xlsx"
        book = Workbook()
        data = book.active
        data.title = "Data"
        data["A1"] = 1
        chart = BarChart()
        chart.add_data(
            Reference(data, min_col=1, min_row=1, max_row=1)
        )
        chart_sheet = book.create_chartsheet("Chart")
        chart_sheet.add_chart(chart)
        book.save(path)
        book.close()

        inspected = inspect_artifact(
            "chartsheet.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert [
            (sheet["name"], sheet["type"])
            for sheet in inspected["sheets"]
        ] == [("Data", "worksheet"), ("Chart", "chartsheet")]
        assert inspected["sheets"][1]["charts"] == [
            {"type": "BarChart", "anchor": {}}
        ]
        assert inspected["sheets"][1]["drawings"] == [
            {"kind": "chart", "type": "BarChart", "anchor": {}}
        ]

    def test_xlsx_extraction_omits_oversized_cell_with_budget_metadata(
        self, tmp_path
    ) -> None:
        from openpyxl import Workbook

        evidence, scratch = _roots(tmp_path)
        path = evidence / "large-cell.xlsx"
        book = Workbook()
        book.active.title = "Summary"
        book.active["A1"] = "x" * 10_000
        book.save(path)
        book.close()

        extracted = extract_artifact(
            "large-cell.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["sheet:Summary"],
            max_extract_chars=2_000,
            max_response_bytes=2_000,
        )

        page = extracted["sheets"][0]
        assert page["cells"] == {}
        assert page["cell_page"]["returned"] == 0
        assert page["cell_page"]["omitted_due_to_budget"] == 1

    @pytest.mark.parametrize(
        "unsafe_name",
        [
            "../escape.xml",
            "/absolute.xml",
            "xl\\evil.xml",
            "xl//evil.xml",
            "C:/absolute.xml",
        ],
    )
    def test_xlsx_preflight_rejects_unsafe_zip_entries(
        self, tmp_path, unsafe_name
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "unsafe.xlsx"
        self._save_xlsx(path)
        with zipfile.ZipFile(path, "a") as archive:
            archive.writestr(unsafe_name, "unsafe")

        with pytest.raises(InspectionError, match="unsafe OOXML entry"):
            preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        "colliding_name",
        [
            "xl/workbook.xml",
            "XL/workbook.xml",
            "xl/WORKBOOK.XML",
            "xl/cafe\u0301.xml",
        ],
    )
    def test_xlsx_preflight_rejects_duplicate_or_unicode_case_collision(
        self, tmp_path, colliding_name
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "collision.xlsx"
        self._save_xlsx(path)
        if colliding_name == "xl/workbook.xml":
            with pytest.warns(UserWarning, match="Duplicate name"):
                with zipfile.ZipFile(path, "a") as archive:
                    archive.writestr(colliding_name, "second")
        else:
            with zipfile.ZipFile(path, "a") as archive:
                if "cafe" in colliding_name:
                    archive.writestr("xl/caf\u00e9.xml", "first")
                archive.writestr(colliding_name, "second")

        with pytest.raises(InspectionError, match="colliding OOXML entry"):
            preflight_xlsx(str(path))

    def test_xlsx_preflight_rejects_entry_count_size_and_ratio_limits(
        self, tmp_path, monkeypatch
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / "limits.xlsx"
        self._save_xlsx(path)
        with zipfile.ZipFile(path, "a", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("xl/media/repetitive.bin", b"0" * 1_000_000)

        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            declared = sum(info.file_size for info in members)

        monkeypatch.setattr(spreadsheet, "_MAX_OOXML_ENTRIES", len(members) - 1)
        with pytest.raises(InspectionError, match="entry count"):
            spreadsheet.preflight_xlsx(str(path))

        monkeypatch.setattr(spreadsheet, "_MAX_OOXML_ENTRIES", len(members))
        monkeypatch.setattr(
            spreadsheet,
            "_MAX_OOXML_UNCOMPRESSED_BYTES",
            declared - 1,
        )
        with pytest.raises(InspectionError, match="uncompressed"):
            spreadsheet.preflight_xlsx(str(path))

        monkeypatch.setattr(
            spreadsheet,
            "_MAX_OOXML_UNCOMPRESSED_BYTES",
            declared,
        )
        monkeypatch.setattr(spreadsheet, "_MAX_COMPRESSION_RATIO", 2.0)
        with pytest.raises(InspectionError, match="compression ratio"):
            spreadsheet.preflight_xlsx(str(path))

    def test_xlsx_structure_limit_rejects_before_openpyxl_load(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "too-many-cells.xlsx"
        self._save_xlsx(path)
        from skillopt.envs.skilleval.inspectors import spreadsheet

        monkeypatch.setattr(
            spreadsheet,
            "_MAX_WORKSHEET_CELLS",
            1,
            raising=False,
        )
        load_calls = []

        def forbidden_load(*args, **kwargs):
            load_calls.append((args, kwargs))
            raise AssertionError("load_workbook must not be called")

        monkeypatch.setattr(
            spreadsheet.openpyxl,
            "load_workbook",
            forbidden_load,
        )

        with pytest.raises(InspectionError, match="worksheet cell"):
            inspect_artifact(
                "too-many-cells.xlsx",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )
        assert load_calls == []

    def test_xlsx_merge_area_rejects_before_openpyxl_load(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "huge-merge.xlsx"
        self._save_xlsx(path)
        from skillopt.envs.skilleval.inspectors import spreadsheet

        self._rewrite_xlsx_member(
            path,
            "xl/worksheets/sheet1.xml",
            lambda payload: payload.replace(
                b'ref="A5:B5"',
                b'ref="A1:XFD1048576"',
                1,
            ),
        )
        load_calls = []

        def forbidden_load(*args, **kwargs):
            load_calls.append((args, kwargs))
            raise AssertionError("load_workbook must not be called")

        monkeypatch.setattr(
            spreadsheet.openpyxl,
            "load_workbook",
            forbidden_load,
        )

        with pytest.raises(InspectionError, match="merged range area"):
            inspect_artifact(
                "huge-merge.xlsx",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )
        assert load_calls == []

    @pytest.mark.parametrize(
        "merge_ref",
        ["A0:B1", "A1:XFE1", "A1:A1048577", "not-a-range"],
    )
    def test_xlsx_preflight_rejects_invalid_merge_coordinates(
        self, tmp_path, merge_ref
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "invalid-merge.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_member(
            path,
            "xl/worksheets/sheet1.xml",
            lambda payload: payload.replace(
                b'ref="A5:B5"',
                f'ref="{merge_ref}"'.encode(),
                1,
            ),
        )

        with pytest.raises(InspectionError, match="merge reference"):
            preflight_xlsx(str(path))

    def test_xlsx_preflight_limits_cumulative_merged_cell_area(
        self, tmp_path, monkeypatch
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / "merged-area.xlsx"
        self._save_xlsx(path)

        def add_merge(payload: bytes) -> bytes:
            marker = b'<mergeCells count="1"><mergeCell ref="A5:B5"/></mergeCells>'
            replacement = (
                b'<mergeCells count="2"><mergeCell ref="A5:B5"/>'
                b'<mergeCell ref="C5:D5"/></mergeCells>'
            )
            assert marker in payload
            return payload.replace(marker, replacement, 1)

        self._rewrite_xlsx_member(
            path,
            "xl/worksheets/sheet1.xml",
            add_merge,
        )
        monkeypatch.setattr(
            spreadsheet,
            "_MAX_MERGED_CELLS",
            3,
            raising=False,
        )

        with pytest.raises(InspectionError, match="merged cell area"):
            spreadsheet.preflight_xlsx(str(path))

    def test_xlsx_preflight_accepts_100001_plain_rows(self, tmp_path) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "many-plain-rows.xlsx"
        self._save_xlsx(path)
        rows = b"".join(
            (
                f'<row r="{row}"><c r="A{row}" t="n">'
                "<v>1</v></c></row>"
            ).encode()
            for row in range(1, 100_002)
        )
        worksheet = (
            b'<worksheet xmlns="http://schemas.openxmlformats.org/'
            b'spreadsheetml/2006/main"><sheetData>'
            + rows
            + b"</sheetData></worksheet>"
        )
        self._rewrite_xlsx_member(
            path,
            "xl/worksheets/sheet1.xml",
            lambda _payload: worksheet,
        )

        preflight_xlsx(str(path))

    def test_xlsx_preflight_limits_materialized_row_dimensions(
        self, tmp_path, monkeypatch
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / "row-dimension.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_member(
            path,
            "xl/worksheets/sheet1.xml",
            lambda payload: payload.replace(
                b'<row r="1">',
                b'<row r="1" ht="15" customHeight="1">',
                1,
            ),
        )
        monkeypatch.setattr(
            spreadsheet,
            "_MAX_ROW_DIMENSIONS",
            0,
            raising=False,
        )

        with pytest.raises(InspectionError, match="row dimension"):
            spreadsheet.preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        ("member_name", "limit_name", "valid_tags", "closing_tag"),
        [
            (
                "xl/worksheets/sheet1.xml",
                "_MAX_WORKSHEET_CELLS",
                {"c"},
                b"</sheetData>",
            ),
            (
                "xl/styles.xml",
                "_MAX_STYLE_RECORDS",
                {
                    "numFmt",
                    "font",
                    "fill",
                    "border",
                    "xf",
                    "cellStyle",
                    "dxf",
                    "tableStyle",
                },
                b"</styleSheet>",
            ),
        ],
    )
    def test_xlsx_extension_namespace_nodes_do_not_count_as_structures(
        self,
        tmp_path,
        monkeypatch,
        member_name,
        limit_name,
        valid_tags,
        closing_tag,
    ) -> None:
        from xml.etree import ElementTree

        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / "extension-node.xlsx"
        self._save_xlsx(path)
        with zipfile.ZipFile(path) as archive:
            payload = archive.read(member_name)
        namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        valid_qnames = {f"{{{namespace}}}{name}" for name in valid_tags}
        valid_count = sum(
            element.tag in valid_qnames
            for element in ElementTree.fromstring(payload).iter()
        )
        extension_name = next(iter(valid_tags))
        extension = (
            f'<x:{extension_name} xmlns:x="urn:extension"/>'.encode()
        )
        self._rewrite_xlsx_member(
            path,
            member_name,
            lambda content: content.replace(
                closing_tag,
                extension + closing_tag,
                1,
            ),
        )
        monkeypatch.setattr(spreadsheet, limit_name, valid_count)

        spreadsheet.preflight_xlsx(str(path))

    def test_xlsx_extension_drawing_node_does_not_count_as_object(
        self, tmp_path, monkeypatch
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from xml.etree import ElementTree

        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / "extension-drawing.xlsx"
        book = Workbook()
        sheet = book.active
        sheet["A1"] = 1
        chart = BarChart()
        chart.add_data(Reference(sheet, min_col=1, min_row=1))
        sheet.add_chart(chart, "C1")
        book.save(path)
        book.close()
        member_name = "xl/drawings/drawing1.xml"
        with zipfile.ZipFile(path) as archive:
            payload = archive.read(member_name)
        namespace = (
            "http://schemas.openxmlformats.org/drawingml/2006/"
            "spreadsheetDrawing"
        )
        valid_qnames = {
            f"{{{namespace}}}{name}"
            for name in {"sp", "pic", "graphicFrame", "cxnSp", "grpSp"}
        }
        valid_count = sum(
            element.tag in valid_qnames
            for element in ElementTree.fromstring(payload).iter()
        )
        self._rewrite_xlsx_member(
            path,
            member_name,
            lambda content: content.replace(
                b"</wsDr>",
                b'<x:graphicFrame xmlns:x="urn:extension"/></wsDr>',
                1,
            ),
        )
        monkeypatch.setattr(
            spreadsheet,
            "_MAX_DRAWING_OBJECTS",
            valid_count,
        )

        spreadsheet.preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        ("probe", "member_name", "marker", "injection", "reason"),
        [
            (
                "worksheet",
                "xl/worksheets/sheet1.xml",
                b"</row>",
                b'<x:c xmlns:x="urn:extension" r="C1" t="n">'
                b"<x:v>1</x:v></x:c></row>",
                "worksheet",
            ),
            (
                "styles",
                "xl/styles.xml",
                b"</cellXfs>",
                b'<x:xf xmlns:x="urn:extension" numFmtId="0"/>'
                b"</cellXfs>",
                "styles",
            ),
        ],
    )
    def test_xlsx_rejects_foreign_parser_sensitive_child_before_load(
        self,
        tmp_path,
        monkeypatch,
        probe,
        member_name,
        marker,
        injection,
        reason,
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / f"foreign-{probe}-child.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_member(
            path,
            member_name,
            lambda payload: payload.replace(marker, injection, 1),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            reason,
        )

    def test_xlsx_rejects_parser_sensitive_child_at_wrong_parent_before_load(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "misnested-style-record.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_member(
            path,
            "xl/styles.xml",
            lambda payload: payload.replace(
                b"</styleSheet>",
                b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>'
                b"</styleSheet>",
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            "styles record has invalid nesting",
        )

    @pytest.mark.parametrize(
        ("probe", "member_name", "marker", "injection", "reason"),
        [
            (
                "drawing",
                "xl/drawings/drawing1.xml",
                b"</oneCellAnchor>",
                b'<x:graphicFrame xmlns:x="urn:extension"/>'
                b"</oneCellAnchor>",
                "drawing",
            ),
            (
                "chart",
                "xl/charts/chart1.xml",
                b"</plotArea>",
                b'<x:barChart xmlns:x="urn:extension"/>'
                b"</plotArea>",
                "chart",
            ),
        ],
    )
    def test_xlsx_rejects_foreign_drawing_chart_child_before_load(
        self,
        tmp_path,
        monkeypatch,
        probe,
        member_name,
        marker,
        injection,
        reason,
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference

        evidence, scratch = _roots(tmp_path)
        path = evidence / f"foreign-{probe}-child.xlsx"
        book = Workbook()
        sheet = book.active
        sheet["A1"] = 1
        chart = BarChart()
        chart.add_data(Reference(sheet, min_col=1, min_row=1))
        sheet.add_chart(chart, "C1")
        book.save(path)
        book.close()
        self._rewrite_xlsx_member(
            path,
            member_name,
            lambda payload: payload.replace(marker, injection, 1),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            reason,
        )

    def test_xlsx_preflight_accepts_extension_payload_in_extlst(
        self, tmp_path
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "worksheet-extension.xlsx"
        self._save_xlsx(path)
        extension = (
            b'<extLst><ext uri="{00000000-0000-0000-0000-000000000000}">'
            b'<x14:row xmlns:x14="http://schemas.microsoft.com/office/'
            b'spreadsheetml/2009/9/main"><x14:c/><x14:mergeCells/>'
            b"<x14:si/></x14:row>"
            b"</ext></extLst>"
        )
        self._rewrite_xlsx_member(
            path,
            "xl/worksheets/sheet1.xml",
            lambda payload: payload.replace(
                b"</worksheet>",
                extension + b"</worksheet>",
                1,
            ),
        )

        preflight_xlsx(str(path))

    def test_xlsx_rejects_main_namespace_row_in_extlst_before_load(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "extlst-main-row.xlsx"
        self._save_xlsx(path)
        extension = (
            b'<extLst><ext uri="{00000000-0000-0000-0000-000000000000}">'
            b'<row r="1048576"><c r="XFD1048576" t="n">'
            b"<v>1</v></c></row></ext></extLst>"
        )
        self._rewrite_xlsx_member(
            path,
            "xl/worksheets/sheet1.xml",
            lambda payload: payload.replace(
                b"</worksheet>",
                extension + b"</worksheet>",
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            "worksheet",
        )

    def test_xlsx_rejects_huge_main_namespace_merge_in_extlst_before_load(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "extlst-main-merge.xlsx"
        self._save_xlsx(path)
        extension = (
            b'<extLst><ext uri="{00000000-0000-0000-0000-000000000000}">'
            b'<mergeCells count="1">'
            b'<mergeCell ref="A1:XFD1048576"/>'
            b"</mergeCells></ext></extLst>"
        )
        self._rewrite_xlsx_member(
            path,
            "xl/worksheets/sheet1.xml",
            lambda payload: payload.replace(
                b"</worksheet>",
                extension + b"</worksheet>",
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            "worksheet",
        )

    def test_xlsx_rejects_main_namespace_shared_string_in_extlst_before_load(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "extlst-main-shared-string.xlsx"
        self._save_xlsx(path)
        self._add_shared_strings_part(path)
        extension = (
            b'<extLst><ext uri="{00000000-0000-0000-0000-000000000000}">'
            b"<si><t>hidden</t></si></ext></extLst>"
        )
        self._rewrite_xlsx_member(
            path,
            "xl/sharedStrings.xml",
            lambda payload: payload.replace(
                b"</sst>",
                extension + b"</sst>",
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            "shared string",
        )

    def test_xlsx_accepts_x14_shared_string_in_extlst_without_counting_it(
        self, tmp_path, monkeypatch
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / "extlst-x14-shared-string.xlsx"
        self._save_xlsx(path)
        self._add_shared_strings_part(path)
        extension = (
            b'<extLst><ext uri="{00000000-0000-0000-0000-000000000000}">'
            b'<x14:si xmlns:x14="http://schemas.microsoft.com/office/'
            b'spreadsheetml/2009/9/main"><x14:t>ignored</x14:t>'
            b"</x14:si></ext></extLst>"
        )
        self._rewrite_xlsx_member(
            path,
            "xl/sharedStrings.xml",
            lambda payload: payload.replace(
                b"</sst>",
                extension + b"</sst>",
                1,
            ),
        )
        monkeypatch.setattr(
            spreadsheet,
            "_MAX_SHARED_STRING_ITEMS",
            1,
        )

        spreadsheet.preflight_xlsx(str(path))

    def test_xlsx_structure_limits_follow_content_type_not_part_path(
        self, tmp_path, monkeypatch
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / "renamed-worksheet.xlsx"
        self._save_xlsx(path)
        renamed = tmp_path / "renamed-package.xlsx"
        old_name = "xl/worksheets/sheet1.xml"
        new_name = "xl/custom/sheet-data.xml"
        with zipfile.ZipFile(path) as source:
            members = [
                (member.filename, source.read(member))
                for member in source.infolist()
            ]
        with zipfile.ZipFile(renamed, "w") as destination:
            for name, payload in members:
                if name == "[Content_Types].xml":
                    payload = payload.replace(
                        f"/{old_name}".encode(),
                        f"/{new_name}".encode(),
                    )
                elif name == "xl/_rels/workbook.xml.rels":
                    payload = payload.replace(
                        f"/{old_name}".encode(),
                        f"/{new_name}".encode(),
                    )
                destination.writestr(
                    new_name if name == old_name else name,
                    payload,
                )
        os.replace(renamed, path)
        monkeypatch.setattr(spreadsheet, "_MAX_WORKSHEET_CELLS", 0)

        with pytest.raises(InspectionError, match="worksheet cell"):
            spreadsheet.preflight_xlsx(str(path))

    def test_xlsx_relationship_graph_accepts_renamed_worksheet(
        self, tmp_path
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "renamed-worksheet.xlsx"
        self._save_xlsx(path)
        old_name = "xl/worksheets/sheet1.xml"
        new_name = "xl/custom/sheet-data.bin"

        def rename(name: str, payload: bytes) -> tuple[str, bytes]:
            if name == "[Content_Types].xml":
                payload = payload.replace(
                    f"/{old_name}".encode(),
                    f"/{new_name}".encode(),
                )
            elif name == "xl/_rels/workbook.xml.rels":
                payload = payload.replace(
                    f"/{old_name}".encode(),
                    f"/{new_name}".encode(),
                )
            return (new_name if name == old_name else name), payload

        self._rewrite_xlsx_package(path, rename)

        preflight_xlsx(str(path))

    def test_xlsx_workbook_rejects_second_sheets_container_before_load(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "second-sheets.xlsx"
        self._save_xlsx(path)
        bypass_part = "xl/worksheets/bypass.xml"
        bypass_sheet = (
            b'<worksheet xmlns="http://schemas.openxmlformats.org/'
            b'spreadsheetml/2006/main"><sheetData/>'
            b'<mergeCells count="1"><mergeCell ref="A1:XFD1048576"/>'
            b"</mergeCells></worksheet>"
        )
        with zipfile.ZipFile(path, "a") as archive:
            archive.writestr(bypass_part, bypass_sheet)
        override = (
            f'<Override PartName="/{bypass_part}" '
            'ContentType="application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.worksheet+xml"/>'
        ).encode()
        relationship = (
            b'<Relationship Type="http://schemas.openxmlformats.org/'
            b'officeDocument/2006/relationships/worksheet" '
            b'Target="/xl/worksheets/bypass.xml" Id="rIdBypass"/>'
        )
        second_container = (
            b'<sheets><sheet xmlns:r="http://schemas.openxmlformats.org/'
            b'officeDocument/2006/relationships" name="Bypass" '
            b'sheetId="2" state="visible" r:id="rIdBypass"/></sheets>'
        )
        self._rewrite_xlsx_member(
            path,
            "[Content_Types].xml",
            lambda payload: payload.replace(
                b"</Types>",
                override + b"</Types>",
                1,
            ),
        )
        self._rewrite_xlsx_member(
            path,
            "xl/_rels/workbook.xml.rels",
            lambda payload: payload.replace(
                b"</Relationships>",
                relationship + b"</Relationships>",
                1,
            ),
        )
        self._rewrite_xlsx_member(
            path,
            "xl/workbook.xml",
            lambda payload: payload.replace(
                b"</sheets>",
                b"</sheets>" + second_container,
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            "workbook sheets",
        )

    def test_xlsx_workbook_rejects_foreign_sheet_before_load(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "foreign-sheet.xlsx"
        self._save_xlsx(path)
        foreign_sheet = (
            b'<x:sheet xmlns:x="urn:extension" '
            b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/'
            b'2006/relationships" name="Foreign" sheetId="2" '
            b'state="visible" r:id="rId1"/>'
        )
        self._rewrite_xlsx_member(
            path,
            "xl/workbook.xml",
            lambda payload: payload.replace(
                b"</sheets>",
                foreign_sheet + b"</sheets>",
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            "workbook sheet",
        )

    def test_xlsx_workbook_rejects_nested_sheet_before_load(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "nested-sheet.xlsx"
        self._save_xlsx(path)
        nested_sheet = (
            b'<extLst><ext uri="{00000000-0000-0000-0000-000000000000}">'
            b'<sheet xmlns:r="http://schemas.openxmlformats.org/'
            b'officeDocument/2006/relationships" name="Nested" '
            b'sheetId="2" state="visible" r:id="rId1"/>'
            b"</ext></extLst>"
        )
        self._rewrite_xlsx_member(
            path,
            "xl/workbook.xml",
            lambda payload: payload.replace(
                b"</workbook>",
                nested_sheet + b"</workbook>",
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            "workbook sheet",
        )

    def test_xlsx_workbook_rejects_external_sheet_relationship_before_load(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "external-sheet.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_member(
            path,
            "xl/_rels/workbook.xml.rels",
            lambda payload: payload.replace(
                b'Target="/xl/worksheets/sheet1.xml"',
                (
                    b'Target="https://example.invalid/sheet.xml" '
                    b'TargetMode="External"'
                ),
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            "worksheet relationship must target an internal part",
        )

    @pytest.mark.parametrize(
        ("role", "relationship_suffix"),
        [
            ("styles", "/styles"),
            ("shared strings", "/sharedStrings"),
        ],
    )
    def test_xlsx_rejects_parser_part_without_workbook_relationship(
        self,
        tmp_path,
        monkeypatch,
        role,
        relationship_suffix,
    ) -> None:
        from xml.etree import ElementTree

        evidence, scratch = _roots(tmp_path)
        path = evidence / f"missing-{role.replace(' ', '-')}-relationship.xlsx"
        self._save_xlsx(path)
        if role == "shared strings":
            self._add_shared_strings_part(path)

        def remove_relationship(payload: bytes) -> bytes:
            root = ElementTree.fromstring(payload)
            for child in list(root):
                if child.attrib.get("Type", "").endswith(relationship_suffix):
                    root.remove(child)
            return ElementTree.tostring(root)

        self._rewrite_xlsx_member(
            path,
            "xl/_rels/workbook.xml.rels",
            remove_relationship,
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            role,
        )

    @pytest.mark.parametrize(
        ("role", "relationship_type", "target"),
        [
            (
                "styles",
                "http://schemas.openxmlformats.org/officeDocument/"
                "2006/relationships/styles",
                "/xl/styles.xml",
            ),
            (
                "shared strings",
                "http://schemas.openxmlformats.org/officeDocument/"
                "2006/relationships/sharedStrings",
                "/xl/sharedStrings.xml",
            ),
        ],
    )
    def test_xlsx_rejects_duplicate_parser_part_relationship(
        self,
        tmp_path,
        monkeypatch,
        role,
        relationship_type,
        target,
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / f"duplicate-{role.replace(' ', '-')}-relationship.xlsx"
        self._save_xlsx(path)
        if role == "shared strings":
            self._add_shared_strings_part(path)
        duplicate = (
            f'<Relationship Type="{relationship_type}" '
            f'Target="{target}" Id="rIdDuplicate"/>'
        ).encode()
        self._rewrite_xlsx_member(
            path,
            "xl/_rels/workbook.xml.rels",
            lambda payload: payload.replace(
                b"</Relationships>",
                duplicate + b"</Relationships>",
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            role,
        )

    @pytest.mark.parametrize(
        ("role", "source_part", "duplicate_part", "content_type"),
        [
            (
                "styles",
                "xl/styles.xml",
                "xl/custom/styles2.xml",
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.styles+xml",
            ),
            (
                "shared strings",
                "xl/sharedStrings.xml",
                "xl/custom/shared2.xml",
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sharedStrings+xml",
            ),
        ],
    )
    def test_xlsx_rejects_multiple_parser_parts_by_content_type(
        self,
        tmp_path,
        monkeypatch,
        role,
        source_part,
        duplicate_part,
        content_type,
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / f"multiple-{role.replace(' ', '-')}-parts.xlsx"
        self._save_xlsx(path)
        if role == "shared strings":
            self._add_shared_strings_part(path)
        with zipfile.ZipFile(path) as archive:
            duplicate_payload = archive.read(source_part)
        with zipfile.ZipFile(path, "a") as archive:
            archive.writestr(duplicate_part, duplicate_payload)
        override = (
            f'<Override PartName="/{duplicate_part}" '
            f'ContentType="{content_type}"/>'
        ).encode()
        self._rewrite_xlsx_member(
            path,
            "[Content_Types].xml",
            lambda payload: payload.replace(
                b"</Types>",
                override + b"</Types>",
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            role,
        )

    def test_xlsx_rejects_styles_relationship_targeting_custom_part(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "custom-styles-target.xlsx"
        self._save_xlsx(path)
        custom_part = "xl/custom/styles.xml"
        with zipfile.ZipFile(path) as archive:
            styles = archive.read("xl/styles.xml")
        with zipfile.ZipFile(path, "a") as archive:
            archive.writestr(custom_part, styles)
        override = (
            f'<Override PartName="/{custom_part}" '
            'ContentType="application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.styles+xml"/>'
        ).encode()
        self._rewrite_xlsx_member(
            path,
            "[Content_Types].xml",
            lambda payload: payload.replace(
                b"</Types>",
                override + b"</Types>",
                1,
            ),
        )
        self._rewrite_xlsx_member(
            path,
            "xl/_rels/workbook.xml.rels",
            lambda payload: payload.replace(
                b'Target="styles.xml"',
                b'Target="/xl/custom/styles.xml"',
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            "styles",
        )

    def test_xlsx_relationship_graph_rejects_sheet_content_type_mismatch(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "sheet-type-mismatch.xlsx"
        self._save_xlsx(path)
        from skillopt.envs.skilleval.inspectors import spreadsheet

        old_name = "xl/worksheets/sheet1.xml"
        new_name = "xl/custom/sheet-data.bin"
        worksheet_type = (
            b"application/vnd.openxmlformats-officedocument."
            b"spreadsheetml.worksheet+xml"
        )

        def mismatch(name: str, payload: bytes) -> tuple[str, bytes]:
            if name == "[Content_Types].xml":
                payload = payload.replace(
                    f"/{old_name}".encode(),
                    f"/{new_name}".encode(),
                ).replace(
                    worksheet_type,
                    b"application/octet-stream",
                    1,
                )
            elif name == "xl/_rels/workbook.xml.rels":
                payload = payload.replace(
                    f"/{old_name}".encode(),
                    f"/{new_name}".encode(),
                )
            return (new_name if name == old_name else name), payload

        self._rewrite_xlsx_package(path, mismatch)
        load_calls = []

        def forbidden_load(*args, **kwargs):
            load_calls.append((args, kwargs))
            raise AssertionError("load_workbook must not be called")

        monkeypatch.setattr(
            spreadsheet.openpyxl,
            "load_workbook",
            forbidden_load,
        )

        with pytest.raises(
            InspectionError,
            match="worksheet relationship content type",
        ):
            inspect_artifact(
                "sheet-type-mismatch.xlsx",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )
        assert load_calls == []

    def test_xlsx_relationship_graph_rejects_sheet_type_mismatch(
        self, tmp_path
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "sheet-relationship-type.xlsx"
        self._save_xlsx(path)
        worksheet_rel = (
            b"http://schemas.openxmlformats.org/officeDocument/"
            b"2006/relationships/worksheet"
        )
        chart_rel = (
            b"http://schemas.openxmlformats.org/officeDocument/"
            b"2006/relationships/chart"
        )
        self._rewrite_xlsx_member(
            path,
            "xl/_rels/workbook.xml.rels",
            lambda payload: payload.replace(
                worksheet_rel,
                chart_rel,
                1,
            ),
        )

        with pytest.raises(
            InspectionError,
            match="sheet relationship type",
        ):
            preflight_xlsx(str(path))

    def test_xlsx_relationship_graph_rejects_duplicate_sheet_relationship_id(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "duplicate-sheet-id.xlsx"
        self._save_xlsx(path)
        duplicate = (
            b'<sheet xmlns:r="http://schemas.openxmlformats.org/'
            b'officeDocument/2006/relationships" name="Other" '
            b'sheetId="2" state="visible" r:id="rId1"/>'
        )
        self._rewrite_xlsx_member(
            path,
            "xl/workbook.xml",
            lambda payload: payload.replace(
                b"</sheets>",
                duplicate + b"</sheets>",
                1,
            ),
        )

        self._assert_inspect_rejected_before_openpyxl(
            evidence,
            scratch,
            path.name,
            monkeypatch,
            "sheet relationship id",
        )

    @pytest.mark.parametrize(
        "target",
        ["../../xl/worksheets/sheet1.xml", "%2e%2e/%2e%2e/escape.xml"],
    )
    def test_xlsx_relationship_graph_rejects_traversal_target(
        self, tmp_path, target
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "relationship-traversal.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_member(
            path,
            "xl/_rels/workbook.xml.rels",
            lambda payload: payload.replace(
                b"/xl/worksheets/sheet1.xml",
                target.encode(),
                1,
            ),
        )

        with pytest.raises(
            InspectionError,
            match="relationship target",
        ):
            preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        ("part_name", "content_type", "relationship_role"),
        [
            (
                "xl/drawings/drawing1.xml",
                (
                    b"application/vnd.openxmlformats-officedocument."
                    b"drawing+xml"
                ),
                "drawing",
            ),
            (
                "xl/charts/chart1.xml",
                (
                    b"application/vnd.openxmlformats-officedocument."
                    b"drawingml.chart+xml"
                ),
                "chart",
            ),
        ],
    )
    def test_xlsx_relationship_graph_validates_drawing_and_chart_types(
        self,
        tmp_path,
        part_name,
        content_type,
        relationship_role,
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / f"{relationship_role}-type-mismatch.xlsx"
        book = Workbook()
        sheet = book.active
        sheet["A1"] = 1
        chart = BarChart()
        chart.add_data(Reference(sheet, min_col=1, min_row=1))
        sheet.add_chart(chart, "C1")
        book.save(path)
        book.close()

        def mismatch(name: str, payload: bytes) -> tuple[str, bytes]:
            if name == "[Content_Types].xml":
                marker = (
                    f'<Override PartName="/{part_name}" '.encode()
                    + b'ContentType="'
                    + content_type
                    + b'"/>'
                )
                assert marker in payload
                payload = payload.replace(
                    marker,
                    (
                        f'<Override PartName="/{part_name}" '
                        'ContentType="application/octet-stream"/>'
                    ).encode(),
                    1,
                )
            return name, payload

        self._rewrite_xlsx_package(path, mismatch)

        with pytest.raises(
            InspectionError,
            match=f"{relationship_role} relationship content type",
        ):
            preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        ("limit_name", "reason", "fixture_kind"),
        [
            ("_MAX_ROW_DIMENSIONS", "row dimension", "row-dimension"),
            ("_MAX_WORKSHEET_CELLS", "worksheet cell", "basic"),
            ("_MAX_WORKSHEET_MERGES", "worksheet merge", "basic"),
            ("_MAX_SHARED_STRING_ITEMS", "shared string item", "shared"),
            ("_MAX_SHARED_STRING_CHARS", "shared string character", "shared"),
            ("_MAX_STYLE_RECORDS", "style record", "basic"),
            ("_MAX_RELATIONSHIPS", "relationship", "basic"),
            ("_MAX_DRAWING_OBJECTS", "drawing object", "chart"),
            ("_MAX_CHART_OBJECTS", "chart object", "chart"),
        ],
    )
    def test_xlsx_preflight_enforces_xml_structure_limits(
        self,
        tmp_path,
        monkeypatch,
        limit_name,
        reason,
        fixture_kind,
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / f"{fixture_kind}.xlsx"
        self._save_xlsx(path)
        if fixture_kind == "row-dimension":
            self._rewrite_xlsx_member(
                path,
                "xl/worksheets/sheet1.xml",
                lambda payload: payload.replace(
                    b'<row r="1">',
                    b'<row r="1" hidden="1">',
                    1,
                ),
            )
        elif fixture_kind == "shared":
            shared_strings = (
                b'<?xml version="1.0" encoding="UTF-8"?>'
                b'<sst xmlns="http://schemas.openxmlformats.org/'
                b'spreadsheetml/2006/main" count="1" uniqueCount="1">'
                b"<si><t>Revenue</t></si></sst>"
            )
            with zipfile.ZipFile(path, "a") as archive:
                archive.writestr(
                    "xl/sharedStrings.xml",
                    shared_strings,
                )

            def add_shared_override(payload: bytes) -> bytes:
                override = (
                    b'<Override PartName="/xl/sharedStrings.xml" '
                    b'ContentType="application/vnd.openxmlformats-'
                    b'officedocument.spreadsheetml.sharedStrings+xml"/>'
                )
                return payload.replace(b"</Types>", override + b"</Types>")

            self._rewrite_xlsx_member(
                path,
                "[Content_Types].xml",
                add_shared_override,
            )
            shared_relationship = (
                b'<Relationship Type="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships/sharedStrings" '
                b'Target="/xl/sharedStrings.xml" Id="rIdShared"/>'
            )
            self._rewrite_xlsx_member(
                path,
                "xl/_rels/workbook.xml.rels",
                lambda payload: payload.replace(
                    b"</Relationships>",
                    shared_relationship + b"</Relationships>",
                    1,
                ),
            )
        elif fixture_kind == "chart":
            book = Workbook()
            data = book.active
            data["A1"] = 1
            chart = BarChart()
            chart.add_data(
                Reference(data, min_col=1, min_row=1, max_row=1)
            )
            data.add_chart(chart, "C1")
            book.save(path)
            book.close()

        monkeypatch.setattr(
            spreadsheet,
            limit_name,
            0,
            raising=False,
        )
        with pytest.raises(InspectionError, match=reason):
            spreadsheet.preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        ("limit_name", "reason"),
        [
            ("_MAX_OOXML_XML_PART_BYTES", "XML part bytes"),
            ("_MAX_OOXML_XML_NODES", "XML node"),
        ],
    )
    def test_xlsx_preflight_enforces_general_xml_limits(
        self,
        tmp_path,
        monkeypatch,
        limit_name,
        reason,
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / "xml-limits.xlsx"
        self._save_xlsx(path)
        monkeypatch.setattr(
            spreadsheet,
            limit_name,
            1,
            raising=False,
        )

        with pytest.raises(InspectionError, match=reason):
            spreadsheet.preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        "invalid_case",
        ["namespace", "child", "duplicate", "unmapped"],
    )
    def test_xlsx_content_types_schema_is_strict(
        self, tmp_path, invalid_case
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / f"content-types-{invalid_case}.xlsx"
        self._save_xlsx(path)
        if invalid_case == "unmapped":
            with zipfile.ZipFile(path, "a") as archive:
                archive.writestr("xl/unmapped.payload", b"payload")
        else:
            def invalidate(payload: bytes) -> bytes:
                if invalid_case == "namespace":
                    return payload.replace(
                        b"http://schemas.openxmlformats.org/package/"
                        b"2006/content-types",
                        b"urn:invalid-content-types",
                    )
                if invalid_case == "child":
                    return payload.replace(
                        b"</Types>",
                        b"<Bogus/></Types>",
                    )
                marker = b'<Override PartName="/xl/workbook.xml"'
                start = payload.index(marker)
                end = payload.index(b"/>", start) + 2
                return payload[:end] + payload[start:end] + payload[end:]

            self._rewrite_xlsx_member(
                path,
                "[Content_Types].xml",
                invalidate,
            )

        with pytest.raises(InspectionError, match="content types"):
            preflight_xlsx(str(path))

    def test_xlsx_preflight_rejects_unsupported_compression_and_content_type(
        self, tmp_path
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        unsupported = tmp_path / "unsupported.xlsx"
        self._save_xlsx(unsupported)
        with zipfile.ZipFile(
            unsupported,
            "a",
            compression=zipfile.ZIP_BZIP2,
        ) as archive:
            archive.writestr("xl/media/unsupported.bin", "payload")
        with pytest.raises(InspectionError, match="compression"):
            preflight_xlsx(str(unsupported))

        mismatched = tmp_path / "mismatched.xlsx"
        content_types = (
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Override PartName="/word/document.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.'
            'wordprocessingml.document.main+xml"/>'
            "</Types>"
        )
        with zipfile.ZipFile(mismatched, "w") as archive:
            archive.writestr("[Content_Types].xml", content_types)
            archive.writestr("xl/workbook.xml", "<workbook/>")
        with pytest.raises(InspectionError, match="content types"):
            preflight_xlsx(str(mismatched))

    def test_xlsx_preflight_wraps_decompression_errors(
        self, tmp_path, monkeypatch
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / "decompression-error.xlsx"
        self._save_xlsx(path)
        real_open = spreadsheet.zipfile.ZipFile.open

        def broken_open(archive, member, *args, **kwargs):
            if getattr(member, "filename", member) == "[Content_Types].xml":
                raise zlib.error("invalid distance code")
            return real_open(archive, member, *args, **kwargs)

        monkeypatch.setattr(
            spreadsheet.zipfile.ZipFile,
            "open",
            broken_open,
        )

        with pytest.raises(
            InspectionError,
            match="OOXML ZIP could not be validated",
        ):
            spreadsheet.preflight_xlsx(str(path))

    def test_xlsx_preflight_rejects_encrypted_nonregular_and_nul_entries(
        self,
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        encrypted = zipfile.ZipInfo("xl/encrypted.bin")
        encrypted.flag_bits = 0x1
        with pytest.raises(InspectionError, match="encrypted"):
            spreadsheet._validate_member_type(encrypted)

        linked = zipfile.ZipInfo("xl/linked.bin")
        linked.create_system = 3
        linked.external_attr = 0o120777 << 16
        with pytest.raises(InspectionError, match="non-regular"):
            spreadsheet._validate_member_type(linked)

        nul_name = zipfile.ZipInfo("xl/good.xml")
        nul_name.orig_filename = "xl/\x00evil.xml"
        with pytest.raises(InspectionError, match="unsafe"):
            spreadsheet._validate_member_name(nul_name)

        data_descriptor = zipfile.ZipInfo("xl/descriptor.bin")
        data_descriptor.flag_bits = 0x8
        spreadsheet._validate_member_type(data_descriptor)

    @pytest.mark.parametrize(
        ("zip64", "signature"),
        [
            (False, False),
            (False, True),
            (True, False),
            (True, True),
        ],
        ids=[
            "zip32-unsigned",
            "zip32-signed",
            "zip64-unsigned",
            "zip64-signed",
        ],
    )
    def test_xlsx_preflight_accepts_data_descriptors(
        self,
        tmp_path,
        zip64,
        signature,
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "descriptor.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_with_data_descriptors(
            path,
            zip64=zip64,
            signature=signature,
        )

        preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        (
            "zip64",
            "field_offset",
            "field_format",
            "central_attribute",
            "reason",
        ),
        [
            (False, 4, "<L", "CRC", "CRC"),
            (False, 8, "<L", "compress_size", "compressed size"),
            (False, 12, "<L", "file_size", "uncompressed size"),
            (True, 4, "<L", "CRC", "CRC"),
            (True, 8, "<Q", "compress_size", "compressed size"),
            (True, 16, "<Q", "file_size", "uncompressed size"),
        ],
        ids=[
            "zip32-crc",
            "zip32-compressed-size",
            "zip32-uncompressed-size",
            "zip64-crc",
            "zip64-compressed-size",
            "zip64-uncompressed-size",
        ],
    )
    def test_xlsx_preflight_rejects_corrupt_data_descriptor(
        self,
        tmp_path,
        zip64,
        field_offset,
        field_format,
        central_attribute,
        reason,
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "corrupt-descriptor.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_with_data_descriptors(
            path,
            zip64=zip64,
            signature=True,
        )
        member_name = "[Content_Types].xml"
        descriptor_offset = self._data_descriptor_offset(path, member_name)
        with zipfile.ZipFile(path) as archive:
            expected = getattr(archive.getinfo(member_name), central_attribute)
        with path.open("r+b") as target:
            target.seek(descriptor_offset + field_offset)
            target.write(struct.pack(field_format, expected ^ 1))

        with pytest.raises(
            InspectionError,
            match=f"data descriptor {reason}",
        ):
            preflight_xlsx(str(path))

    def test_xlsx_preflight_rejects_truncated_data_descriptor(
        self, tmp_path
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "truncated-descriptor.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_with_data_descriptors(
            path,
            zip64=False,
            signature=True,
            descriptor_transform=lambda name, descriptor: (
                descriptor[:-1]
                if name == "[Content_Types].xml"
                else descriptor
            ),
        )

        with pytest.raises(
            InspectionError,
            match="data descriptor.*truncated",
        ):
            preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        ("zip64", "descriptor_format"),
        [(False, "<3L"), (True, "<LQQ")],
        ids=["zip32", "zip64"],
    )
    def test_xlsx_data_descriptor_accepts_unsigned_signature_crc(
        self,
        tmp_path,
        zip64,
        descriptor_format,
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        member = zipfile.ZipInfo("xl/collision.xml")
        member.CRC = 0x08074B50
        member.compress_size = 7
        member.file_size = 11
        descriptor = struct.pack(
            descriptor_format,
            member.CRC,
            member.compress_size,
            member.file_size,
        )
        path = tmp_path / "descriptor.bin"
        path.write_bytes(descriptor)
        file_descriptor = os.open(path, os.O_RDONLY)
        try:
            assert spreadsheet._validate_data_descriptor(
                file_descriptor,
                member,
                0,
                len(descriptor),
                zip64=zip64,
            ) == len(descriptor)
        finally:
            os.close(file_descriptor)

    def test_xlsx_preflight_rejects_wrong_data_descriptor_width(
        self, tmp_path
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        def widen_descriptor(name, descriptor):
            if name != "[Content_Types].xml":
                return descriptor
            signature, crc, compressed, uncompressed = struct.unpack(
                "<4s3L",
                descriptor,
            )
            return struct.pack(
                "<4sLQQ",
                signature,
                crc,
                compressed,
                uncompressed,
            )

        path = tmp_path / "wrong-width-descriptor.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_with_data_descriptors(
            path,
            zip64=False,
            signature=True,
            descriptor_transform=widen_descriptor,
        )

        with pytest.raises(
            InspectionError,
            match="data descriptor has wrong width",
        ):
            preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        ("relative_offset", "replacement"),
        [
            (14, struct.pack("<L", 1)),
            (18, struct.pack("<L", 1)),
        ],
        ids=["crc", "size"],
    )
    def test_xlsx_preflight_rejects_conflicting_descriptor_placeholders(
        self,
        tmp_path,
        relative_offset,
        replacement,
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "descriptor-placeholders.xlsx"
        self._save_xlsx(path)
        self._rewrite_xlsx_with_data_descriptors(
            path,
            zip64=False,
            signature=True,
        )
        self._mutate_local_header(
            path,
            "[Content_Types].xml",
            relative_offset,
            replacement,
        )

        with pytest.raises(
            InspectionError,
            match="local file header descriptor placeholders",
        ):
            preflight_xlsx(str(path))

    @pytest.mark.parametrize(
        ("relative_offset", "replacement"),
        [
            (0, b"BAD!"),
            (6, struct.pack("<H", 0x1)),
            (6, struct.pack("<H", 0x8)),
            (8, struct.pack("<H", zipfile.ZIP_STORED)),
            (14, struct.pack("<L", 0)),
            (18, struct.pack("<L", 0)),
            (22, struct.pack("<L", 0)),
            (30, b"X"),
        ],
        ids=[
            "signature",
            "encrypted-flags",
            "descriptor-flags",
            "method",
            "crc",
            "compressed-size",
            "uncompressed-size",
            "name",
        ],
    )
    def test_xlsx_preflight_rejects_local_header_mismatch(
        self,
        tmp_path,
        relative_offset,
        replacement,
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "local-header.xlsx"
        self._save_xlsx(path)
        self._mutate_local_header(
            path,
            "[Content_Types].xml",
            relative_offset,
            replacement,
        )

        with pytest.raises(InspectionError, match="local file header"):
            preflight_xlsx(str(path))

    def test_xlsx_preflight_rejects_overlapping_local_regions(
        self, tmp_path
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "overlapping-local-regions.xlsx"
        self._save_xlsx(path)
        with zipfile.ZipFile(path) as archive:
            first, second = sorted(
                archive.infolist(),
                key=lambda member: member.header_offset,
            )[:2]
        with path.open("r+b") as target:
            target.seek(first.header_offset)
            fields = struct.unpack("<4s5H3L2H", target.read(30))
            filename_size, extra_size = fields[-2:]
            assert extra_size == 0
            data_start = first.header_offset + 30 + filename_size
            assert data_start + first.compress_size == second.header_offset
            target.seek(first.header_offset + 28)
            target.write(struct.pack("<H", 4))
            target.seek(data_start)
            target.write(struct.pack("<HH", 0xCAFE, 0))

        with pytest.raises(InspectionError, match="regions overlap"):
            preflight_xlsx(str(path))

    def test_xlsx_preflight_rejects_local_header_out_of_bounds(
        self, tmp_path
    ) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        path = tmp_path / "out-of-bounds-local-header.xlsx"
        self._save_xlsx(path)
        with zipfile.ZipFile(path) as archive:
            final_member = max(
                archive.infolist(),
                key=lambda member: member.header_offset,
            )
        self._mutate_local_header(
            path,
            final_member.filename,
            28,
            struct.pack("<H", 0xFFFF),
        )

        with pytest.raises(InspectionError, match="outside archive bounds"):
            preflight_xlsx(str(path))

    def test_xlsx_preflight_accepts_force_zip64_entries(self, tmp_path) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            preflight_xlsx,
        )

        source_path = tmp_path / "source.xlsx"
        zip64_path = tmp_path / "zip64.xlsx"
        self._save_xlsx(source_path)
        with (
            zipfile.ZipFile(source_path) as source,
            zipfile.ZipFile(zip64_path, "w") as destination,
        ):
            for source_info in source.infolist():
                target_info = zipfile.ZipInfo(
                    source_info.filename,
                    source_info.date_time,
                )
                target_info.compress_type = source_info.compress_type
                target_info.external_attr = source_info.external_attr
                target_info.create_system = source_info.create_system
                with destination.open(
                    target_info,
                    "w",
                    force_zip64=True,
                ) as target:
                    target.write(source.read(source_info))

        preflight_xlsx(str(zip64_path))

    def test_xlsx_corrupt_workbook_raises_controlled_error(self, tmp_path) -> None:
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            evaluate_xlsx_check,
        )

        path = tmp_path / "corrupt.xlsx"
        path.write_bytes(b"PK\x03\x04corrupt")
        check = {
            "id": "cell",
            "type": "xlsx_cell",
            "path": "corrupt.xlsx",
            "spec": {"sheet": "Summary", "cell": "A1", "value": 1},
        }

        with pytest.raises(InspectionError, match="OOXML|workbook"):
            evaluate_xlsx_check(str(path), check)

    def test_xlsx_render_converts_to_pdf_and_delegates_with_budget(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        self._save_xlsx(evidence / "report.xlsx")

        outputs = render_artifact(
            "report.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["page:1"],
            max_pixels=321,
        )

        render_call = fake_pdf_inspector.calls[-1]
        assert render_call[0] == "render"
        assert render_call[1].endswith(".pdf")
        assert render_call[3] == ["page:1"]
        assert render_call[4].max_pixels == 321
        assert all("lo-profile" not in output for output in outputs)
        assert all("lo-render" not in output for output in outputs)

    def test_xlsx_render_preflights_before_libreoffice(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        path = evidence / "unsafe-render.xlsx"
        self._save_xlsx(path)
        with zipfile.ZipFile(path, "a") as archive:
            archive.writestr("XL/workbook.xml", "collision")
        from skillopt.envs.skilleval.inspectors import spreadsheet

        monkeypatch.setattr(
            spreadsheet,
            "_convert_with_libreoffice",
            lambda *args, **kwargs: pytest.fail(
                "converter must not run before XLSX preflight"
            ),
        )

        with pytest.raises(InspectionError, match="colliding OOXML entry"):
            render_artifact(
                "unsafe-render.xlsx",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_xls_render_does_not_run_ooxml_preflight(
        self,
        tmp_path,
        fake_pdf_inspector,
        monkeypatch,
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        (evidence / "legacy.xls").write_bytes(
            b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy"
        )
        from skillopt.envs.skilleval.inspectors import spreadsheet

        def fake_convert(path, scratch_dir, target_format):
            output = Path(scratch_dir) / "legacy.pdf"
            output.write_bytes(b"%PDF-1.4\n")
            return str(output)

        monkeypatch.setattr(
            spreadsheet,
            "preflight_xlsx",
            lambda path: pytest.fail("XLS must not use OOXML preflight"),
        )
        monkeypatch.setattr(
            spreadsheet,
            "_convert_with_libreoffice",
            fake_convert,
        )

        outputs = render_artifact(
            "legacy.xls",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert len(outputs) == 1
        assert fake_pdf_inspector.calls[-1][0] == "render"

    def test_xlsx_render_rejects_non_pdf_selector_before_conversion(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        self._save_xlsx(evidence / "report.xlsx")
        from skillopt.envs.skilleval.inspectors import spreadsheet

        monkeypatch.setattr(
            spreadsheet,
            "_convert_with_libreoffice",
            lambda *args, **kwargs: pytest.fail("conversion must not run"),
        )

        with pytest.raises(InspectionError, match="render selector"):
            render_artifact(
                "report.xlsx",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                selectors=["sheet:Summary"],
            )

    def test_libreoffice_xls_conversion_uses_fresh_hardened_profile(
        self, tmp_path, monkeypatch
    ) -> None:
        from openpyxl import Workbook
        from skillopt.envs.skilleval.inspectors import spreadsheet
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            evaluate_xlsx_check,
        )

        path = tmp_path / "legacy.xls"
        path.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy")
        observed = {}

        def fake_run(command, *, timeout, cwd, home, pass_fds=()):
            observed["command"] = command
            observed["cwd"] = cwd
            observed["home"] = home
            profile_arg = next(
                part
                for part in command
                if part.startswith("-env:UserInstallation=")
            )
            observed["profile_arg"] = profile_arg
            profile_path = profile_arg.split("file://", 1)[1]
            observed["profile_config"] = (
                Path(profile_path) / "user" / "registrymodifications.xcu"
            ).read_text(encoding="utf-8")
            replacement = tmp_path / "replacement.xls"
            replacement.write_bytes(b"REPLACED")
            os.replace(replacement, path)
            observed["input_path"] = command[-1]
            observed["input_payload"] = Path(command[-1]).read_bytes()
            out_dir = Path(command[command.index("--outdir") + 1])
            book = Workbook()
            book.active.title = "Summary"
            book.active["A1"] = 7
            book.save(out_dir / f"{Path(command[-1]).stem}.xlsx")
            book.close()
            return subprocess.CompletedProcess(command, 0, "", "")

        monkeypatch.setattr(
            spreadsheet,
            "_find_libreoffice",
            lambda: "/mock/libreoffice",
        )
        monkeypatch.setattr(spreadsheet, "safe_run", fake_run)
        check = {
            "id": "legacy",
            "type": "xlsx_cell",
            "path": "legacy.xls",
            "spec": {"sheet": "Summary", "cell": "A1", "value": 7},
        }

        assert evaluate_xlsx_check(str(path), check)["passed"] is True
        assert "--headless" in observed["command"]
        assert observed["profile_arg"].startswith(
            "-env:UserInstallation=file://"
        )
        assert str(tmp_path) not in observed["profile_arg"]
        assert observed["home"] != str(tmp_path)
        assert "MacroSecurityLevel" in observed["profile_config"]
        assert "org.openoffice.Office.Calc/Content/Update" in (
            observed["profile_config"]
        )
        assert '<prop oor:name="Link"' in observed["profile_config"]
        assert observed["input_path"].startswith("/proc/self/fd/")
        assert observed["input_payload"].startswith(
            b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
        )

    def test_libreoffice_conversion_requires_expected_output(
        self, tmp_path, monkeypatch
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            evaluate_xlsx_check,
        )

        path = tmp_path / "legacy.xls"
        path.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy")
        monkeypatch.setattr(
            spreadsheet,
            "_find_libreoffice",
            lambda: "/mock/libreoffice",
        )
        monkeypatch.setattr(
            spreadsheet,
            "safe_run",
            lambda command, **kwargs: subprocess.CompletedProcess(
                command, 0, "", ""
            ),
        )
        check = {
            "id": "legacy",
            "type": "xlsx_cell",
            "path": "legacy.xls",
            "spec": {"sheet": "Summary", "cell": "A1", "value": 7},
        }

        with pytest.raises(InspectionError, match="expected output"):
            evaluate_xlsx_check(str(path), check)

    @pytest.mark.parametrize("transient_error_kind", ["entry", "profile-dir"])
    def test_libreoffice_conversion_retries_transient_scratch_disappearance(
        self, tmp_path, monkeypatch, transient_error_kind
    ) -> None:
        from openpyxl import Workbook
        from skillopt.envs.skilleval.inspectors import spreadsheet
        from skillopt.envs.skilleval.inspectors.spreadsheet import (
            evaluate_xlsx_check,
        )

        path = tmp_path / "legacy.xls"
        path.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy")
        attempts = []

        def fake_run(command, **kwargs):
            attempts.append((command, kwargs))
            if len(attempts) == 1:
                if transient_error_kind == "entry":
                    detail = (
                        "[Errno 2] No such file or directory: "
                        "'backenddb.xml'"
                    )
                else:
                    profile_arg = next(
                        part
                        for part in command
                        if part.startswith("-env:UserInstallation=")
                    )
                    profile_name = Path(
                        profile_arg.split("file://", 1)[1]
                    ).name
                    detail = (
                        "scratch directory could not be scanned: "
                        f"work/{profile_name}/user/uno_packages/cache/"
                        "registry/PackageRegistryBackend"
                    )
                raise InspectionError(
                    f"inspector command scratch failure: {detail}"
                )
            out_dir = Path(command[command.index("--outdir") + 1])
            book = Workbook()
            book.active.title = "Summary"
            book.active["A1"] = 7
            book.save(out_dir / f"{Path(command[-1]).stem}.xlsx")
            book.close()
            return subprocess.CompletedProcess(command, 0, "", "")

        monkeypatch.setattr(
            spreadsheet,
            "_find_libreoffice",
            lambda: "/mock/libreoffice",
        )
        monkeypatch.setattr(spreadsheet, "safe_run", fake_run)
        check = {
            "id": "legacy-retry",
            "type": "xlsx_cell",
            "path": "legacy.xls",
            "spec": {"sheet": "Summary", "cell": "A1", "value": 7},
        }

        assert evaluate_xlsx_check(str(path), check)["passed"] is True
        assert len(attempts) == 2
        assert attempts[1][1]["timeout"] < attempts[0][1]["timeout"]

    def test_libreoffice_lookup_failure_does_not_leak_input_fd(
        self, tmp_path, monkeypatch
    ) -> None:
        from skillopt.envs.skilleval.inspectors import spreadsheet

        path = tmp_path / "legacy.xls"
        path.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy")
        scratch = tmp_path / "scratch"
        scratch.mkdir()

        def missing_libreoffice():
            raise InspectionError("LibreOffice is unavailable")

        monkeypatch.setattr(
            spreadsheet,
            "_find_libreoffice",
            missing_libreoffice,
        )
        with scratch_transaction(
            str(scratch),
            max_bytes=inspector_base.DEFAULT_SCRATCH_BYTES,
            max_entries=inspector_base.DEFAULT_SCRATCH_ENTRIES,
            max_depth=inspector_base.DEFAULT_SCRATCH_DEPTH,
        ) as transaction:
            before = set(os.listdir("/proc/self/fd"))
            for _ in range(3):
                with pytest.raises(
                    InspectionError,
                    match="LibreOffice is unavailable",
                ):
                    spreadsheet._convert_with_libreoffice(
                        str(path),
                        transaction.proc_path,
                        "xlsx",
                    )
            assert set(os.listdir("/proc/self/fd")) == before

    @pytest.mark.skipif(
        shutil.which("libreoffice") is None
        and shutil.which("soffice") is None,
        reason="LibreOffice is not installed",
    )
    def test_libreoffice_xls_round_trip_inspect_smoke(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        seed = tmp_path / "seed.xlsx"
        self._save_xlsx(seed)
        profile = tmp_path / "xls-seed-profile"
        output_dir = tmp_path / "xls-seed-output"
        profile.mkdir()
        output_dir.mkdir()
        executable = shutil.which("libreoffice") or shutil.which("soffice")
        command = [
            executable,
            "--headless",
            "--invisible",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--norestore",
            f"-env:UserInstallation={profile.as_uri()}",
            "--convert-to",
            "xls:MS Excel 97",
            "--outdir",
            str(output_dir),
            str(seed),
        ]
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
        )
        assert completed.returncode == 0, (
            completed.stdout,
            completed.stderr,
        )
        generated = output_dir / "seed.xls"
        assert generated.read_bytes().startswith(
            b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
        )
        os.replace(generated, evidence / "report.xls")

        inspected = inspect_artifact(
            "report.xls",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert inspected["kind"] == "spreadsheet"
        assert inspected["opens"] is True
        assert inspected["sheets"][0]["name"] == "Summary"
        assert inspected["sheets"][0]["cells"]["A1"]["value"] == "Revenue"

    @pytest.mark.skipif(
        shutil.which("libreoffice") is None,
        reason="LibreOffice is not installed",
    )
    def test_libreoffice_xlsx_render_smoke(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        self._save_xlsx(evidence / "report.xlsx")

        outputs = render_artifact(
            "report.xlsx",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["page:1"],
            max_pixels=10_000,
        )

        assert len(outputs) == 1
        assert Path(outputs[0]).is_file()
        assert fake_pdf_inspector.calls[-1][0] == "render"


class TestPdfImageInspectors:
    def test_image_inspection_reports_rgba_metadata(self, tmp_path) -> None:
        evidence, scratch = _roots(tmp_path)
        PillowImage.new(
            "RGBA",
            (40, 20),
            color=(255, 0, 0, 128),
        ).save(evidence / "preview.png")

        result = inspect_artifact(
            "preview.png",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert result == {
            "kind": "image",
            "opens": True,
            "format": "PNG",
            "width": 40,
            "height": 20,
            "mode": "RGBA",
            "frames": 1,
            "has_transparency": True,
        }

    def test_pdf_inspection_indexes_metadata_and_extracts_text_separately(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        commands = []

        def fake_run(command, **kwargs):
            commands.append((command, kwargs))
            if command[0] == "pdfinfo":
                stdout = "Pages: 2\nTitle: Report\n"
            else:
                stdout = "Quarterly report\fSecond page\f"
            return subprocess.CompletedProcess(command, 0, stdout, "")

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)

        inspected = inspect_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )
        extracted = extract_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["page:1-2"],
        )

        assert inspected["kind"] == "pdf"
        assert inspected["opens"] is True
        assert inspected["pages"] == 2
        assert inspected["metadata"]["Title"] == "Report"
        assert inspected["page_index"] == {
            "items": [1, 2],
            "returned": 2,
            "omitted": 0,
        }
        assert "text" not in inspected
        assert [page["text"] for page in extracted["page_text"]] == [
            "Quarterly report",
            "Second page",
        ]
        assert [command[0][0] for command in commands] == [
            "pdfinfo",
            "pdfinfo",
            "pdftotext",
        ]

    def test_multiframe_image_index_and_extract_selectors(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        first = PillowImage.new("RGBA", (8, 6), (255, 0, 0, 128))
        second = PillowImage.new("RGBA", (4, 3), (0, 0, 255, 64))
        first.save(
            evidence / "animation.tiff",
            save_all=True,
            append_images=[second],
        )
        first.close()
        second.close()

        inspected = inspect_artifact(
            "animation.tiff",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )
        extracted = extract_artifact(
            "animation.tiff",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["frame:2", "metadata"],
        )

        assert inspected["frames"] == 2
        assert inspected["has_transparency"] is True
        assert inspected["frame_index"] == {
            "items": [
                {
                    "frame": 1,
                    "width": 8,
                    "height": 6,
                    "mode": "RGBA",
                    "has_transparency": True,
                },
                {
                    "frame": 2,
                    "width": 4,
                    "height": 3,
                    "mode": "RGBA",
                    "has_transparency": True,
                },
            ],
            "returned": 2,
            "omitted": 0,
        }
        assert extracted["frame_data"] == [
            inspected["frame_index"]["items"][1]
        ]
        assert extracted["metadata"]["format"] == "TIFF"
        assert extracted["units_inspected"] == ["frame:2", "metadata"]
        assert extracted["omitted"] == {"frames": 1, "metadata": 0}
        assert extracted["truncated"] is False
        assert extracted["next_cursor"] is None

    def test_multiframe_image_extract_defaults_to_first_frame_page(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        first = PillowImage.new("RGB", (8, 6), "red")
        second = PillowImage.new("RGB", (4, 3), "blue")
        first.save(
            evidence / "animation.tiff",
            save_all=True,
            append_images=[second],
        )
        first.close()
        second.close()

        result = extract_artifact(
            "animation.tiff",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert [frame["frame"] for frame in result["frame_data"]] == [1]
        assert result["omitted"]["frames"] == 1
        assert result["truncated"] is True
        assert result["next_cursor"] == "frame:2"

    def test_image_extract_reports_each_selected_metadata_unit(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        PillowImage.new("RGB", (8, 6), "red").save(
            evidence / "preview.png"
        )

        result = extract_artifact(
            "preview.png",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["metadata:format", "metadata:width"],
        )

        assert result["metadata"] == {"format": "PNG", "width": 8}
        assert result["units_inspected"] == [
            "metadata:format",
            "metadata:width",
        ]
        assert result["omitted"]["metadata"] == 4

    def test_image_render_normalizes_selected_frame_to_png(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        first = PillowImage.new("RGBA", (8, 6), (255, 0, 0, 128))
        second = PillowImage.new("RGBA", (4, 3), (0, 0, 255, 64))
        first.save(
            evidence / "animation.tiff",
            save_all=True,
            append_images=[second],
        )
        first.close()
        second.close()

        outputs = render_artifact(
            "animation.tiff",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["frame:2"],
            max_pixels=12,
        )

        assert len(outputs) == 1
        output = Path(outputs[0])
        assert output.is_file()
        assert output.parent.parent == scratch
        with PillowImage.open(output) as rendered:
            assert rendered.format == "PNG"
            assert rendered.size == (4, 3)
            assert rendered.mode == "RGBA"
            assert rendered.getpixel((0, 0)) == (0, 0, 255, 64)
        assert sorted(path.name for path in evidence.iterdir()) == [
            "animation.tiff"
        ]

    def test_pdf_render_runs_pdftoppm_per_selected_page_in_order(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        commands = []

        def fake_run(command, **kwargs):
            commands.append((command, kwargs))
            if command[0] == "pdfinfo":
                return subprocess.CompletedProcess(
                    command,
                    0,
                    "Pages: 2\n",
                    "",
                )
            page = int(command[2])
            size = (3, 4) if page == 1 else (5, 2)
            PillowImage.new("RGB", size, (page, 0, 0)).save(
                f"{command[-1]}.png"
            )
            return subprocess.CompletedProcess(command, 0, "", "")

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)

        outputs = render_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["page:2", "page:1"],
            max_pixels=22,
        )

        ppm_commands = [
            command for command, _kwargs in commands
            if command[0] == "pdftoppm"
        ]
        assert [command[2] for command in ppm_commands] == ["1", "2"]
        assert [
            command[:8] for command in ppm_commands
        ] == [
            [
                "pdftoppm",
                "-f",
                "1",
                "-singlefile",
                "-png",
                "-r",
                "144",
                ppm_commands[0][7],
            ],
            [
                "pdftoppm",
                "-f",
                "2",
                "-singlefile",
                "-png",
                "-r",
                "144",
                ppm_commands[1][7],
            ],
        ]
        assert [Path(output).name.endswith(name) for output, name in zip(
            outputs,
            ("page-0001.png", "page-0002.png"),
        )] == [True, True]
        sizes = []
        for output in outputs:
            with PillowImage.open(output) as rendered:
                sizes.append(rendered.size)
        assert sizes == [(3, 4), (5, 2)]

    def test_pdf_render_rejects_page_ranges_before_pdftoppm(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        commands = []

        def fake_run(command, **kwargs):
            commands.append(command)
            return subprocess.CompletedProcess(
                command,
                0,
                "Pages: 100000\n",
                "",
            )

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)

        with pytest.raises(InspectionError, match="render selector"):
            render_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                selectors=["page:1-100000"],
            )

        assert not any(command[0] == "pdftoppm" for command in commands)

    def test_pdf_render_default_is_bounded_to_first_page(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        rendered_pages = []

        def fake_run(command, **kwargs):
            if command[0] == "pdfinfo":
                stdout = "Pages: 100000\n"
            else:
                rendered_pages.append(int(command[2]))
                PillowImage.new("RGB", (1, 1)).save(
                    f"{command[-1]}.png"
                )
                stdout = ""
            return subprocess.CompletedProcess(command, 0, stdout, "")

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)

        outputs = render_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            max_pixels=1,
        )

        assert rendered_pages == [1]
        assert len(outputs) == 1

    def test_pdf_extract_truncates_page_text_with_explicit_pagination(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        commands = []

        def fake_run(command, **kwargs):
            commands.append(command)
            if command[0] == "pdfinfo":
                stdout = "Pages: 3\n"
            else:
                stdout = ("A" * 2_000) + "\f"
            return subprocess.CompletedProcess(command, 0, stdout, "")

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)

        result = extract_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            max_response_bytes=700,
            max_extract_chars=600,
        )

        assert len(result["page_text"]) == 1
        page = result["page_text"][0]
        assert page["page"] == 1
        assert 0 < len(page["text"]) < 2_000
        assert page["truncated"] is True
        assert page["omitted_characters"] == 2_000 - len(page["text"])
        assert result["units_inspected"] == ["page:1"]
        assert result["omitted"]["pages"] == 2
        assert result["omitted"]["characters"] == page["omitted_characters"]
        assert result["truncated"] is True
        assert result["next_cursor"] == "page:2"
        text_command = next(
            command for command in commands if command[0] == "pdftotext"
        )
        assert text_command[2:6] == ["-f", "1", "-l", "1"]

    def test_pdf_contains_text_streams_pages_under_total_budgets(
        self, tmp_path, monkeypatch
    ) -> None:
        from skillopt.envs.skilleval.inspectors import pdf_image

        path = tmp_path / "report.pdf"
        _write_pdf(path)
        scratch = tmp_path / "scratch"
        scratch.mkdir()
        commands = []

        def fake_run(command, **kwargs):
            commands.append(command)
            if command[0] == "pdfinfo":
                stdout = "Pages: 3\n"
            else:
                page = int(command[3])
                stdout = (
                    "needle is here\f"
                    if page == 3
                    else f"page {page}\f"
                )
            return subprocess.CompletedProcess(command, 0, stdout, "")

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)
        inspector = pdf_image.PdfInspector()

        assert inspector.contains_text(
            str(path),
            str(scratch),
            "needle",
            max_pages=3,
            max_bytes=100,
            timeout=10,
        ) is True
        assert [
            command[3]
            for command in commands
            if command[0] == "pdftotext"
        ] == ["1", "2", "3"]

        with pytest.raises(InspectionError, match="page budget"):
            inspector.contains_text(
                str(path),
                str(scratch),
                "missing",
                max_pages=2,
                max_bytes=100,
                timeout=10,
            )

    def test_pdf_contains_text_timeout_includes_pdfinfo(
        self, tmp_path, monkeypatch
    ) -> None:
        from skillopt.envs.skilleval.inspectors import pdf_image

        path = tmp_path / "report.pdf"
        _write_pdf(path)
        scratch = tmp_path / "scratch"
        scratch.mkdir()
        clock = [0.0]
        commands = []

        def fake_run(command, **kwargs):
            commands.append(command)
            if command[0] == "pdfinfo":
                clock[0] = 6.0
                stdout = "Pages: 1\n"
            else:
                stdout = "missing\f"
            return subprocess.CompletedProcess(command, 0, stdout, "")

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)
        monkeypatch.setattr(
            pdf_image.time,
            "monotonic",
            lambda: clock[0],
        )

        with pytest.raises(InspectionError, match="time budget"):
            pdf_image.PdfInspector().contains_text(
                str(path),
                str(scratch),
                "needle",
                max_pages=1,
                max_bytes=100,
                timeout=5,
            )

        assert [command[0] for command in commands] == ["pdfinfo"]

    def test_pdf_nonzero_tool_result_fails_closed(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        monkeypatch.setattr(
            pdf_image,
            "safe_run",
            lambda command, **kwargs: subprocess.CompletedProcess(
                command,
                1,
                "Pages: 1\n",
                "broken PDF",
            ),
        )

        with pytest.raises(InspectionError, match="exited 1"):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    @pytest.mark.parametrize(
        ("limit_name", "limit_value", "match"),
        [
            ("MAX_IMAGE_PIXELS", 399, "pixel limit"),
            ("MAX_IMAGE_DECODED_BYTES", 1_599, "decoded-byte limit"),
        ],
    )
    def test_image_explicit_decode_limits_are_controlled(
        self,
        tmp_path,
        monkeypatch,
        limit_name,
        limit_value,
        match,
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        PillowImage.new("RGBA", (20, 20), (1, 2, 3, 4)).save(
            evidence / "large.png"
        )
        from skillopt.envs.skilleval.inspectors import pdf_image

        monkeypatch.setattr(pdf_image, limit_name, limit_value)

        with pytest.raises(InspectionError, match=match):
            inspect_artifact(
                "large.png",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_image_frame_limit_and_pillow_bomb_are_controlled(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        first = PillowImage.new("RGB", (5, 5), (1, 2, 3))
        second = PillowImage.new("RGB", (5, 5), (4, 5, 6))
        first.save(
            evidence / "many.tiff",
            save_all=True,
            append_images=[second],
        )
        first.close()
        second.close()
        from skillopt.envs.skilleval.inspectors import pdf_image

        monkeypatch.setattr(pdf_image, "MAX_IMAGE_FRAMES", 1)
        with pytest.raises(InspectionError, match="frame count"):
            inspect_artifact(
                "many.tiff",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

        monkeypatch.setattr(pdf_image, "MAX_IMAGE_FRAMES", 256)
        monkeypatch.setattr(PillowImage, "MAX_IMAGE_PIXELS", 10)
        with pytest.raises(InspectionError, match="decoded safely"):
            inspect_artifact(
                "many.tiff",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_image_decoded_limit_accounts_for_sample_width(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        PillowImage.new("F", (10, 10), 1.0).save(
            evidence / "float.tiff"
        )
        from skillopt.envs.skilleval.inspectors import pdf_image

        monkeypatch.setattr(
            pdf_image,
            "MAX_IMAGE_DECODED_BYTES",
            399,
        )

        with pytest.raises(InspectionError, match="decoded-byte limit"):
            inspect_artifact(
                "float.tiff",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_image_truncated_input_and_invalid_selectors_fail_closed(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        valid = tmp_path / "valid.png"
        PillowImage.new("RGB", (10, 10), (1, 2, 3)).save(valid)
        truncated = evidence / "truncated.png"
        payload = valid.read_bytes()
        truncated.write_bytes(payload[: len(payload) // 2])
        from skillopt.envs.skilleval.inspectors.pdf_image import ImageInspector

        with pytest.raises(InspectionError):
            ImageInspector().inspect(
                str(truncated),
                str(scratch),
                response_budget=ResponseBudget(),
            )

        os.replace(valid, evidence / "valid.png")
        with pytest.raises(InspectionError, match="image extract selector"):
            extract_artifact(
                "valid.png",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                selectors=["frame:0"],
            )
        with pytest.raises(InspectionError, match="image render selector"):
            render_artifact(
                "valid.png",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                selectors=["frame:0"],
            )

    @pytest.mark.parametrize(
        ("image_format", "suffix"),
        [("PNG", "png"), ("GIF", "gif")],
    )
    def test_truncated_image_never_leaks_pillow_format_errors(
        self, tmp_path, image_format, suffix
    ) -> None:
        from skillopt.envs.skilleval.inspectors.pdf_image import ImageInspector

        complete = tmp_path / f"complete.{suffix}"
        PillowImage.new("RGBA", (12, 9), (1, 2, 3, 4)).save(
            complete,
            format=image_format,
        )
        truncated = tmp_path / f"truncated.{suffix}"
        truncated.write_bytes(complete.read_bytes()[:-5])

        with pytest.raises(InspectionError, match="decoded safely"):
            ImageInspector().inspect(
                str(truncated),
                str(tmp_path),
                response_budget=ResponseBudget(),
            )

    def test_image_render_applies_exif_orientation_and_strips_icc(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        image = PillowImage.new("RGB", (3, 2), (10, 20, 30))
        exif = image.getexif()
        exif[274] = 6
        image.save(
            evidence / "oriented.jpg",
            exif=exif,
            icc_profile=b"bounded-test-profile",
        )
        image.close()

        metadata = extract_artifact(
            "oriented.jpg",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=[
                "metadata:orientation",
                "metadata:icc_profile_bytes",
            ],
        )
        outputs = render_artifact(
            "oriented.jpg",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            max_pixels=6,
        )

        assert metadata["metadata"] == {
            "orientation": 6,
            "icc_profile_bytes": len(b"bounded-test-profile"),
        }
        with PillowImage.open(outputs[0]) as rendered:
            assert rendered.size == (2, 3)
            assert rendered.mode == "RGB"
            assert "icc_profile" not in rendered.info
            assert rendered.getexif().get(274) is None

    @pytest.mark.parametrize("kind", ["image", "pdf"])
    def test_render_pixel_overflow_rolls_back_transaction(
        self, tmp_path, monkeypatch, kind
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        if kind == "image":
            PillowImage.new("RGB", (10, 10), (1, 2, 3)).save(
                evidence / "large.png"
            )
            filename = "large.png"
        else:
            _write_pdf(evidence / "large.pdf")
            filename = "large.pdf"
            from skillopt.envs.skilleval.inspectors import pdf_image

            def fake_run(command, **kwargs):
                if command[0] == "pdfinfo":
                    stdout = "Pages: 1\n"
                else:
                    PillowImage.new("RGB", (10, 10)).save(
                        f"{command[-1]}.png"
                    )
                    stdout = ""
                return subprocess.CompletedProcess(command, 0, stdout, "")

            monkeypatch.setattr(pdf_image, "safe_run", fake_run)

        with pytest.raises(InspectionError, match="pixel budget"):
            render_artifact(
                filename,
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                max_pixels=99,
            )

        assert list(scratch.iterdir()) == []
        assert [path.name for path in evidence.iterdir()] == [filename]

    def test_pdf_render_checks_png_budget_before_verify_or_load(
        self, tmp_path, monkeypatch
    ) -> None:
        from PIL import PngImagePlugin
        from skillopt.envs.skilleval.inspectors import pdf_image

        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "large.pdf")
        verify_calls = []
        load_calls = []
        real_verify = PngImagePlugin.PngImageFile.verify

        def recording_verify(image):
            verify_calls.append(image.size)
            return real_verify(image)

        def forbidden_load(image, *args, **kwargs):
            load_calls.append(image.size)
            raise AssertionError("over-budget PNG must not be decoded")

        def fake_run(command, **kwargs):
            if command[0] == "pdfinfo":
                stdout = "Pages: 1\n"
            else:
                image = PillowImage.new("RGB", (100, 100))
                image.save(f"{command[-1]}.png")
                image.close()
                stdout = ""
            return subprocess.CompletedProcess(command, 0, stdout, "")

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)
        monkeypatch.setattr(
            PngImagePlugin.PngImageFile,
            "verify",
            recording_verify,
        )
        monkeypatch.setattr(
            PngImagePlugin.PngImageFile,
            "load",
            forbidden_load,
        )

        with pytest.raises(InspectionError, match="pixel budget"):
            render_artifact(
                "large.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                max_pixels=99,
            )

        assert verify_calls == []
        assert load_calls == []
        assert list(scratch.iterdir()) == []

    def test_pdf_tool_receives_live_stable_evidence_descriptor(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf", b"stable")
        from skillopt.envs.skilleval.inspectors import pdf_image

        observed = {}

        def fake_run(command, **kwargs):
            path = command[-1]
            descriptor = int(path.removeprefix("/proc/self/fd/"))
            observed["path"] = path
            observed["payload"] = os.pread(descriptor, 100, 0)
            observed["passed"] = descriptor in kwargs["pass_fds"]
            return subprocess.CompletedProcess(
                command,
                0,
                "Pages: 1\n",
                "",
            )

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)

        inspect_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert observed["path"].startswith("/proc/self/fd/")
        assert observed["payload"].startswith(b"%PDF-1.4\nstable")
        assert observed["passed"] is True

    def test_pdf_image_inspection_does_not_leak_file_descriptors(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        PillowImage.new("RGBA", (4, 3), (1, 2, 3, 4)).save(
            evidence / "preview.png"
        )
        from skillopt.envs.skilleval.inspectors import pdf_image

        monkeypatch.setattr(
            pdf_image,
            "safe_run",
            lambda command, **kwargs: subprocess.CompletedProcess(
                command,
                0,
                "Pages: 1\n",
                "",
            ),
        )
        before = set(os.listdir("/proc/self/fd"))

        for _ in range(5):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )
            inspect_artifact(
                "preview.png",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

        assert set(os.listdir("/proc/self/fd")) == before

    @pytest.mark.parametrize(
        ("stdout", "match"),
        [
            ("Title: Missing pages\n", "Pages"),
            ("Pages: 0\n", "positive integer"),
            ("Pages: 100001\n", "exceeds maximum"),
            ("Pages: 1\nmalformed\n", "malformed metadata"),
        ],
    )
    def test_pdfinfo_malformed_page_metadata_is_rejected(
        self, tmp_path, monkeypatch, stdout, match
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        monkeypatch.setattr(
            pdf_image,
            "safe_run",
            lambda command, **kwargs: subprocess.CompletedProcess(
                command,
                0,
                stdout,
                "",
            ),
        )

        with pytest.raises(InspectionError, match=match):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_pdf_inspect_marks_truncated_metadata_values(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        monkeypatch.setattr(
            pdf_image,
            "safe_run",
            lambda command, **kwargs: subprocess.CompletedProcess(
                command,
                0,
                f"Pages: 1\nTitle: {'x' * 3_000}\n",
                "",
            ),
        )

        result = inspect_artifact(
            "report.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )

        assert len(result["metadata"]["Title"]) == 2_048
        assert result["metadata_truncated"] == 1
        assert result["metadata_omitted"] == 0

    @pytest.mark.parametrize(
        ("failure", "match"),
        [
            (InspectionError("inspector command timed out"), "timed out"),
            (FileNotFoundError("pdfinfo"), "PDF tool failed"),
        ],
    )
    def test_pdf_missing_or_timed_out_tool_is_controlled(
        self, tmp_path, monkeypatch, failure, match
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        def fail_tool(command, **kwargs):
            raise failure

        monkeypatch.setattr(pdf_image, "safe_run", fail_tool)

        with pytest.raises(InspectionError, match=match):
            inspect_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
            )

    def test_pdf_extract_rejects_bad_pages_and_control_text(
        self, tmp_path, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        def fake_run(command, **kwargs):
            stdout = (
                "Pages: 2\n"
                if command[0] == "pdfinfo"
                else "unsafe\x00text\f"
            )
            return subprocess.CompletedProcess(command, 0, stdout, "")

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)

        for selector in ("page:0", "page:2-1", "page:3"):
            with pytest.raises(InspectionError):
                extract_artifact(
                    "report.pdf",
                    evidence_dir=str(evidence),
                    scratch_dir=str(scratch),
                    selectors=[selector],
                )
        with pytest.raises(InspectionError, match="control characters"):
            extract_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                selectors=["page:1"],
            )

    @pytest.mark.parametrize("control", ["\x7f", "\x85"])
    def test_pdf_extract_rejects_del_and_c1_controls(
        self, tmp_path, monkeypatch, control
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        from skillopt.envs.skilleval.inspectors import pdf_image

        def fake_run(command, **kwargs):
            stdout = (
                "Pages: 1\n"
                if command[0] == "pdfinfo"
                else f"safe{control}unsafe\f"
            )
            return subprocess.CompletedProcess(command, 0, stdout, "")

        monkeypatch.setattr(pdf_image, "safe_run", fake_run)

        with pytest.raises(InspectionError, match="control characters"):
            extract_artifact(
                "report.pdf",
                evidence_dir=str(evidence),
                scratch_dir=str(scratch),
                selectors=["page:1"],
            )

    @pytest.mark.skipif(
        any(
            shutil.which(tool) is None
            for tool in ("pdfinfo", "pdftotext", "pdftoppm")
        ),
        reason="Poppler pdfinfo/pdftotext/pdftoppm are not installed",
    )
    def test_real_poppler_pdf_inspect_extract_render_smoke(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        first = PillowImage.new("RGB", (16, 12), "white")
        second = PillowImage.new("RGB", (16, 12), "black")
        first.save(
            evidence / "two-pages.pdf",
            save_all=True,
            append_images=[second],
            resolution=72,
        )
        first.close()
        second.close()

        inspected = inspect_artifact(
            "two-pages.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
        )
        extracted = extract_artifact(
            "two-pages.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["page:1"],
        )
        outputs = render_artifact(
            "two-pages.pdf",
            evidence_dir=str(evidence),
            scratch_dir=str(scratch),
            selectors=["page:2"],
            max_pixels=10_000,
        )

        assert inspected["pages"] == 2
        assert extracted["page_text"][0]["page"] == 1
        assert len(outputs) == 1
        with PillowImage.open(outputs[0]) as rendered:
            assert rendered.format == "PNG"
            assert rendered.width > 0
            assert rendered.height > 0


class TestArtifactCtl:
    def _run(self, capsys, argv):
        code = artifactctl_main(argv)
        captured = capsys.readouterr()
        lines = captured.out.splitlines()
        assert len(lines) == 1
        assert captured.err == ""
        return code, json.loads(lines[0])

    def test_inventory_command(self, tmp_path, capsys) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")

        code, payload = self._run(
            capsys,
            [
                "inventory",
                "--evidence",
                str(evidence),
                "--scratch",
                str(scratch),
            ],
        )

        assert code == 0
        assert payload["status"] == "ok"
        assert payload["result"][0]["path"] == "report.pdf"

    @pytest.mark.parametrize(
        ("command", "patched_name", "result"),
        [
            ("inspect", "inspect_artifact", {"metadata": "ok"}),
            ("render", "render_artifact", ["/trusted/scratch/render.png"]),
            ("extract", "extract_artifact", {"text": "ok"}),
        ],
    )
    def test_artifact_commands_emit_one_json_object(
        self,
        tmp_path,
        capsys,
        monkeypatch,
        command,
        patched_name,
        result,
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        cli_module = sys.modules[
            "skillopt.envs.skilleval.inspectors.__main__"
        ]
        monkeypatch.setattr(cli_module, patched_name, lambda *args, **kwargs: result)
        argv = [
            command,
            "report.pdf",
            "--evidence",
            str(evidence),
            "--scratch",
            str(scratch),
            "--selector",
            "page:1",
        ]
        if command == "render":
            argv.extend(["--max-pixels", "1000"])

        code, payload = self._run(capsys, argv)

        assert code == 0
        assert payload == {"status": "ok", "result": result}

    def test_operation_error_is_one_bounded_json_object(
        self, tmp_path, capsys
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        (evidence / "UNTRUSTED.bin").write_bytes(b"unknown")

        code, payload = self._run(
            capsys,
            [
                "inspect",
                "UNTRUSTED.bin",
                "--evidence",
                str(evidence),
                "--scratch",
                str(scratch),
            ],
        )

        assert code == 2
        assert payload["status"] == "error"
        assert payload["error"].startswith("InspectionError:")
        assert len(json.dumps(payload)) < 2_000

    def test_selected_response_budget_bounds_compact_error(
        self, tmp_path, capsys, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        cli_module = sys.modules[
            "skillopt.envs.skilleval.inspectors.__main__"
        ]
        monkeypatch.setattr(
            cli_module,
            "inspect_artifact",
            lambda *args, **kwargs: {"text": "界" * 10_000},
        )

        code = artifactctl_main(
            [
                "inspect",
                "report.pdf",
                "--evidence",
                str(evidence),
                "--scratch",
                str(scratch),
                "--max-response-bytes",
                str(MIN_RESPONSE_BYTES),
            ]
        )
        captured = capsys.readouterr()

        assert code == 2
        assert captured.err == ""
        assert len(captured.out.splitlines()) == 1
        assert len(captured.out.rstrip("\n").encode("utf-8")) <= (
            MIN_RESPONSE_BYTES
        )
        assert json.loads(captured.out)["status"] == "error"

    def test_selected_response_budget_bounds_success(
        self, tmp_path, capsys, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        cli_module = sys.modules[
            "skillopt.envs.skilleval.inspectors.__main__"
        ]
        monkeypatch.setattr(
            cli_module,
            "inspect_artifact",
            lambda *args, **kwargs: {"text": "ok"},
        )

        code = artifactctl_main(
            [
                "inspect",
                "report.pdf",
                "--evidence",
                str(evidence),
                "--scratch",
                str(scratch),
                "--max-response-bytes",
                str(MIN_RESPONSE_BYTES),
            ]
        )
        captured = capsys.readouterr()

        assert code == 0
        assert captured.err == ""
        assert len(captured.out.rstrip("\n").encode("utf-8")) <= (
            MIN_RESPONSE_BYTES
        )
        assert json.loads(captured.out)["status"] == "ok"

    def test_response_budget_below_minimum_rejects_before_dispatch(
        self, tmp_path, capsys, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        called = False
        cli_module = sys.modules[
            "skillopt.envs.skilleval.inspectors.__main__"
        ]

        def inspect_should_not_run(*args, **kwargs):
            nonlocal called
            called = True
            return {}

        monkeypatch.setattr(
            cli_module,
            "inspect_artifact",
            inspect_should_not_run,
        )
        code, payload = self._run(
            capsys,
            [
                "inspect",
                "report.pdf",
                "--evidence",
                str(evidence),
                "--scratch",
                str(scratch),
                "--max-response-bytes",
                str(MIN_RESPONSE_BYTES - 1),
            ],
        )

        assert code == 2
        assert payload["status"] == "error"
        assert "at least" in payload["error"]
        assert called is False

    def test_path_limit_is_checked_before_root_filesystem_access(
        self, tmp_path, capsys
    ) -> None:
        code, payload = self._run(
            capsys,
            [
                "inspect",
                "a" * (MAX_LOGICAL_COMPONENT_BYTES + 1),
                "--evidence",
                str(tmp_path / "missing-evidence"),
                "--scratch",
                str(tmp_path / "missing-scratch"),
            ],
        )

        assert code == 2
        assert payload["status"] == "error"
        assert "component" in payload["error"]
        assert "does not exist" not in payload["error"]

    def test_parse_error_is_json_only(self, capsys) -> None:
        code, payload = self._run(capsys, ["inspect"])
        assert code == 2
        assert payload["status"] == "error"
        assert "required" in payload["error"]

    @pytest.mark.parametrize("argv", [["--help"], ["inspect", "--help"]])
    def test_help_is_one_success_json_object(self, capsys, argv) -> None:
        code, payload = self._run(capsys, argv)
        assert code == 0
        assert payload["status"] == "ok"
        assert "usage:" in payload["result"]["help"]

    def test_selected_response_budget_bounds_help(self, capsys) -> None:
        argv = [
            "render",
            "--max-response-bytes",
            str(MIN_RESPONSE_BYTES),
            "--help",
        ]

        code = artifactctl_main(argv)
        captured = capsys.readouterr()

        assert code == 0
        assert captured.err == ""
        assert len(captured.out.splitlines()) == 1
        assert len(captured.out.rstrip("\n").encode("utf-8")) <= (
            MIN_RESPONSE_BYTES
        )
        assert json.loads(captured.out)["status"] == "ok"

    def test_selected_response_budget_bounds_argparse_error(self, capsys) -> None:
        argv = [
            "inspect",
            "report.pdf",
            "--evidence",
            "/missing/evidence",
            "--scratch",
            "/missing/scratch",
            "--max-response-bytes",
            str(MIN_RESPONSE_BYTES),
            "--unknown=" + "x" * 2_000,
        ]

        code = artifactctl_main(argv)
        captured = capsys.readouterr()

        assert code == 2
        assert captured.err == ""
        assert len(captured.out.splitlines()) == 1
        assert len(captured.out.rstrip("\n").encode("utf-8")) <= (
            MIN_RESPONSE_BYTES
        )
        assert json.loads(captured.out)["status"] == "error"

    @pytest.mark.parametrize(
        "argv",
        [
            [],
            ["unknown-command"],
            ["inventory", "--unknown-option"],
        ],
    )
    def test_all_argparse_errors_are_one_error_json_object(
        self, capsys, argv
    ) -> None:
        code, payload = self._run(capsys, argv)
        assert code == 2
        assert payload["status"] == "error"
        assert payload["error"].startswith("InspectionError:")

    @pytest.mark.parametrize(
        "argv",
        [
            ["render", "x.pdf", "--max-pixels", "0"],
            ["render", "x.pdf", "--max-pixels", str(MAX_RENDER_PIXELS + 1)],
            ["inspect", "x.pdf", "--max-response-bytes", "0"],
            [
                "inspect",
                "x.pdf",
                "--max-response-bytes",
                str(MAX_RESPONSE_BYTES + 1),
            ],
        ],
    )
    def test_cli_rejects_out_of_range_budgets(self, tmp_path, capsys, argv) -> None:
        evidence, scratch = _roots(tmp_path)
        argv.extend(
            ["--evidence", str(evidence), "--scratch", str(scratch)]
        )
        code, payload = self._run(capsys, argv)
        assert code == 2
        assert payload["status"] == "error"

    def test_cli_passes_scratch_budget(self, tmp_path, capsys, monkeypatch) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        cli_module = sys.modules[
            "skillopt.envs.skilleval.inspectors.__main__"
        ]
        observed = {}

        def fake_inspect(*args, **kwargs):
            observed.update(kwargs)
            return {"ok": True}

        monkeypatch.setattr(cli_module, "inspect_artifact", fake_inspect)
        code, payload = self._run(
            capsys,
            [
                "inspect",
                "report.pdf",
                "--evidence",
                str(evidence),
                "--scratch",
                str(scratch),
                "--max-scratch-bytes",
                "1234",
                "--max-scratch-entries",
                "12",
                "--max-scratch-depth",
                "4",
            ],
        )

        assert code == 0
        assert payload["status"] == "ok"
        assert observed["max_scratch_bytes"] == 1234
        assert observed["max_scratch_entries"] == 12
        assert observed["max_scratch_depth"] == 4


def _assert_untrusted_location(value, needles, path=()):
    if isinstance(value, dict):
        for key, child in value.items():
            _assert_untrusted_location(child, needles, (*path, key))
        return
    if isinstance(value, list):
        for index, child in enumerate(value):
            _assert_untrusted_location(child, needles, (*path, index))
        return
    if not isinstance(value, str):
        return
    for needle in needles:
        if needle in value:
            assert "untrusted_evidence" in path
            assert any(part in path for part in ("result", "error"))


def _structured_payload(result):
    assert result.structuredContent is not None
    payload = result.structuredContent
    assert set(payload) == {"untrusted_evidence"}
    return payload


def _serialized_mcp_result_bytes(result) -> int:
    return len(
        result.model_dump_json(
            by_alias=True,
            exclude_none=True,
        ).encode("utf-8")
    )


class TestArtifactMcp:
    @pytest.mark.anyio
    async def test_real_stdio_initialize_list_call_and_image(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        server_script = tmp_path / "artifact_stdio_server.py"
        server_script.write_text(
            """
import os
import sys
import types
from pathlib import Path

from PIL import Image as PillowImage

class FakePdfInspector:
    def inspect(self, path, scratch_dir, *, response_budget):
        return {"filename": Path(path).name, "metadata": "STDIO_META"}

    def render(self, path, scratch_dir, selectors, budget):
        output = Path(scratch_dir) / "stdio-render.png"
        PillowImage.new("RGB", (2, 3), color=(10, 20, 30)).save(output)
        return [str(output)]

    def extract(self, path, scratch_dir, selectors, *, response_budget):
        return {"text": "STDIO_TEXT"}


module = types.ModuleType("skillopt.envs.skilleval.inspectors.pdf_image")
module.PdfInspector = FakePdfInspector
module.ImageInspector = FakePdfInspector
sys.modules[module.__name__] = module

from skillopt.envs.skilleval.artifact_mcp import create_server

server = create_server(
    os.environ["TEST_ARTIFACT_EVIDENCE"],
    os.environ["TEST_ARTIFACT_SCRATCH"],
)
server.run(transport="stdio")
""".lstrip(),
            encoding="utf-8",
        )
        environment = os.environ.copy()
        repo_root = str(Path(__file__).resolve().parents[1])
        environment.update(
            {
                "TEST_ARTIFACT_EVIDENCE": str(evidence),
                "TEST_ARTIFACT_SCRATCH": str(scratch),
                "PYTHONPATH": os.pathsep.join(
                    filter(
                        None,
                        (repo_root, environment.get("PYTHONPATH")),
                    )
                ),
            }
        )
        parameters = StdioServerParameters(
            command=sys.executable,
            args=[str(server_script)],
            env=environment,
            cwd=repo_root,
        )
        stderr_path = tmp_path / "stdio-server.stderr"
        with stderr_path.open("w+", encoding="utf-8") as stderr:
            async with stdio_client(
                parameters,
                errlog=stderr,
            ) as (read_stream, write_stream):
                async with ClientSession(
                    read_stream,
                    write_stream,
                ) as client:
                    initialized = await client.initialize()
                    tools = await client.list_tools()
                    inventory = await client.call_tool(
                        "artifact_inventory",
                        {},
                    )
                    rendered = await client.call_tool(
                        "artifact_render",
                        {"path": "report.pdf", "max_pixels": 100},
                    )
            stderr.seek(0)
            server_stderr = stderr.read()

        assert initialized.serverInfo.name == "skillopt-artifact"
        assert sorted(tool.name for tool in tools.tools) == [
            "artifact_extract",
            "artifact_inspect",
            "artifact_inventory",
            "artifact_render",
        ]
        assert _structured_payload(inventory)["untrusted_evidence"][
            "status"
        ] == "ok"
        assert _structured_payload(rendered)["untrusted_evidence"][
            "status"
        ] == "ok"
        images = [
            content
            for content in rendered.content
            if isinstance(content, ImageContent)
        ]
        assert len(images) == 1
        assert images[0].mimeType == "image/png"
        assert "Traceback" not in server_stderr
        assert "report.pdf" not in server_stderr
        assert "STDIO_META" not in server_stderr

    @pytest.mark.anyio
    async def test_lists_exactly_four_tools_and_no_resources_or_prompts(
        self, tmp_path
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        server = artifact_mcp.create_server(str(evidence), str(scratch))

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            tools = await client.list_tools()
            resources = await client.list_resources()
            prompts = await client.list_prompts()

        assert sorted(tool.name for tool in tools.tools) == [
            "artifact_extract",
            "artifact_inspect",
            "artifact_inventory",
            "artifact_render",
        ]
        assert resources.resources == []
        assert prompts.prompts == []
        assert client.get_server_capabilities().tools is not None
        inspect_tool = next(
            tool for tool in tools.tools if tool.name == "artifact_inspect"
        )
        assert "POSIX-relative" in inspect_tool.description
        assert str(MAX_LOGICAL_PATH_BYTES) in inspect_tool.description
        assert str(MAX_LOGICAL_COMPONENTS) in inspect_tool.description
        assert str(MAX_LOGICAL_COMPONENT_BYTES) in inspect_tool.description
        path_schema = inspect_tool.inputSchema["properties"]["path"]
        assert path_schema["description"].startswith(
            "Normalized POSIX-relative"
        )
        for tool_name in (
            "artifact_inspect",
            "artifact_render",
            "artifact_extract",
        ):
            tool = next(tool for tool in tools.tools if tool.name == tool_name)
            assert tool.inputSchema["required"] == ["path"]
        for tool in tools.tools:
            assert tool.inputSchema["additionalProperties"] is False

    @pytest.mark.anyio
    async def test_calls_representative_tools_and_envelopes_artifact_strings(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        artifact_name = "UNTRUSTED_FILENAME.pdf"
        _write_pdf(evidence / artifact_name)
        server = artifact_mcp.create_server(str(evidence), str(scratch))
        needles = {
            artifact_name,
            "UNTRUSTED_META",
            "UNTRUSTED_TEXT",
        }

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            results = [
                await client.call_tool("artifact_inventory", {}),
                await client.call_tool(
                    "artifact_inspect",
                    {"path": artifact_name},
                ),
                await client.call_tool(
                    "artifact_extract",
                    {"path": artifact_name, "selectors": ["page:1"]},
                ),
            ]

        for result in results:
            payload = _structured_payload(result)
            _assert_untrusted_location(payload, needles)
            for content in result.content:
                if isinstance(content, TextContent):
                    text_payload = json.loads(content.text)
                    _assert_untrusted_location(text_payload, needles)

    @pytest.mark.anyio
    async def test_unsafe_path_and_artifact_error_are_enveloped(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        server = artifact_mcp.create_server(str(evidence), str(scratch))

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            unsafe = await client.call_tool(
                "artifact_inspect",
                {"path": "../UNTRUSTED_ESCAPE.pdf"},
            )
            fake_pdf_inspector.fail_message = "UNTRUSTED_ERROR"
            failed = await client.call_tool(
                "artifact_inspect",
                {"path": "report.pdf"},
            )

        for result, needle in (
            (unsafe, "UNTRUSTED_ESCAPE"),
            (failed, "UNTRUSTED_ERROR"),
        ):
            payload = _structured_payload(result)
            _assert_untrusted_location(payload, {needle})
            envelope = payload["untrusted_evidence"]
            assert envelope["status"] == "error"
            assert "error" in envelope
            assert result.isError is True

    @pytest.mark.anyio
    async def test_server_maxima_reject_oversized_tool_budget(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        server = artifact_mcp.create_server(
            str(evidence),
            str(scratch),
            max_render_pixels=100,
            max_response_bytes=1_000,
            max_extract_chars=500,
        )

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            result = await client.call_tool(
                "artifact_render",
                {"path": "report.pdf", "max_pixels": 101},
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "error"
        assert "maximum" in envelope["error"]
        assert result.isError is True

    @pytest.mark.anyio
    async def test_response_budget_uses_minimum_of_request_and_server(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        server_max = MIN_RESPONSE_BYTES + 128
        server = artifact_mcp.create_server(
            str(evidence),
            str(scratch),
            max_response_bytes=server_max,
        )

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            result = await client.call_tool(
                "artifact_inspect",
                {
                    "path": "report.pdf",
                    "max_response_bytes": MAX_RESPONSE_BYTES,
                },
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "ok"
        text = next(
            content.text
            for content in result.content
            if isinstance(content, TextContent)
        )
        assert len(text.encode("utf-8")) <= server_max

    @pytest.mark.anyio
    async def test_selected_response_budget_bounds_mcp_error(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        fake_pdf_inspector.inspect_value = "界" * 10_000
        server = artifact_mcp.create_server(str(evidence), str(scratch))

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            result = await client.call_tool(
                "artifact_inspect",
                {
                    "path": "report.pdf",
                    "max_response_bytes": MIN_RESPONSE_BYTES,
                },
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "error"
        assert result.isError is True
        text = next(
            content.text
            for content in result.content
            if isinstance(content, TextContent)
        )
        assert len(text.encode("utf-8")) <= MIN_RESPONSE_BYTES
        assert json.loads(text) == result.structuredContent

    @pytest.mark.anyio
    @pytest.mark.parametrize(
        "tool_name",
        ["artifact_render", "artifact_extract"],
    )
    async def test_argument_validation_errors_are_bounded_and_enveloped(
        self, tmp_path, fake_pdf_inspector, tool_name
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        server = artifact_mcp.create_server(str(evidence), str(scratch))
        invalid_selectors = [
            {"UNTRUSTED_SELECTOR": index}
            for index in range(256)
        ]

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            assert client.get_server_capabilities().tools is not None
            tools = await client.list_tools()
            assert tool_name in {tool.name for tool in tools.tools}
            result = await client.call_tool(
                tool_name,
                {
                    "path": "report.pdf",
                    "selectors": invalid_selectors,
                    "max_response_bytes": MIN_RESPONSE_BYTES,
                },
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "error"
        assert "selector" in envelope["error"]
        assert result.isError is True
        text = next(
            content.text
            for content in result.content
            if isinstance(content, TextContent)
        )
        assert _serialized_mcp_result_bytes(result) <= MIN_RESPONSE_BYTES
        assert json.loads(text) == result.structuredContent
        assert fake_pdf_inspector.calls == []

    @pytest.mark.anyio
    @pytest.mark.parametrize(
        "tool_name",
        ["artifact_inspect", "artifact_render", "artifact_extract"],
    )
    async def test_missing_path_errors_are_bounded_and_enveloped(
        self, tmp_path, fake_pdf_inspector, tool_name
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        server = artifact_mcp.create_server(str(evidence), str(scratch))

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            tools = await client.list_tools()
            tool = next(tool for tool in tools.tools if tool.name == tool_name)
            assert tool.inputSchema["required"] == ["path"]
            result = await client.call_tool(
                tool_name,
                {"max_response_bytes": MIN_RESPONSE_BYTES},
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "error"
        assert "path" in envelope["error"]
        assert result.isError is True
        text = next(
            content.text
            for content in result.content
            if isinstance(content, TextContent)
        )
        assert _serialized_mcp_result_bytes(result) <= MIN_RESPONSE_BYTES
        assert json.loads(text) == result.structuredContent
        assert fake_pdf_inspector.calls == []

    @pytest.mark.anyio
    async def test_unknown_argument_is_bounded_enveloped_and_not_dispatched(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        server = artifact_mcp.create_server(str(evidence), str(scratch))
        unknown_name = "UNTRUSTED_UNKNOWN_" + "x" * 1_000

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            result = await client.call_tool(
                "artifact_inspect",
                {
                    "path": "report.pdf",
                    "max_response_bytes": MIN_RESPONSE_BYTES,
                    unknown_name: "UNTRUSTED_VALUE" * 1_000,
                },
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "error"
        assert "unknown" in envelope["error"].lower()
        assert result.isError is True
        assert _serialized_mcp_result_bytes(result) <= MIN_RESPONSE_BYTES
        assert fake_pdf_inspector.calls == []

    @pytest.mark.anyio
    async def test_scratch_budget_is_exposed_and_enforced(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        fake_pdf_inspector.unreturned_scratch_bytes = 128
        server = artifact_mcp.create_server(
            str(evidence),
            str(scratch),
            max_scratch_bytes=100,
            max_scratch_entries=10,
            max_scratch_depth=4,
        )

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            tools = await client.list_tools()
            inspect_tool = next(
                tool for tool in tools.tools
                if tool.name == "artifact_inspect"
            )
            assert "max_scratch_bytes" in inspect_tool.inputSchema["properties"]
            assert "max_scratch_entries" in inspect_tool.inputSchema["properties"]
            assert "max_scratch_depth" in inspect_tool.inputSchema["properties"]
            result = await client.call_tool(
                "artifact_inspect",
                {
                    "path": "report.pdf",
                    "max_response_bytes": 4_096,
                    "max_scratch_bytes": 100,
                    "max_scratch_entries": 10,
                    "max_scratch_depth": 4,
                },
            )
            inventory_result = await client.call_tool(
                "artifact_inventory",
                {
                    "max_response_bytes": 4_096,
                    "max_scratch_bytes": 100,
                    "max_scratch_entries": 10,
                    "max_scratch_depth": 4,
                },
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "error"
        assert "scratch" in envelope["error"]
        assert result.isError is True
        assert _serialized_mcp_result_bytes(result) <= 4_096
        assert list(scratch.iterdir()) == []
        inventory_envelope = _structured_payload(inventory_result)[
            "untrusted_evidence"
        ]
        assert inventory_envelope["status"] == "ok"
        assert inventory_result.isError is not True
        assert (
            _serialized_mcp_result_bytes(inventory_result)
            <= 4_096
        )

    @pytest.mark.anyio
    async def test_mcp_response_budget_below_minimum_skips_inspector(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        server = artifact_mcp.create_server(str(evidence), str(scratch))

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            result = await client.call_tool(
                "artifact_inspect",
                {
                    "path": "report.pdf",
                    "max_response_bytes": MIN_RESPONSE_BYTES - 1,
                },
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "error"
        assert "at least" in envelope["error"]
        assert result.isError is True
        assert fake_pdf_inspector.calls == []

    @pytest.mark.anyio
    async def test_mcp_path_limit_rejects_before_inspector(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        server = artifact_mcp.create_server(str(evidence), str(scratch))

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            result = await client.call_tool(
                "artifact_inspect",
                {"path": "界" * 86},
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "error"
        assert "component" in envelope["error"]
        assert result.isError is True
        assert fake_pdf_inspector.calls == []

    @pytest.mark.anyio
    async def test_render_returns_png_image_content_without_host_path(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "UNTRUSTED_FILENAME.pdf")
        server = artifact_mcp.create_server(str(evidence), str(scratch))

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            result = await client.call_tool(
                "artifact_render",
                {
                    "path": "UNTRUSTED_FILENAME.pdf",
                    "selectors": ["page:1"],
                    "max_pixels": 100,
                },
            )

        payload = _structured_payload(result)
        envelope = payload["untrusted_evidence"]
        assert envelope["status"] == "ok"
        serialized = json.dumps(payload)
        assert str(scratch) not in serialized
        images = [
            content for content in result.content if isinstance(content, ImageContent)
        ]
        assert len(images) == 1
        assert images[0].mimeType == "image/png"
        image_metadata = envelope["result"]["images"][0]
        assert image_metadata == {
            "index": 0,
            "mime": "image/png",
            "width": 2,
            "height": 3,
            "bytes": len(base64.b64decode(images[0].data)),
        }
        assert list(scratch.iterdir()) == []

    @pytest.mark.anyio
    async def test_render_does_not_reopen_replaced_scratch_root(
        self, tmp_path, fake_pdf_inspector, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        held = tmp_path / "held-scratch"
        real_read_png = artifact_mcp._read_png
        replacement_performed = False

        def replace_root_then_read(source, scratch_dir, **kwargs):
            nonlocal replacement_performed
            scratch.rename(held)
            scratch.mkdir()
            PillowImage.new("RGB", (9, 9), color=(200, 1, 1)).save(
                scratch / "render-UNTRUSTED_META.png"
            )
            replacement_performed = True
            return real_read_png(source, scratch_dir, **kwargs)

        monkeypatch.setattr(
            artifact_mcp,
            "_read_png",
            replace_root_then_read,
        )
        server = artifact_mcp.create_server(str(evidence), str(scratch))

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            result = await client.call_tool(
                "artifact_render",
                {"path": "report.pdf", "max_pixels": 100},
            )

        assert replacement_performed is True
        envelope = _structured_payload(result)["untrusted_evidence"]
        if envelope["status"] == "ok":
            assert envelope["result"]["images"][0]["width"] == 2
            assert envelope["result"]["images"][0]["height"] == 3
        else:
            assert result.isError is True
            assert not any(
                isinstance(content, ImageContent) for content in result.content
            )

    @pytest.mark.anyio
    async def test_render_full_result_budget_rejects_before_reading_media(
        self, tmp_path, fake_pdf_inspector, monkeypatch
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        large = scratch / "large.png"
        pixels = os.urandom(128 * 128 * 3)
        PillowImage.frombytes("RGB", (128, 128), pixels).save(
            large,
            compress_level=0,
        )
        fake_pdf_inspector.render_bytes = large.read_bytes()
        read_called = False
        image_called = False

        def unexpected_read(*args, **kwargs):
            nonlocal read_called
            read_called = True
            raise AssertionError("oversized media must not be read")

        def unexpected_image(*args, **kwargs):
            nonlocal image_called
            image_called = True
            raise AssertionError("oversized media must not be encoded")

        monkeypatch.setattr(artifact_mcp, "_read_png", unexpected_read)
        monkeypatch.setattr(artifact_mcp, "Image", unexpected_image)
        server = artifact_mcp.create_server(str(evidence), str(scratch))

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            result = await client.call_tool(
                "artifact_render",
                {
                    "path": "report.pdf",
                    "max_pixels": 128 * 128,
                    "max_response_bytes": MIN_RESPONSE_BYTES,
                },
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "error"
        assert "response" in envelope["error"].lower()
        assert result.isError is True
        assert _serialized_mcp_result_bytes(result) <= MIN_RESPONSE_BYTES
        assert read_called is False
        assert image_called is False
        assert not any(
            isinstance(content, ImageContent) for content in result.content
        )

    @pytest.mark.anyio
    async def test_render_rejects_non_png_and_oversized_media(
        self, tmp_path, fake_pdf_inspector
    ) -> None:
        evidence, scratch = _roots(tmp_path)
        _write_pdf(evidence / "report.pdf")
        bad = scratch / "bad.png"
        bad.write_text("not an image", encoding="utf-8")
        fake_pdf_inspector.render_bytes = bad.read_bytes()
        server = artifact_mcp.create_server(
            str(evidence),
            str(scratch),
            max_media_bytes=10,
        )

        async with create_connected_server_and_client_session(
            server,
            raise_exceptions=True,
        ) as client:
            result = await client.call_tool(
                "artifact_render",
                {"path": "report.pdf"},
            )

        envelope = _structured_payload(result)["untrusted_evidence"]
        assert envelope["status"] == "error"
        assert result.isError is True
        assert not any(
            isinstance(content, ImageContent) for content in result.content
        )
