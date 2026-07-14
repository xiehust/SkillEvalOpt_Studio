"""Bounded spreadsheet structure, checks, and LibreOffice rendering."""
from __future__ import annotations

import datetime as dt
import json
import math
import os
import re
import shutil
import stat
import struct
import tempfile
import time
import unicodedata
import zipfile
import zlib
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import unquote_to_bytes, urlsplit
from xml.etree import ElementTree

import openpyxl
from openpyxl.utils.cell import (
    coordinate_to_tuple,
    get_column_letter,
    range_boundaries,
)
from openpyxl.worksheet.worksheet import Worksheet

from skillopt.envs.skilleval import artifacts as artifact_security

from ._scratch import current_scratch_transaction, scratch_transaction
from .base import (
    DEFAULT_SCRATCH_BYTES,
    DEFAULT_SCRATCH_DEPTH,
    DEFAULT_SCRATCH_ENTRIES,
    InspectionError,
    RenderBudget,
    ResponseBudget,
    bounded_diagnostic,
    safe_run,
    validate_json_result,
)

_OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_MAX_OOXML_ENTRIES = 10_000
_MAX_OOXML_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
_MAX_CONTENT_TYPES_BYTES = 2 * 1024 * 1024
_MAX_OOXML_LOCAL_METADATA_BYTES = 16 * 1024 * 1024
_MAX_OOXML_GRAPH_PART_BYTES = 4 * 1024 * 1024
_MAX_OOXML_XML_PART_BYTES = 32 * 1024 * 1024
_MAX_OOXML_XML_NODES = 1_000_000
_MAX_OOXML_XML_DEPTH = 256
_MAX_ROW_DIMENSIONS = 50_000
_MAX_WORKSHEET_CELLS = 250_000
_MAX_WORKSHEET_MERGES = 50_000
_MAX_MERGED_RANGE_CELLS = 100_000
_MAX_MERGED_CELLS = 250_000
_MAX_SHARED_STRING_ITEMS = 250_000
_MAX_SHARED_STRING_CHARS = 16 * 1024 * 1024
_MAX_STYLE_RECORDS = 100_000
_MAX_RELATIONSHIPS = 100_000
_MAX_DRAWING_OBJECTS = 100_000
_MAX_CHART_OBJECTS = 100_000
_MAX_COMPRESSION_RATIO = 200.0
_ZIP_STREAM_CHUNK = 1024 * 1024
_LOCAL_FILE_HEADER = struct.Struct("<4s5H3L2H")
_LOCAL_FILE_HEADER_SIGNATURE = b"PK\x03\x04"
_DATA_DESCRIPTOR_SIGNATURE = b"PK\x07\x08"
_ZIP32_DATA_DESCRIPTOR = struct.Struct("<3L")
_ZIP64_DATA_DESCRIPTOR = struct.Struct("<LQQ")
_ZIP64_EXTRA_ID = 0x0001
_ZIP32_MAX = 0xFFFFFFFF
_SUPPORTED_ZIP_FLAGS = 0x080E
_CELL_PAGE_SIZE = 128
_LIBREOFFICE_TIMEOUT_SECONDS = 120
_LIBREOFFICE_MAX_TRANSIENT_ATTEMPTS = 3
_TRANSIENT_SCRATCH_ENOENT = (
    "inspector command scratch failure: [Errno 2] No such file or directory:"
)
_TRANSIENT_SCRATCH_PROFILE_SCAN = (
    "inspector command scratch failure: "
    "scratch directory could not be scanned: "
)
_SUPPORTED_ZIP_COMPRESSION = {zipfile.ZIP_STORED, zipfile.ZIP_DEFLATED}
_XLSX_WORKBOOK_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet.main+xml"
)
_WORKSHEET_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.worksheet+xml"
)
_CHARTSHEET_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.chartsheet+xml"
)
_STYLES_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.styles+xml"
)
_SHARED_STRINGS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sharedStrings+xml"
)
_DRAWING_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.drawing+xml"
)
_CHART_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"
)
_RELATIONSHIPS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-package.relationships+xml"
)
_CONTENT_TYPES_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/content-types"
)
_SPREADSHEET_NAMESPACE = (
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
)
_PACKAGE_RELATIONSHIPS_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/relationships"
)
_SPREADSHEET_DRAWING_NAMESPACE = (
    "http://schemas.openxmlformats.org/drawingml/2006/"
    "spreadsheetDrawing"
)
_CHART_NAMESPACE = (
    "http://schemas.openxmlformats.org/drawingml/2006/chart"
)
_OFFICE_RELATIONSHIPS_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_WORKSHEET_RELATIONSHIP_TYPE = (
    f"{_OFFICE_RELATIONSHIPS_NAMESPACE}/worksheet"
)
_CHARTSHEET_RELATIONSHIP_TYPE = (
    f"{_OFFICE_RELATIONSHIPS_NAMESPACE}/chartsheet"
)
_STYLES_RELATIONSHIP_TYPE = f"{_OFFICE_RELATIONSHIPS_NAMESPACE}/styles"
_SHARED_STRINGS_RELATIONSHIP_TYPE = (
    f"{_OFFICE_RELATIONSHIPS_NAMESPACE}/sharedStrings"
)
_DRAWING_RELATIONSHIP_TYPE = f"{_OFFICE_RELATIONSHIPS_NAMESPACE}/drawing"
_CHART_RELATIONSHIP_TYPE = f"{_OFFICE_RELATIONSHIPS_NAMESPACE}/chart"
_CELL_SELECTOR_RE = re.compile(
    r"^sheet:([^:]+?)(?::page:([1-9][0-9]*))?$"
)
_PDF_SELECTOR_RE = re.compile(r"^page:[1-9][0-9]*$")
_CELL_REF_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]{0,6})$")
_WINDOWS_ROOT_RE = re.compile(r"^[A-Za-z]:")
_PROC_FD_RE = re.compile(r"^/proc/self/fd/([0-9]+)$")


def _member_name(info: zipfile.ZipInfo) -> str:
    name = getattr(info, "orig_filename", info.filename)
    if not isinstance(name, str) or name != info.filename:
        raise InspectionError("unsafe OOXML entry name")
    return name


def _validate_member_name(info: zipfile.ZipInfo) -> str:
    name = _member_name(info)
    if not artifact_security._safe_zip_member(name):
        raise InspectionError(f"unsafe OOXML entry: {bounded_diagnostic(name)!r}")
    logical_name = name[:-1] if info.is_dir() else name
    parts = logical_name.split("/")
    if (
        not logical_name
        or any(part in {"", ".", ".."} for part in parts)
        or _WINDOWS_ROOT_RE.match(parts[0])
    ):
        raise InspectionError(f"unsafe OOXML entry: {bounded_diagnostic(name)!r}")
    try:
        encoded = name.encode("utf-8")
    except UnicodeEncodeError as exc:
        raise InspectionError("unsafe OOXML entry encoding") from exc
    if len(encoded) > 4_096 or any(
        len(part.encode("utf-8")) > 255 for part in parts
    ):
        raise InspectionError("unsafe OOXML entry exceeds path limits")
    return name


def _validate_member_type(info: zipfile.ZipInfo) -> None:
    if info.flag_bits & (0x1 | 0x40 | 0x2000):
        raise InspectionError("encrypted OOXML entries are unsupported")
    if info.flag_bits & ~_SUPPORTED_ZIP_FLAGS:
        raise InspectionError("unsupported OOXML general-purpose flags")
    if info.compress_type not in _SUPPORTED_ZIP_COMPRESSION:
        raise InspectionError(
            f"unsupported OOXML compression method {info.compress_type}"
        )
    if info.compress_type != zipfile.ZIP_DEFLATED and info.flag_bits & 0x6:
        raise InspectionError(
            "unsupported OOXML compression-option flags"
        )
    mode = (info.external_attr >> 16) & 0xFFFF
    file_type = stat.S_IFMT(mode)
    if info.create_system == 3 and file_type:
        expected = stat.S_IFDIR if info.is_dir() else stat.S_IFREG
        if file_type != expected:
            raise InspectionError("unsupported non-regular OOXML entry")
    if info.file_size < 0 or info.compress_size < 0:
        raise InspectionError("invalid OOXML entry size metadata")
    if info.file_size:
        if info.compress_size == 0:
            raise InspectionError("suspicious OOXML compression ratio")
        ratio = info.file_size / info.compress_size
        if ratio > _MAX_COMPRESSION_RATIO:
            raise InspectionError(
                "suspicious OOXML compression ratio: "
                f"{ratio:.1f} > {_MAX_COMPRESSION_RATIO:.1f}"
            )


def _read_local_header_bytes(
    descriptor: int,
    size: int,
    offset: int,
) -> bytes:
    payload = bytearray()
    try:
        while len(payload) < size:
            chunk = os.pread(
                descriptor,
                size - len(payload),
                offset + len(payload),
            )
            if not chunk:
                break
            payload.extend(chunk)
    except OSError as exc:
        raise InspectionError(
            "OOXML local file header could not be read: "
            f"{bounded_diagnostic(exc)}"
        ) from exc
    if len(payload) != size:
        raise InspectionError("OOXML local file header is truncated")
    return bytes(payload)


def _local_zip64_sizes(
    extra: bytes,
    compressed_size: int,
    uncompressed_size: int,
) -> tuple[int, int]:
    zip64_payload = _local_zip64_payload(extra)
    needs_uncompressed = uncompressed_size == _ZIP32_MAX
    needs_compressed = compressed_size == _ZIP32_MAX
    if not needs_uncompressed and not needs_compressed:
        return compressed_size, uncompressed_size
    if zip64_payload is None:
        raise InspectionError(
            "OOXML local file header is missing ZIP64 size metadata"
        )

    zip64_offset = 0

    def read_size() -> int:
        nonlocal zip64_offset
        if len(zip64_payload) - zip64_offset < 8:
            raise InspectionError(
                "OOXML local file header has truncated ZIP64 size metadata"
            )
        value = struct.unpack_from("<Q", zip64_payload, zip64_offset)[0]
        zip64_offset += 8
        return value

    if needs_uncompressed:
        uncompressed_size = read_size()
    if needs_compressed:
        compressed_size = read_size()
    return compressed_size, uncompressed_size


def _local_zip64_payload(extra: bytes) -> bytes | None:
    offset = 0
    zip64_payload: bytes | None = None
    while offset < len(extra):
        if len(extra) - offset < 4:
            raise InspectionError(
                "OOXML local file header has a truncated extra field"
            )
        field_id, field_size = struct.unpack_from("<HH", extra, offset)
        offset += 4
        field_end = offset + field_size
        if field_end > len(extra):
            raise InspectionError(
                "OOXML local file header has a truncated extra field"
            )
        if field_id == _ZIP64_EXTRA_ID:
            if zip64_payload is not None:
                raise InspectionError(
                    "OOXML local file header has duplicate ZIP64 metadata"
                )
            zip64_payload = extra[offset:field_end]
        offset = field_end
    return zip64_payload


def _descriptor_uses_zip64(
    crc: int,
    compressed_size: int,
    uncompressed_size: int,
    extra: bytes,
) -> bool:
    zip64_payload = _local_zip64_payload(extra)
    if crc != 0:
        raise InspectionError(
            "OOXML local file header descriptor placeholders conflict"
        )
    if compressed_size == 0 and uncompressed_size == 0:
        if zip64_payload is not None:
            raise InspectionError(
                "OOXML local file header descriptor placeholders conflict"
            )
        return False
    if (
        compressed_size == _ZIP32_MAX
        and uncompressed_size == _ZIP32_MAX
        and zip64_payload is not None
        and len(zip64_payload) == 16
        and struct.unpack("<QQ", zip64_payload) == (0, 0)
    ):
        return True
    raise InspectionError(
        "OOXML local file header descriptor placeholders conflict"
    )


def _validate_data_descriptor(
    descriptor: int,
    member: zipfile.ZipInfo,
    descriptor_start: int,
    boundary: int,
    *,
    zip64: bool,
) -> int:
    if descriptor_start > boundary:
        raise InspectionError(
            "OOXML data descriptor overlaps the next archive region"
        )
    body_struct = (
        _ZIP64_DATA_DESCRIPTOR
        if zip64
        else _ZIP32_DATA_DESCRIPTOR
    )
    unsigned_size = body_struct.size
    signed_size = len(_DATA_DESCRIPTOR_SIGNATURE) + unsigned_size
    available = boundary - descriptor_start
    if available < unsigned_size:
        raise InspectionError("OOXML data descriptor is truncated")
    if available > signed_size:
        raise InspectionError("OOXML data descriptor has wrong width")
    payload = _read_local_header_bytes(
        descriptor,
        available,
        descriptor_start,
    )
    if available not in {unsigned_size, signed_size}:
        if payload.startswith(_DATA_DESCRIPTOR_SIGNATURE):
            raise InspectionError("OOXML data descriptor is truncated")
        raise InspectionError("OOXML data descriptor has wrong width")
    has_signature = available == signed_size
    if has_signature and not payload.startswith(_DATA_DESCRIPTOR_SIGNATURE):
        raise InspectionError("OOXML data descriptor has wrong width")
    body = (
        payload[len(_DATA_DESCRIPTOR_SIGNATURE):]
        if has_signature
        else payload
    )
    actual_crc, actual_compressed, actual_uncompressed = body_struct.unpack(body)
    if actual_crc != member.CRC:
        raise InspectionError(
            "OOXML data descriptor CRC differs from central directory"
        )
    if actual_compressed != member.compress_size:
        raise InspectionError(
            "OOXML data descriptor compressed size differs "
            "from central directory"
        )
    if actual_uncompressed != member.file_size:
        raise InspectionError(
            "OOXML data descriptor uncompressed size differs "
            "from central directory"
        )
    return boundary


def _central_filename_bytes(member: zipfile.ZipInfo, name: str) -> bytes:
    encoding = "utf-8" if member.flag_bits & 0x800 else "cp437"
    try:
        return name.encode(encoding)
    except UnicodeEncodeError as exc:
        raise InspectionError(
            "OOXML central directory filename encoding is invalid"
        ) from exc


def _validate_local_headers(
    descriptor: int,
    archive: zipfile.ZipFile,
    members: list[zipfile.ZipInfo],
) -> None:
    file_size = os.fstat(descriptor).st_size
    central_start = getattr(archive, "start_dir", None)
    if (
        not isinstance(central_start, int)
        or central_start < 0
        or central_start > file_size
    ):
        raise InspectionError(
            "OOXML local file header bounds are invalid"
        )

    intervals: list[tuple[int, int]] = []
    local_metadata_bytes = 0
    for member in members:
        if (
            not isinstance(member.header_offset, int)
            or member.header_offset < 0
        ):
            raise InspectionError(
                "OOXML local file header is outside archive bounds"
            )
    ordered_members = sorted(
        members,
        key=lambda member: member.header_offset,
    )
    for index, member in enumerate(ordered_members):
        name = _member_name(member)
        header_offset = member.header_offset
        boundary = (
            ordered_members[index + 1].header_offset
            if index + 1 < len(ordered_members)
            else central_start
        )
        if boundary <= header_offset or boundary > central_start:
            raise InspectionError(
                "OOXML local file header regions overlap"
            )
        fixed_end = header_offset + _LOCAL_FILE_HEADER.size
        if fixed_end > central_start or fixed_end > file_size:
            raise InspectionError(
                "OOXML local file header is outside archive bounds"
            )

        fixed = _read_local_header_bytes(
            descriptor,
            _LOCAL_FILE_HEADER.size,
            header_offset,
        )
        (
            signature,
            _extract_version,
            flags,
            compression_method,
            _modified_time,
            _modified_date,
            crc,
            compressed_size,
            uncompressed_size,
            filename_size,
            extra_size,
        ) = _LOCAL_FILE_HEADER.unpack(fixed)
        if signature != _LOCAL_FILE_HEADER_SIGNATURE:
            raise InspectionError(
                "OOXML local file header signature is invalid"
            )

        central_filename = _central_filename_bytes(member, name)
        if filename_size != len(central_filename):
            raise InspectionError(
                "OOXML local file header filename length differs "
                "from central directory"
            )
        local_metadata_bytes += (
            _LOCAL_FILE_HEADER.size + filename_size + extra_size
        )
        if local_metadata_bytes > _MAX_OOXML_LOCAL_METADATA_BYTES:
            raise InspectionError(
                "OOXML local file header metadata exceeds configured limits"
            )

        variable_size = filename_size + extra_size
        data_start = fixed_end + variable_size
        if data_start > central_start or data_start > file_size:
            raise InspectionError(
                "OOXML local file header extends outside archive bounds"
            )
        variable = _read_local_header_bytes(
            descriptor,
            variable_size,
            fixed_end,
        )
        local_filename = variable[:filename_size]
        local_extra = variable[filename_size:]
        if local_filename != central_filename:
            raise InspectionError(
                "OOXML local file header filename bytes differ "
                "from central directory"
            )
        encoding = "utf-8" if flags & 0x800 else "cp437"
        try:
            decoded_filename = local_filename.decode(encoding)
        except UnicodeDecodeError as exc:
            raise InspectionError(
                "OOXML local file header filename encoding is invalid"
            ) from exc
        if decoded_filename != name:
            raise InspectionError(
                "OOXML local file header logical filename differs "
                "from central directory"
            )

        if (
            flags != member.flag_bits
            or compression_method != member.compress_type
        ):
            raise InspectionError(
                "OOXML local file header metadata differs "
                "from central directory"
            )
        uses_descriptor = bool(flags & 0x8)
        descriptor_is_zip64 = False
        if uses_descriptor:
            descriptor_is_zip64 = _descriptor_uses_zip64(
                crc,
                compressed_size,
                uncompressed_size,
                local_extra,
            )
        else:
            if crc != member.CRC:
                raise InspectionError(
                    "OOXML local file header metadata differs "
                    "from central directory"
                )
            local_compressed, local_uncompressed = _local_zip64_sizes(
                local_extra,
                compressed_size,
                uncompressed_size,
            )
            if (
                local_compressed != member.compress_size
                or local_uncompressed != member.file_size
            ):
                raise InspectionError(
                    "OOXML local file header sizes differ "
                    "from central directory"
                )

        data_end = data_start + member.compress_size
        if data_end > boundary:
            raise InspectionError(
                "OOXML local file header regions overlap"
            )
        if data_end > central_start or data_end > file_size:
            raise InspectionError(
                "OOXML local file header data extends outside archive bounds"
            )
        interval_end = (
            _validate_data_descriptor(
                descriptor,
                member,
                data_end,
                boundary,
                zip64=descriptor_is_zip64,
            )
            if uses_descriptor
            else data_end
        )
        intervals.append((header_offset, interval_end))

    previous_end = -1
    for interval_start, interval_end in sorted(intervals):
        if interval_start < previous_end:
            raise InspectionError(
                "OOXML local file header regions overlap"
            )
        previous_end = interval_end


def _validate_content_types(
    payload: bytes,
    names: set[str],
) -> dict[str, str]:
    if len(payload) > _MAX_CONTENT_TYPES_BYTES:
        raise InspectionError("OOXML content types exceed size limit")
    lowered = payload.lower()
    if b"<!doctype" in lowered or b"<!entity" in lowered:
        raise InspectionError("OOXML content types contain forbidden declarations")
    try:
        root = ElementTree.fromstring(payload)
    except ElementTree.ParseError as exc:
        raise InspectionError("OOXML content types are malformed") from exc

    types_tag = f"{{{_CONTENT_TYPES_NAMESPACE}}}Types"
    default_tag = f"{{{_CONTENT_TYPES_NAMESPACE}}}Default"
    override_tag = f"{{{_CONTENT_TYPES_NAMESPACE}}}Override"
    if root.tag != types_tag or root.attrib:
        raise InspectionError("OOXML content types root is invalid")

    defaults: dict[str, str] = {}
    overrides: dict[str, str] = {}
    for element in root:
        if len(element):
            raise InspectionError("OOXML content types child is invalid")
        if element.tag == default_tag:
            if set(element.attrib) != {"Extension", "ContentType"}:
                raise InspectionError("OOXML content types Default is invalid")
            extension = element.attrib["Extension"]
            content_type = element.attrib["ContentType"]
            key = extension.casefold()
            if (
                not extension
                or extension.startswith(".")
                or "/" in extension
                or "\\" in extension
                or key in defaults
            ):
                raise InspectionError(
                    "OOXML content types contain duplicate or invalid Default"
                )
            defaults[key] = content_type
        elif element.tag == override_tag:
            if set(element.attrib) != {"PartName", "ContentType"}:
                raise InspectionError("OOXML content types Override is invalid")
            part_name = element.attrib["PartName"]
            content_type = element.attrib["ContentType"]
            if (
                not part_name.startswith("/")
                or part_name == "/"
                or part_name in overrides
            ):
                raise InspectionError(
                    "OOXML content types contain duplicate or invalid PartName"
                )
            archive_name = part_name[1:]
            if (
                not artifact_security._safe_zip_member(archive_name)
                or archive_name not in names
                or archive_name.endswith("/")
            ):
                raise InspectionError(
                    "OOXML content types reference a missing or invalid part"
                )
            overrides[part_name] = content_type
        else:
            raise InspectionError("OOXML content types child is invalid")

        if (
            not content_type
            or content_type != content_type.strip()
            or any(ord(character) < 0x20 for character in content_type)
        ):
            raise InspectionError("OOXML content types mapping is invalid")

    mapped: dict[str, str] = {}
    for name in names:
        if name == "[Content_Types].xml" or name.endswith("/"):
            continue
        content_type = overrides.get(f"/{name}")
        if content_type is None:
            extension = (
                name.rsplit(".", 1)[1].casefold()
                if "." in name.rsplit("/", 1)[-1]
                else ""
            )
            content_type = defaults.get(extension)
        if content_type is None:
            raise InspectionError(
                f"OOXML content types do not map part {bounded_diagnostic(name)!r}"
            )
        mapped[name] = content_type

    if (
        mapped.get("xl/workbook.xml") != _XLSX_WORKBOOK_CONTENT_TYPE
        or overrides.get("/xl/workbook.xml")
        != _XLSX_WORKBOOK_CONTENT_TYPE
    ):
        raise InspectionError("OOXML content types do not describe an XLSX workbook")
    return mapped


def _read_bounded_member(
    archive: zipfile.ZipFile,
    member: zipfile.ZipInfo,
    limit: int,
    label: str,
) -> bytes:
    payload = bytearray()
    with archive.open(member, "r") as entry:
        while True:
            chunk = entry.read(min(_ZIP_STREAM_CHUNK, limit - len(payload) + 1))
            if not chunk:
                break
            payload.extend(chunk)
            if len(payload) > limit:
                raise InspectionError(f"{label} exceed size limit")
    if len(payload) != member.file_size:
        raise InspectionError("OOXML entry size metadata does not match content")
    return bytes(payload)


def _is_xml_content_type(content_type: str | None) -> bool:
    if content_type is None:
        return False
    media_type = content_type.split(";", 1)[0].strip().casefold()
    return (
        media_type in {"application/xml", "text/xml"}
        or media_type.endswith("+xml")
    )


@dataclass(frozen=True)
class _Relationship:
    id: str
    type: str
    target: str | None
    external: bool


def _parse_graph_xml(
    archive: zipfile.ZipFile,
    members: dict[str, zipfile.ZipInfo],
    name: str,
    label: str,
):
    member = members.get(name)
    if member is None or member.is_dir():
        raise InspectionError(f"OOXML {label} part is missing")
    payload = _read_bounded_member(
        archive,
        member,
        _MAX_OOXML_GRAPH_PART_BYTES,
        f"OOXML {label} part",
    )
    lowered = payload.lower()
    if b"<!doctype" in lowered or b"<!entity" in lowered:
        raise InspectionError(
            f"OOXML {label} part contains forbidden declarations"
        )
    try:
        return ElementTree.fromstring(payload)
    except ElementTree.ParseError as exc:
        raise InspectionError(f"OOXML {label} part is malformed") from exc


def _relationship_part_name(source_part: str) -> str:
    parent, _, filename = source_part.rpartition("/")
    prefix = f"{parent}/" if parent else ""
    return f"{prefix}_rels/{filename}.rels"


def _normalize_relationship_target(source_part: str, target: str) -> str:
    if not target or "\x00" in target or "\\" in target:
        raise InspectionError("OOXML relationship target is invalid")
    split = urlsplit(target)
    if split.scheme or split.netloc or split.query or split.fragment:
        raise InspectionError("OOXML relationship target is not an internal part")
    try:
        decoded = unquote_to_bytes(split.path).decode("utf-8")
    except UnicodeDecodeError as exc:
        raise InspectionError(
            "OOXML relationship target encoding is invalid"
        ) from exc
    if not decoded or "\x00" in decoded or "\\" in decoded:
        raise InspectionError("OOXML relationship target is invalid")

    parts = [] if decoded.startswith("/") else source_part.split("/")[:-1]
    for part in decoded.lstrip("/").split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if not parts:
                raise InspectionError(
                    "OOXML relationship target escapes the package"
                )
            parts.pop()
            continue
        parts.append(part)
    normalized = "/".join(parts)
    if not artifact_security._safe_zip_member(normalized):
        raise InspectionError("OOXML relationship target is invalid")
    return normalized


def _parse_relationships(
    archive: zipfile.ZipFile,
    members: dict[str, zipfile.ZipInfo],
    names: set[str],
    part_content_types: dict[str, str],
    source_part: str,
    *,
    required: bool = False,
) -> tuple[str | None, dict[str, _Relationship]]:
    relationship_part = _relationship_part_name(source_part)
    if relationship_part not in members:
        if required:
            raise InspectionError("OOXML workbook relationships are missing")
        return None, {}
    if (
        part_content_types.get(relationship_part)
        != _RELATIONSHIPS_CONTENT_TYPE
    ):
        raise InspectionError(
            "OOXML relationship part content type is invalid"
        )
    root = _parse_graph_xml(
        archive,
        members,
        relationship_part,
        "relationship",
    )
    relationships_tag = (
        f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationships"
    )
    relationship_tag = (
        f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship"
    )
    if root.tag != relationships_tag:
        raise InspectionError("OOXML relationship root is invalid")

    relationships: dict[str, _Relationship] = {}
    for element in root:
        if element.tag != relationship_tag or len(element):
            raise InspectionError("OOXML relationship entry is invalid")
        if not set(element.attrib).issubset(
            {"Id", "Type", "Target", "TargetMode"}
        ):
            raise InspectionError("OOXML relationship entry is invalid")
        relationship_id = element.attrib.get("Id")
        relationship_type = element.attrib.get("Type")
        target = element.attrib.get("Target")
        target_mode = element.attrib.get("TargetMode")
        if (
            not relationship_id
            or not relationship_type
            or not target
            or relationship_id in relationships
        ):
            raise InspectionError(
                "OOXML relationship id, type, or target is invalid"
            )
        if len(relationships) >= _MAX_RELATIONSHIPS:
            raise InspectionError(
                "OOXML relationship count exceeds configured limit"
            )
        if target_mode not in {None, "External"}:
            raise InspectionError("OOXML relationship target mode is invalid")
        external = target_mode == "External"
        normalized_target = (
            None
            if external
            else _normalize_relationship_target(source_part, target)
        )
        if normalized_target is not None and normalized_target not in names:
            raise InspectionError(
                "OOXML relationship target part is missing"
            )
        relationships[relationship_id] = _Relationship(
            id=relationship_id,
            type=relationship_type,
            target=normalized_target,
            external=external,
        )
    return relationship_part, relationships


def _require_relationship_role(
    roles: dict[str, set[str]],
    part_content_types: dict[str, str],
    relationship: _Relationship,
    role: str,
    expected_content_type: str,
) -> str:
    if relationship.external or relationship.target is None:
        raise InspectionError(
            f"OOXML {role} relationship must target an internal part"
        )
    if part_content_types.get(relationship.target) != expected_content_type:
        raise InspectionError(
            f"OOXML {role} relationship content type is invalid"
        )
    roles.setdefault(relationship.target, set()).add(role)
    return relationship.target


def _build_part_roles(
    archive: zipfile.ZipFile,
    members: dict[str, zipfile.ZipInfo],
    names: set[str],
    part_content_types: dict[str, str],
) -> dict[str, set[str]]:
    workbook_name = "xl/workbook.xml"
    workbook = _parse_graph_xml(
        archive,
        members,
        workbook_name,
        "workbook",
    )
    workbook_tag = f"{{{_SPREADSHEET_NAMESPACE}}}workbook"
    sheets_tag = f"{{{_SPREADSHEET_NAMESPACE}}}sheets"
    sheet_tag = f"{{{_SPREADSHEET_NAMESPACE}}}sheet"
    relationship_id_attr = (
        f"{{{_OFFICE_RELATIONSHIPS_NAMESPACE}}}id"
    )
    if workbook.tag != workbook_tag:
        raise InspectionError("OOXML workbook root is invalid")
    parents = {
        id(child): parent
        for parent in workbook.iter()
        for child in parent
    }
    sheets_nodes = []
    sheet_nodes = []
    for element in workbook.iter():
        local_name = _xml_local_name(element.tag)
        if local_name == "sheets":
            if (
                element.tag != sheets_tag
                or parents.get(id(element)) is not workbook
            ):
                raise InspectionError(
                    "OOXML workbook sheets container is invalid"
                )
            sheets_nodes.append(element)
        elif local_name == "sheet":
            sheet_nodes.append(element)
    if len(sheets_nodes) != 1:
        raise InspectionError(
            "OOXML workbook sheets container must appear exactly once"
        )
    sheets = sheets_nodes[0]
    if any(
        sheet.tag != sheet_tag or parents.get(id(sheet)) is not sheets
        for sheet in sheet_nodes
    ) or any(child.tag != sheet_tag for child in sheets):
        raise InspectionError("OOXML workbook sheet element is invalid")
    if len(sheet_nodes) != len(sheets):
        raise InspectionError("OOXML workbook sheet structure is invalid")

    roles: dict[str, set[str]] = {workbook_name: {"workbook"}}
    relationship_parts: set[str] = set()
    relationship_cache: dict[str, dict[str, _Relationship]] = {}

    def relationships_for(
        source_part: str,
        *,
        required: bool = False,
    ) -> dict[str, _Relationship]:
        if source_part in relationship_cache:
            return relationship_cache[source_part]
        part_name, parsed = _parse_relationships(
            archive,
            members,
            names,
            part_content_types,
            source_part,
            required=required,
        )
        relationship_cache[source_part] = parsed
        if part_name is not None:
            relationship_parts.add(part_name)
        return parsed

    workbook_relationships = relationships_for(
        workbook_name,
        required=True,
    )
    sheet_relationship_ids: set[str] = set()
    sheet_targets: set[str] = set()
    sheet_parts: list[str] = []
    for sheet in sheets:
        relationship_id = sheet.attrib.get(relationship_id_attr)
        if (
            not relationship_id
            or relationship_id in sheet_relationship_ids
        ):
            raise InspectionError(
                "OOXML sheet relationship id is missing or duplicated"
            )
        sheet_relationship_ids.add(relationship_id)
        relationship = workbook_relationships.get(relationship_id)
        if relationship is None:
            raise InspectionError("OOXML sheet relationship is missing")
        if relationship.type == _WORKSHEET_RELATIONSHIP_TYPE:
            role = "worksheet"
            content_type = _WORKSHEET_CONTENT_TYPE
        elif relationship.type == _CHARTSHEET_RELATIONSHIP_TYPE:
            role = "chartsheet"
            content_type = _CHARTSHEET_CONTENT_TYPE
        else:
            raise InspectionError("OOXML sheet relationship type is invalid")
        target = _require_relationship_role(
            roles,
            part_content_types,
            relationship,
            role,
            content_type,
        )
        if target in sheet_targets:
            raise InspectionError("OOXML sheet relationship target is duplicated")
        sheet_targets.add(target)
        sheet_parts.append(target)

    def require_unique_parser_part(
        *,
        label: str,
        role: str,
        relationship_type: str,
        content_type: str,
        parser_path: str | None = None,
    ) -> None:
        declared_parts = [
            part
            for part, mapped_type in part_content_types.items()
            if mapped_type == content_type
        ]
        relationships = [
            relationship
            for relationship in workbook_relationships.values()
            if relationship.type == relationship_type
        ]
        parser_part_exists = parser_path is not None and parser_path in names
        if not declared_parts and not relationships and not parser_part_exists:
            return
        if len(declared_parts) != 1 or len(relationships) != 1:
            raise InspectionError(
                f"OOXML {label} part and relationship must be unique"
            )
        target = _require_relationship_role(
            roles,
            part_content_types,
            relationships[0],
            role,
            content_type,
        )
        if target != declared_parts[0] or (
            parser_path is not None and target != parser_path
        ):
            raise InspectionError(
                f"OOXML {label} part, relationship, and parser target "
                "are inconsistent"
            )

    require_unique_parser_part(
        label="styles",
        role="styles",
        relationship_type=_STYLES_RELATIONSHIP_TYPE,
        content_type=_STYLES_CONTENT_TYPE,
        parser_path="xl/styles.xml",
    )
    require_unique_parser_part(
        label="shared strings",
        role="shared_strings",
        relationship_type=_SHARED_STRINGS_RELATIONSHIP_TYPE,
        content_type=_SHARED_STRINGS_CONTENT_TYPE,
    )

    drawing_parts: set[str] = set()
    for sheet_part in sheet_parts:
        for relationship in relationships_for(sheet_part).values():
            if relationship.type == _DRAWING_RELATIONSHIP_TYPE:
                drawing_parts.add(
                    _require_relationship_role(
                        roles,
                        part_content_types,
                        relationship,
                        "drawing",
                        _DRAWING_CONTENT_TYPE,
                    )
                )
            elif relationship.type == _CHART_RELATIONSHIP_TYPE:
                _require_relationship_role(
                    roles,
                    part_content_types,
                    relationship,
                    "chart",
                    _CHART_CONTENT_TYPE,
                )

    for drawing_part in drawing_parts:
        for relationship in relationships_for(drawing_part).values():
            if relationship.type == _CHART_RELATIONSHIP_TYPE:
                _require_relationship_role(
                    roles,
                    part_content_types,
                    relationship,
                    "chart",
                    _CHART_CONTENT_TYPE,
                )

    for part_name in relationship_parts:
        roles.setdefault(part_name, set()).add("relationships")
    return roles


def _xml_local_name(tag: object) -> str:
    if not isinstance(tag, str):
        return ""
    return tag.rsplit("}", 1)[-1]


class _XmlStructureValidator:
    _WORKSHEET_ROOT_TAG = f"{{{_SPREADSHEET_NAMESPACE}}}worksheet"
    _STYLE_ROOT_TAG = f"{{{_SPREADSHEET_NAMESPACE}}}styleSheet"
    _SHARED_STRING_ROOT_TAG = f"{{{_SPREADSHEET_NAMESPACE}}}sst"
    _DRAWING_ROOT_TAG = f"{{{_SPREADSHEET_DRAWING_NAMESPACE}}}wsDr"
    _CHART_ROOT_TAG = f"{{{_CHART_NAMESPACE}}}chartSpace"
    _STYLE_TAGS = {
        f"{{{_SPREADSHEET_NAMESPACE}}}{name}"
        for name in {
            "numFmt",
            "font",
            "fill",
            "border",
            "xf",
            "cellStyle",
            "dxf",
            "tableStyle",
        }
    }
    _STYLE_CONTAINERS = {
        name: f"{{{_SPREADSHEET_NAMESPACE}}}{name}"
        for name in {
            "numFmts",
            "fonts",
            "fills",
            "borders",
            "cellStyleXfs",
            "cellXfs",
            "cellStyles",
            "dxfs",
            "tableStyles",
        }
    }
    _STYLE_SEQUENCE_CHILDREN = {
        _STYLE_CONTAINERS["numFmts"]: f"{{{_SPREADSHEET_NAMESPACE}}}numFmt",
        _STYLE_CONTAINERS["fonts"]: f"{{{_SPREADSHEET_NAMESPACE}}}font",
        _STYLE_CONTAINERS["fills"]: f"{{{_SPREADSHEET_NAMESPACE}}}fill",
        _STYLE_CONTAINERS["borders"]: f"{{{_SPREADSHEET_NAMESPACE}}}border",
        _STYLE_CONTAINERS["cellStyleXfs"]: f"{{{_SPREADSHEET_NAMESPACE}}}xf",
        _STYLE_CONTAINERS["cellXfs"]: f"{{{_SPREADSHEET_NAMESPACE}}}xf",
        _STYLE_CONTAINERS["cellStyles"]: (
            f"{{{_SPREADSHEET_NAMESPACE}}}cellStyle"
        ),
        _STYLE_CONTAINERS["dxfs"]: f"{{{_SPREADSHEET_NAMESPACE}}}dxf",
        _STYLE_CONTAINERS["tableStyles"]: (
            f"{{{_SPREADSHEET_NAMESPACE}}}tableStyle"
        ),
    }
    _STYLE_ENTRY_PARENTS = {
        "numFmt": {
            _STYLE_CONTAINERS["numFmts"],
            f"{{{_SPREADSHEET_NAMESPACE}}}dxf",
        },
        "font": {
            _STYLE_CONTAINERS["fonts"],
            f"{{{_SPREADSHEET_NAMESPACE}}}dxf",
        },
        "fill": {
            _STYLE_CONTAINERS["fills"],
            f"{{{_SPREADSHEET_NAMESPACE}}}dxf",
        },
        "border": {
            _STYLE_CONTAINERS["borders"],
            f"{{{_SPREADSHEET_NAMESPACE}}}dxf",
        },
        "xf": {
            _STYLE_CONTAINERS["cellStyleXfs"],
            _STYLE_CONTAINERS["cellXfs"],
        },
        "cellStyle": {_STYLE_CONTAINERS["cellStyles"]},
        "dxf": {_STYLE_CONTAINERS["dxfs"]},
        "tableStyle": {_STYLE_CONTAINERS["tableStyles"]},
    }
    _DRAWING_TAGS = {
        f"{{{_SPREADSHEET_DRAWING_NAMESPACE}}}{name}"
        for name in {"sp", "pic", "graphicFrame", "cxnSp", "grpSp"}
    }
    _DRAWING_ANCHOR_NAMES = {
        "absoluteAnchor",
        "oneCellAnchor",
        "twoCellAnchor",
    }
    _DRAWING_ANCHOR_TAGS = {
        f"{{{_SPREADSHEET_DRAWING_NAMESPACE}}}{name}"
        for name in _DRAWING_ANCHOR_NAMES
    }
    _DRAWING_OBJECT_NAMES = {
        "sp",
        "pic",
        "graphicFrame",
        "cxnSp",
        "grpSp",
    }
    _CHART_TYPE_NAMES = {
        "areaChart",
        "area3DChart",
        "barChart",
        "bar3DChart",
        "bubbleChart",
        "doughnutChart",
        "lineChart",
        "line3DChart",
        "ofPieChart",
        "pieChart",
        "pie3DChart",
        "radarChart",
        "scatterChart",
        "stockChart",
        "surfaceChart",
        "surface3DChart",
    }
    _CHART_AXIS_NAMES = {"catAx", "dateAx", "serAx", "valAx"}
    _ROW_TAG = f"{{{_SPREADSHEET_NAMESPACE}}}row"
    _CELL_TAG = f"{{{_SPREADSHEET_NAMESPACE}}}c"
    _MERGE_TAG = f"{{{_SPREADSHEET_NAMESPACE}}}mergeCell"
    _SHARED_STRING_ITEM_TAG = f"{{{_SPREADSHEET_NAMESPACE}}}si"
    _TEXT_TAG = f"{{{_SPREADSHEET_NAMESPACE}}}t"
    _RELATIONSHIP_TAG = (
        f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship"
    )
    _CHART_TAG = f"{{{_CHART_NAMESPACE}}}chart"
    _CHART_SERIES_TAG = f"{{{_CHART_NAMESPACE}}}ser"

    def __init__(
        self,
        name: str,
        roles: set[str],
        totals: dict[str, int],
    ) -> None:
        self._name = name
        self._totals = totals
        self._parser = ElementTree.XMLPullParser(events=("start", "end"))
        self._bytes = 0
        self._depth = 0
        self._tag_stack: list[str] = []
        self._declaration_tail = b""
        self._worksheet = "worksheet" in roles
        self._shared_strings = "shared_strings" in roles
        self._styles = "styles" in roles
        self._relationships = "relationships" in roles
        self._drawing = "drawing" in roles
        self._chart = "chart" in roles

    def _increment(self, key: str, limit: int, label: str) -> None:
        self._totals[key] = self._totals.get(key, 0) + 1
        if self._totals[key] > limit:
            raise InspectionError(f"OOXML {label} count exceeds configured limit")

    def _count_merge_area(self, element) -> None:
        merge_ref = element.attrib.get("ref")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(merge_ref)
        except (TypeError, ValueError) as exc:
            raise InspectionError("OOXML merge reference is invalid") from exc
        if (
            min_col is None
            or min_row is None
            or max_col is None
            or max_row is None
            or not 1 <= min_col <= max_col <= 16_384
            or not 1 <= min_row <= max_row <= 1_048_576
        ):
            raise InspectionError("OOXML merge reference is invalid or out of range")
        area = (max_col - min_col + 1) * (max_row - min_row + 1)
        if area > _MAX_MERGED_RANGE_CELLS:
            raise InspectionError(
                "OOXML merged range area exceeds configured limit"
            )
        self._totals["merged_cells"] = (
            self._totals.get("merged_cells", 0) + area
        )
        if self._totals["merged_cells"] > _MAX_MERGED_CELLS:
            raise InspectionError(
                "OOXML cumulative merged cell area exceeds configured limit"
            )

    def _validate_root(self, tag: str, parent: str | None) -> None:
        if parent is not None:
            return
        expected = None
        label = ""
        if self._worksheet:
            expected = self._WORKSHEET_ROOT_TAG
            label = "worksheet"
        elif self._shared_strings:
            expected = self._SHARED_STRING_ROOT_TAG
            label = "shared strings"
        elif self._styles:
            expected = self._STYLE_ROOT_TAG
            label = "styles"
        elif self._drawing:
            expected = self._DRAWING_ROOT_TAG
            label = "drawing"
        elif self._chart:
            expected = self._CHART_ROOT_TAG
            label = "chart"
        if expected is not None and tag != expected:
            raise InspectionError(f"OOXML {label} root is invalid")

    def _validate_shared_string_element(
        self,
        tag: str,
        parent: str | None,
    ) -> None:
        if tag != self._SHARED_STRING_ITEM_TAG:
            return
        if parent != self._SHARED_STRING_ROOT_TAG:
            raise InspectionError(
                "OOXML shared string item has invalid namespace or nesting"
            )

    def _validate_worksheet_element(
        self,
        tag: str,
        local_name: str,
        parent: str | None,
    ) -> None:
        sheet_data = f"{{{_SPREADSHEET_NAMESPACE}}}sheetData"
        merge_cells = f"{{{_SPREADSHEET_NAMESPACE}}}mergeCells"
        if parent == self._ROW_TAG and tag != self._CELL_TAG:
            raise InspectionError(
                "OOXML worksheet row contains a non-cell child"
            )
        if parent == merge_cells and tag != self._MERGE_TAG:
            raise InspectionError(
                "OOXML worksheet merge container contains an invalid child"
            )
        expected = {
            "sheetData": (sheet_data, self._WORKSHEET_ROOT_TAG),
            "row": (self._ROW_TAG, sheet_data),
            "c": (self._CELL_TAG, self._ROW_TAG),
            "mergeCells": (merge_cells, self._WORKSHEET_ROOT_TAG),
            "mergeCell": (self._MERGE_TAG, merge_cells),
        }.get(local_name)
        if expected is None:
            return
        expected_tag, expected_parent = expected
        if tag == expected_tag and parent != expected_parent:
            raise InspectionError(
                "OOXML worksheet parser-sensitive element has invalid nesting"
            )
        if parent == expected_parent and tag != expected_tag:
            raise InspectionError(
                "OOXML worksheet parser-sensitive element has invalid namespace"
            )

    def _validate_style_element(
        self,
        tag: str,
        local_name: str,
        parent: str | None,
    ) -> None:
        if local_name in self._STYLE_CONTAINERS:
            if (
                tag != self._STYLE_CONTAINERS[local_name]
                or parent != self._STYLE_ROOT_TAG
            ):
                raise InspectionError(
                    "OOXML styles container has invalid namespace or nesting"
                )
        expected_child = self._STYLE_SEQUENCE_CHILDREN.get(parent)
        if expected_child is not None and tag != expected_child:
            raise InspectionError(
                "OOXML styles sequence contains an invalid child"
            )
        allowed_parents = self._STYLE_ENTRY_PARENTS.get(local_name)
        if allowed_parents is None:
            return
        expected_tag = f"{{{_SPREADSHEET_NAMESPACE}}}{local_name}"
        if tag == expected_tag and parent not in allowed_parents:
            raise InspectionError(
                "OOXML styles record has invalid nesting"
            )
        if parent in allowed_parents and tag != expected_tag:
            raise InspectionError(
                "OOXML styles record has invalid namespace"
            )

    def _validate_drawing_element(
        self,
        tag: str,
        local_name: str,
        parent: str | None,
    ) -> None:
        expected_tag = (
            f"{{{_SPREADSHEET_DRAWING_NAMESPACE}}}{local_name}"
        )
        if local_name in self._DRAWING_ANCHOR_NAMES:
            if tag == expected_tag and parent != self._DRAWING_ROOT_TAG:
                raise InspectionError("OOXML drawing anchor has invalid nesting")
            if parent == self._DRAWING_ROOT_TAG and tag != expected_tag:
                raise InspectionError(
                    "OOXML drawing anchor has invalid namespace"
                )
        if local_name not in self._DRAWING_OBJECT_NAMES:
            return
        allowed_parents = self._DRAWING_ANCHOR_TAGS | {
            f"{{{_SPREADSHEET_DRAWING_NAMESPACE}}}grpSp"
        }
        if tag == expected_tag and parent not in allowed_parents:
            raise InspectionError("OOXML drawing object has invalid nesting")
        if parent in allowed_parents and tag != expected_tag:
            raise InspectionError("OOXML drawing object has invalid namespace")

    def _validate_chart_element(
        self,
        tag: str,
        local_name: str,
        parent: str | None,
    ) -> None:
        chart = f"{{{_CHART_NAMESPACE}}}chart"
        plot_area = f"{{{_CHART_NAMESPACE}}}plotArea"
        expected_tag = f"{{{_CHART_NAMESPACE}}}{local_name}"
        expected_parent = None
        if local_name == "chart":
            expected_parent = self._CHART_ROOT_TAG
        elif local_name == "plotArea":
            expected_parent = chart
        elif (
            local_name in self._CHART_TYPE_NAMES
            or local_name in self._CHART_AXIS_NAMES
        ):
            expected_parent = plot_area
        elif local_name == "ser":
            if parent is not None and _xml_local_name(parent) in self._CHART_TYPE_NAMES:
                if tag != expected_tag or not parent.startswith(
                    f"{{{_CHART_NAMESPACE}}}"
                ):
                    raise InspectionError(
                        "OOXML chart series has invalid namespace or nesting"
                    )
            elif tag == expected_tag:
                raise InspectionError("OOXML chart series has invalid nesting")
            return
        if expected_parent is None:
            return
        if tag == expected_tag and parent != expected_parent:
            raise InspectionError(
                "OOXML chart element has invalid nesting"
            )
        if parent == expected_parent and tag != expected_tag:
            raise InspectionError(
                "OOXML chart element has invalid namespace"
            )

    def _validate_parser_sensitive_element(
        self,
        tag: str,
        local_name: str,
        parent: str | None,
    ) -> None:
        self._validate_root(tag, parent)
        if self._worksheet:
            self._validate_worksheet_element(tag, local_name, parent)
        if self._shared_strings:
            self._validate_shared_string_element(tag, parent)
        if self._styles:
            self._validate_style_element(tag, local_name, parent)
        if self._drawing:
            self._validate_drawing_element(tag, local_name, parent)
        if self._chart:
            self._validate_chart_element(tag, local_name, parent)

    def _drain_events(self) -> None:
        for event, element in self._parser.read_events():
            local_name = _xml_local_name(element.tag)
            if event == "start":
                parent = self._tag_stack[-1] if self._tag_stack else None
                self._depth += 1
                if self._depth > _MAX_OOXML_XML_DEPTH:
                    raise InspectionError(
                        "OOXML XML depth exceeds configured limit"
                    )
                self._increment(
                    "nodes",
                    _MAX_OOXML_XML_NODES,
                    "XML node",
                )
                self._validate_parser_sensitive_element(
                    element.tag,
                    local_name,
                    parent,
                )
                self._tag_stack.append(element.tag)
                if self._worksheet:
                    if element.tag == self._ROW_TAG:
                        dimension_attributes = {
                            key
                            for key in element.attrib
                            if not key.startswith("{") and key not in {"r", "spans"}
                        }
                        if dimension_attributes:
                            self._increment(
                                "row_dimensions",
                                _MAX_ROW_DIMENSIONS,
                                "row dimension",
                            )
                    elif element.tag == self._CELL_TAG:
                        self._increment(
                            "worksheet_cells",
                            _MAX_WORKSHEET_CELLS,
                            "worksheet cell",
                        )
                    elif element.tag == self._MERGE_TAG:
                        self._increment(
                            "worksheet_merges",
                            _MAX_WORKSHEET_MERGES,
                            "worksheet merge",
                        )
                        self._count_merge_area(element)
                if (
                    self._shared_strings
                    and element.tag == self._SHARED_STRING_ITEM_TAG
                ):
                    self._increment(
                        "shared_string_items",
                        _MAX_SHARED_STRING_ITEMS,
                        "shared string item",
                    )
                if (
                    self._styles
                    and element.tag in self._STYLE_TAGS
                ):
                    self._increment(
                        "style_records",
                        _MAX_STYLE_RECORDS,
                        "style record",
                    )
                if (
                    self._relationships
                    and element.tag == self._RELATIONSHIP_TAG
                ):
                    self._increment(
                        "relationships",
                        _MAX_RELATIONSHIPS,
                        "relationship",
                    )
                if (
                    self._drawing
                    and element.tag in self._DRAWING_TAGS
                ):
                    self._increment(
                        "drawing_objects",
                        _MAX_DRAWING_OBJECTS,
                        "drawing object",
                    )
                if self._chart and (
                    element.tag in {self._CHART_TAG, self._CHART_SERIES_TAG}
                    or (
                        element.tag.startswith(f"{{{_CHART_NAMESPACE}}}")
                        and local_name.endswith("Chart")
                    )
                ):
                    self._increment(
                        "chart_objects",
                        _MAX_CHART_OBJECTS,
                        "chart object",
                    )
            else:
                if (
                    self._shared_strings
                    and element.tag == self._TEXT_TAG
                ):
                    chars = len(element.text or "")
                    self._totals["shared_string_chars"] = (
                        self._totals.get("shared_string_chars", 0) + chars
                    )
                    if (
                        self._totals["shared_string_chars"]
                        > _MAX_SHARED_STRING_CHARS
                    ):
                        raise InspectionError(
                            "OOXML shared string character count "
                            "exceeds configured limit"
                        )
                if not self._tag_stack or self._tag_stack[-1] != element.tag:
                    raise InspectionError("OOXML XML part has invalid nesting")
                self._tag_stack.pop()
                self._depth -= 1
                element.clear()

    def feed(self, chunk: bytes) -> None:
        self._bytes += len(chunk)
        if self._bytes > _MAX_OOXML_XML_PART_BYTES:
            raise InspectionError(
                f"OOXML XML part bytes exceed configured limit: "
                f"{bounded_diagnostic(self._name)!r}"
            )
        declaration_scan = (self._declaration_tail + chunk).lower()
        if b"<!doctype" in declaration_scan or b"<!entity" in declaration_scan:
            raise InspectionError(
                "OOXML XML part contains forbidden declarations"
            )
        self._declaration_tail = declaration_scan[-16:]
        try:
            self._parser.feed(chunk)
            self._drain_events()
        except ElementTree.ParseError as exc:
            raise InspectionError(
                f"OOXML XML part is malformed: "
                f"{bounded_diagnostic(self._name)!r}"
            ) from exc

    def finish(self) -> None:
        try:
            self._parser.close()
            self._drain_events()
        except ElementTree.ParseError as exc:
            raise InspectionError(
                f"OOXML XML part is malformed: "
                f"{bounded_diagnostic(self._name)!r}"
            ) from exc
        if self._depth != 0 or self._tag_stack:
            raise InspectionError("OOXML XML part has invalid nesting")


def _open_stable_descriptor(path: str, label: str) -> int:
    flags = os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0)
    try:
        proc_fd = _PROC_FD_RE.fullmatch(os.fspath(path))
        descriptor = (
            os.dup(int(proc_fd.group(1)))
            if proc_fd is not None
            else os.open(os.fspath(path), flags)
        )
        if not stat.S_ISREG(os.fstat(descriptor).st_mode):
            os.close(descriptor)
            raise InspectionError(f"{label} must be a regular file")
        return descriptor
    except InspectionError:
        raise
    except OSError as exc:
        raise InspectionError(
            f"{label} could not be opened: {bounded_diagnostic(exc)}"
        ) from exc


def preflight_xlsx(path: str) -> None:
    """Validate an XLSX ZIP before a parser opens it."""
    descriptor = _open_stable_descriptor(path, "OOXML workbook")
    try:
        if not artifact_security._preflight_zip_descriptor(
            descriptor,
            validate_local_bounds=False,
        ):
            raise InspectionError("OOXML ZIP central directory failed preflight")

        with os.fdopen(os.dup(descriptor), "rb") as source:
            try:
                with zipfile.ZipFile(source) as archive:
                    members = archive.infolist()
                    if len(members) > _MAX_OOXML_ENTRIES:
                        raise InspectionError(
                            "OOXML entry count exceeds maximum "
                            f"{_MAX_OOXML_ENTRIES}"
                        )

                    names: set[str] = set()
                    members_by_name: dict[str, zipfile.ZipInfo] = {}
                    collision_keys: set[str] = set()
                    declared_total = 0
                    content_types_info: zipfile.ZipInfo | None = None
                    for member in members:
                        name = _validate_member_name(member)
                        collision_key = unicodedata.normalize(
                            "NFC", name
                        ).casefold()
                        if name in names or collision_key in collision_keys:
                            raise InspectionError(
                                "duplicate or colliding OOXML entry name"
                            )
                        names.add(name)
                        members_by_name[name] = member
                        collision_keys.add(collision_key)
                        declared_total += member.file_size
                        if declared_total > _MAX_OOXML_UNCOMPRESSED_BYTES:
                            raise InspectionError(
                                "OOXML declared uncompressed bytes exceed maximum "
                                f"{_MAX_OOXML_UNCOMPRESSED_BYTES}"
                            )
                        _validate_member_type(member)
                        if name == "[Content_Types].xml":
                            content_types_info = member

                    _validate_local_headers(descriptor, archive, members)
                    if content_types_info is None:
                        raise InspectionError("OOXML content types are missing")
                    if content_types_info.file_size > _MAX_CONTENT_TYPES_BYTES:
                        raise InspectionError(
                            "OOXML content types exceed size limit"
                        )

                    content_types = _read_bounded_member(
                        archive,
                        content_types_info,
                        _MAX_CONTENT_TYPES_BYTES,
                        "OOXML content types",
                    )
                    part_content_types = _validate_content_types(
                        content_types,
                        names,
                    )
                    part_roles = _build_part_roles(
                        archive,
                        members_by_name,
                        names,
                        part_content_types,
                    )

                    actual_total = 0
                    xml_totals: dict[str, int] = {}
                    for member in members:
                        if member.is_dir():
                            if member.file_size != 0:
                                raise InspectionError(
                                    "OOXML directory has invalid size metadata"
                                )
                            continue
                        actual_entry = 0
                        name = _member_name(member)
                        content_type = part_content_types.get(name)
                        xml_validator = (
                            _XmlStructureValidator(
                                name,
                                part_roles.get(name, set()),
                                xml_totals,
                            )
                            if (
                                name == "[Content_Types].xml"
                                or _is_xml_content_type(content_type)
                            )
                            else None
                        )
                        with archive.open(member, "r") as entry:
                            while True:
                                chunk = entry.read(_ZIP_STREAM_CHUNK)
                                if not chunk:
                                    break
                                actual_entry += len(chunk)
                                actual_total += len(chunk)
                                if (
                                    actual_entry > member.file_size
                                    or actual_total
                                    > _MAX_OOXML_UNCOMPRESSED_BYTES
                                ):
                                    raise InspectionError(
                                        "OOXML actual uncompressed bytes exceed "
                                        "declared or configured limits"
                                    )
                                if xml_validator is not None:
                                    xml_validator.feed(chunk)
                        if actual_entry != member.file_size:
                            raise InspectionError(
                                "OOXML entry size metadata does not match content"
                            )
                        if xml_validator is not None:
                            xml_validator.finish()
            except InspectionError:
                raise
            except (
                OSError,
                RuntimeError,
                NotImplementedError,
                ValueError,
                zlib.error,
                zipfile.BadZipFile,
                zipfile.LargeZipFile,
            ) as exc:
                raise InspectionError(
                    f"OOXML ZIP could not be validated: {bounded_diagnostic(exc)}"
                ) from exc
        if actual_total != declared_total:
            raise InspectionError(
                "OOXML total size metadata does not match content"
            )
    finally:
        os.close(descriptor)


def _is_legacy_xls(path: str) -> bool:
    descriptor = _open_stable_descriptor(path, "spreadsheet")
    try:
        return os.pread(descriptor, len(_OLE_SIGNATURE), 0) == _OLE_SIGNATURE
    finally:
        os.close(descriptor)


def _find_libreoffice() -> str:
    executable = shutil.which("libreoffice") or shutil.which("soffice")
    if executable is None:
        raise InspectionError("LibreOffice is not installed or not on PATH")
    return executable


def _write_libreoffice_profile(profile: Path) -> None:
    user = profile / "user"
    user.mkdir(parents=True)
    config = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry">
  <item oor:path="/org.openoffice.Office.Common/Security/Scripting">
    <prop oor:name="MacroSecurityLevel" oor:op="fuse"><value>3</value></prop>
    <prop oor:name="DisableMacrosExecution" oor:op="fuse"><value>true</value></prop>
    <prop oor:name="DisableActiveContent" oor:op="fuse"><value>true</value></prop>
    <prop oor:name="BlockUntrustedRefererLinks" oor:op="fuse"><value>true</value></prop>
  </item>
  <item oor:path="/org.openoffice.Office.Calc/Content/Update">
    <prop oor:name="Link" oor:op="fuse"><value>2</value></prop>
  </item>
</oor:items>
"""
    (user / "registrymodifications.xcu").write_text(config, encoding="utf-8")
    fontconfig = """<?xml version="1.0"?>
<!DOCTYPE fontconfig SYSTEM "urn:fontconfig:fonts.dtd">
<fontconfig>
  <dir>/usr/share/fonts</dir>
  <dir>/usr/local/share/fonts</dir>
  <cachedir>/var/cache/fontconfig</cachedir>
  <include ignore_missing="yes">/etc/fonts/conf.d</include>
  <config><rescan><int>0</int></rescan></config>
</fontconfig>
"""
    (profile / "fontconfig.xml").write_text(fontconfig, encoding="utf-8")


def _is_transient_libreoffice_scratch_error(
    exc: InspectionError,
    profile: Path,
) -> bool:
    message = str(exc)
    if message.startswith(_TRANSIENT_SCRATCH_ENOENT):
        return True
    profile_scan_prefix = (
        f"{_TRANSIENT_SCRATCH_PROFILE_SCAN}work/{profile.name}/"
    )
    return message.startswith(profile_scan_prefix)


def _convert_with_libreoffice(
    path: str,
    scratch_dir: str,
    target_format: str,
) -> str:
    if current_scratch_transaction() is None:
        raise InspectionError(
            "LibreOffice conversion requires an active scratch transaction"
        )
    if target_format not in {"xlsx", "pdf"}:
        raise InspectionError("unsupported LibreOffice conversion target")

    libreoffice_executable = _find_libreoffice()
    env_executable = shutil.which("env")
    if env_executable is None:
        raise InspectionError("env executable is unavailable for LibreOffice")
    profile = Path(tempfile.mkdtemp(prefix="lo-profile-", dir=scratch_dir))
    output_dir = Path(tempfile.mkdtemp(prefix="lo-output-", dir=scratch_dir))
    _write_libreoffice_profile(profile)
    input_descriptor = _open_stable_descriptor(
        path,
        "LibreOffice spreadsheet input",
    )
    input_path = f"/proc/self/fd/{input_descriptor}"
    command = [
        env_executable,
        f"FONTCONFIG_FILE={profile / 'fontconfig.xml'}",
        libreoffice_executable,
        "--headless",
        "--invisible",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--norestore",
        f"-env:UserInstallation={profile.as_uri()}",
        "--convert-to",
        target_format,
        "--outdir",
        str(output_dir),
        input_path,
    ]
    try:
        deadline = time.monotonic() + _LIBREOFFICE_TIMEOUT_SECONDS
        for attempt in range(_LIBREOFFICE_MAX_TRANSIENT_ATTEMPTS):
            remaining_timeout = deadline - time.monotonic()
            if remaining_timeout <= 0:
                raise InspectionError(
                    "LibreOffice conversion exceeded its total timeout"
                )
            try:
                safe_run(
                    command,
                    timeout=remaining_timeout,
                    cwd=scratch_dir,
                    home=scratch_dir,
                    pass_fds=(input_descriptor,),
                )
                break
            except InspectionError as exc:
                if (
                    attempt + 1 >= _LIBREOFFICE_MAX_TRANSIENT_ATTEMPTS
                    or not _is_transient_libreoffice_scratch_error(
                        exc,
                        profile,
                    )
                ):
                    raise

        expected = output_dir / f"{input_descriptor}.{target_format}"
        try:
            output_info = expected.lstat()
        except FileNotFoundError as exc:
            raise InspectionError(
                "LibreOffice did not create the expected output "
                f"{expected.name!r}"
            ) from exc
        if not stat.S_ISREG(output_info.st_mode) or output_info.st_nlink != 1:
            raise InspectionError(
                "LibreOffice output is not a single-link regular file"
            )
        if output_info.st_size <= 0:
            raise InspectionError("LibreOffice output is empty")
        if target_format == "pdf":
            try:
                with expected.open("rb") as source:
                    signature = source.read(5)
            except OSError as exc:
                raise InspectionError(
                    "LibreOffice PDF output could not be read: "
                    f"{bounded_diagnostic(exc)}"
                ) from exc
            if signature != b"%PDF-":
                raise InspectionError("LibreOffice output is not a PDF")
        return str(expected)
    finally:
        os.close(input_descriptor)


@contextmanager
def _opened_workbooks(path: str):
    descriptor = -1
    formula_book = None
    cached_book = None
    try:
        descriptor = _open_stable_descriptor(path, "spreadsheet workbook")
        preflight_xlsx(f"/proc/self/fd/{descriptor}")
        os.lseek(descriptor, 0, os.SEEK_SET)
        with os.fdopen(os.dup(descriptor), "rb") as formula_source:
            formula_book = openpyxl.load_workbook(
                formula_source,
                data_only=False,
                read_only=False,
                keep_links=False,
            )
        os.lseek(descriptor, 0, os.SEEK_SET)
        with os.fdopen(os.dup(descriptor), "rb") as cached_source:
            cached_book = openpyxl.load_workbook(
                cached_source,
                data_only=True,
                read_only=False,
                keep_links=False,
            )
        yield formula_book, cached_book
    except InspectionError:
        raise
    except Exception as exc:
        raise InspectionError(
            f"spreadsheet workbook could not be parsed: {bounded_diagnostic(exc)}"
        ) from exc
    finally:
        if cached_book is not None:
            cached_book.close()
        if formula_book is not None:
            formula_book.close()
        if descriptor >= 0:
            os.close(descriptor)


def _json_scalar(value):
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    return str(value)


def _anchor_metadata(anchor) -> dict:
    if isinstance(anchor, str):
        return {"from": anchor}
    start = getattr(anchor, "_from", None)
    end = getattr(anchor, "to", None)
    metadata = {}
    if start is not None:
        metadata["from"] = (
            f"{get_column_letter(start.col + 1)}{start.row + 1}"
        )
    if end is not None:
        metadata["to"] = f"{get_column_letter(end.col + 1)}{end.row + 1}"
    return metadata


def _sheet_inventory(sheet, total_cells: int, page: int | None) -> dict:
    charts = [
        {
            "type": type(chart).__name__,
            "anchor": _anchor_metadata(getattr(chart, "anchor", None)),
        }
        for chart in getattr(sheet, "_charts", ())
    ]
    images = [
        {
            "kind": "image",
            "type": type(image).__name__,
            "anchor": _anchor_metadata(getattr(image, "anchor", None)),
        }
        for image in getattr(sheet, "_images", ())
    ]
    drawings = [
        {"kind": "chart", **chart}
        for chart in charts
    ] + images
    return {
        "name": sheet.title,
        "type": "worksheet",
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "used_range": sheet.calculate_dimension(),
        "merged_ranges": [
            str(value) for value in sheet.merged_cells.ranges
        ],
        "charts": charts,
        "drawings": drawings,
        "cells": {},
        "cell_page": {
            "page": page,
            "page_size": _CELL_PAGE_SIZE,
            "total": total_cells,
            "returned": 0,
            "omitted": total_cells,
            "omitted_due_to_budget": 0,
        },
    }


def _chartsheet_inventory(sheet) -> dict:
    charts = [
        {
            "type": type(chart).__name__,
            "anchor": _anchor_metadata(getattr(chart, "anchor", None)),
        }
        for chart in getattr(sheet, "_charts", ())
    ]
    return {
        "name": sheet.title,
        "type": "chartsheet",
        "charts": charts,
        "drawings": [
            {"kind": "chart", **chart}
            for chart in charts
        ],
        "cells": {},
        "cell_page": {
            "page": None,
            "page_size": _CELL_PAGE_SIZE,
            "total": 0,
            "returned": 0,
            "omitted": 0,
            "omitted_due_to_budget": 0,
        },
    }


def _iter_content_cells(sheet):
    for cell in sheet._cells.values():
        if getattr(cell, "value", None) is not None:
            yield cell


def _formula_text_and_metadata(
    value: object,
) -> tuple[str | None, dict | None]:
    if isinstance(value, str):
        return value, None
    if value is None:
        return None, None
    if type(value).__module__ != "openpyxl.worksheet.formula":
        return None, {"kind": "unsupported"}
    text = getattr(value, "text", None)
    if not isinstance(text, str):
        return None, {"kind": "unsupported"}
    raw_kind = getattr(value, "t", None)
    kind = (
        raw_kind
        if isinstance(raw_kind, str) and raw_kind
        else "formula"
    )
    metadata = {"kind": kind}
    ref = getattr(value, "ref", None)
    if isinstance(ref, str):
        metadata["ref"] = ref
    return text, metadata


def _cell_detail(formula_cell, cached_sheet) -> dict:
    cached_cell = cached_sheet._cells.get(
        (formula_cell.row, formula_cell.column)
    )
    cached_value = None if cached_cell is None else cached_cell.value
    formula = None
    formula_metadata = None
    if formula_cell.data_type == "f":
        formula, formula_metadata = _formula_text_and_metadata(
            formula_cell.value
        )
    detail = {
        "coordinate": formula_cell.coordinate,
        "value": _json_scalar(cached_value),
        "cached_value": _json_scalar(cached_value),
        "formula": formula,
        "number_format": formula_cell.number_format,
        "style_id": formula_cell.style_id,
    }
    if formula_metadata is not None:
        detail["formula_metadata"] = formula_metadata
    return detail


def _text_chars(value) -> int:
    if isinstance(value, str):
        return len(value)
    if isinstance(value, list):
        return sum(_text_chars(item) for item in value)
    if isinstance(value, dict):
        return sum(len(key) + _text_chars(item) for key, item in value.items())
    return 0


def _json_bytes(value) -> int:
    return len(
        json.dumps(
            value,
            ensure_ascii=False,
            allow_nan=False,
            separators=(",", ":"),
        ).encode("utf-8")
    )


def _selector_pages(
    selectors: list[str],
    sheet_names: set[str],
) -> dict[str, int]:
    pages: dict[str, int] = {}
    for selector in selectors:
        match = _CELL_SELECTOR_RE.fullmatch(selector)
        if match is None:
            raise InspectionError(
                "spreadsheet selector must be "
                "'sheet:<name>' or 'sheet:<name>:page:<n>'"
            )
        name = match.group(1)
        if name not in sheet_names:
            raise InspectionError(f"spreadsheet selector names missing sheet {name!r}")
        if name in pages:
            raise InspectionError(
                f"spreadsheet selector repeats sheet {name!r}"
            )
        pages[name] = int(match.group(2) or "1")
    return pages


def _build_structure(
    formula_book,
    cached_book,
    pages: dict[str, int | None],
    budget: ResponseBudget,
    *,
    enforce_extract_chars: bool,
) -> dict:
    cached_sheets = {sheet.title: sheet for sheet in cached_book.worksheets}
    rows: list[dict] = []
    source_sheets = {
        sheet.title: sheet for sheet in formula_book.worksheets
    }
    for sheet in formula_book._sheets:
        if isinstance(sheet, Worksheet):
            if sheet.title not in pages:
                continue
            total = sum(1 for _cell in _iter_content_cells(sheet))
            rows.append(_sheet_inventory(sheet, total, pages[sheet.title]))
        else:
            rows.append(_chartsheet_inventory(sheet))

    result = {"kind": "spreadsheet", "opens": True, "sheets": rows}
    used_bytes = _json_bytes(result)
    used_chars = _text_chars(result)
    if used_bytes > budget.max_bytes or (
        enforce_extract_chars and used_chars > budget.max_extract_chars
    ):
        return validate_json_result(
            result,
            budget,
            enforce_extract_chars=enforce_extract_chars,
        )

    byte_limit = max(0, budget.max_bytes - 128)
    char_limit = max(0, budget.max_extract_chars - 128)
    for row in rows:
        if row["type"] != "worksheet":
            continue
        page = row["cell_page"]["page"]
        if page is None:
            continue
        formula_sheet = source_sheets[row["name"]]
        cached_sheet = cached_sheets[row["name"]]
        start = (page - 1) * _CELL_PAGE_SIZE
        stop = start + _CELL_PAGE_SIZE
        for index, formula_cell in enumerate(
            _iter_content_cells(formula_sheet)
        ):
            if index < start:
                continue
            if index >= stop:
                break
            detail = _cell_detail(formula_cell, cached_sheet)
            coordinate = formula_cell.coordinate
            fragment = {coordinate: detail}
            byte_cost = _json_bytes(fragment) - 2
            if row["cells"]:
                byte_cost += 1
            char_cost = len(coordinate) + _text_chars(detail)
            if (
                used_bytes + byte_cost > byte_limit
                or (
                    enforce_extract_chars
                    and used_chars + char_cost > char_limit
                )
            ):
                row["cell_page"]["omitted_due_to_budget"] += 1
                continue
            row["cells"][coordinate] = detail
            row["cell_page"]["returned"] += 1
            used_bytes += byte_cost
            used_chars += char_cost
        row["cell_page"]["omitted"] = (
            row["cell_page"]["total"]
            - row["cell_page"]["returned"]
        )

    return validate_json_result(
        result,
        budget,
        enforce_extract_chars=enforce_extract_chars,
    )


def _prepare_xlsx(path: str, scratch_dir: str) -> str:
    if _is_legacy_xls(path):
        return _convert_with_libreoffice(path, scratch_dir, "xlsx")
    return path


def _validate_render_selectors(selectors: list[str]) -> None:
    for selector in selectors:
        if _PDF_SELECTOR_RE.fullmatch(selector) is None:
            raise InspectionError(
                "spreadsheet render selector must be 'page:<n>'"
            )


class SpreadsheetInspector:
    """Inspect and render XLS/XLSX artifacts within registry budgets."""

    def inspect(
        self,
        path: str,
        scratch_dir: str,
        *,
        response_budget: ResponseBudget,
    ):
        xlsx_path = _prepare_xlsx(path, scratch_dir)
        with _opened_workbooks(xlsx_path) as (formula_book, cached_book):
            pages = {sheet.title: 1 for sheet in formula_book.worksheets}
            return _build_structure(
                formula_book,
                cached_book,
                pages,
                response_budget,
                enforce_extract_chars=False,
            )

    def extract(
        self,
        path: str,
        scratch_dir: str,
        selectors: list[str],
        *,
        response_budget: ResponseBudget,
    ):
        xlsx_path = _prepare_xlsx(path, scratch_dir)
        with _opened_workbooks(xlsx_path) as (formula_book, cached_book):
            sheet_names = {
                sheet.title for sheet in formula_book.worksheets
            }
            selected = _selector_pages(selectors, sheet_names)
            pages: dict[str, int | None]
            if selected:
                pages = {
                    sheet.title: selected[sheet.title]
                    for sheet in formula_book.worksheets
                    if sheet.title in selected
                }
            else:
                pages = {
                    sheet.title: None
                    for sheet in formula_book.worksheets
                }
            return _build_structure(
                formula_book,
                cached_book,
                pages,
                response_budget,
                enforce_extract_chars=True,
            )

    def render(
        self,
        path: str,
        scratch_dir: str,
        selectors: list[str],
        budget: RenderBudget,
    ) -> list[str]:
        _validate_render_selectors(selectors)
        if not _is_legacy_xls(path):
            preflight_xlsx(path)
        pdf_path = _convert_with_libreoffice(path, scratch_dir, "pdf")
        try:
            from .pdf_image import PdfInspector
        except (ImportError, ModuleNotFoundError) as exc:
            raise InspectionError(
                "PDF renderer is unavailable for spreadsheet rendering"
            ) from exc
        return PdfInspector().render(pdf_path, scratch_dir, selectors, budget)


def _canonical_cell_ref(value: object) -> tuple[str, int, int]:
    if not isinstance(value, str):
        raise InspectionError("spreadsheet check cell must be a string")
    match = _CELL_REF_RE.fullmatch(value)
    if match is None:
        raise InspectionError(f"invalid spreadsheet cell reference {value!r}")
    row, column = coordinate_to_tuple(value.replace("$", ""))
    if row > 1_048_576 or column > 16_384:
        raise InspectionError(f"spreadsheet cell reference is out of range {value!r}")
    return f"{get_column_letter(column)}{row}", row, column


def _criterion(
    check: dict,
    sheet_name: str,
    cell_ref: str,
    passed: bool,
    reason: str,
) -> dict:
    return {
        "id": check["id"],
        "passed": passed,
        "score": 1.0 if passed else 0.0,
        "reason": reason,
        "evidence": [
            {
                "path": check["path"],
                "locator": f"sheet={sheet_name},cell={cell_ref}",
                "source": "structure",
            }
        ],
    }


def _formula_body(value: object) -> str | None:
    text, _metadata = _formula_text_and_metadata(value)
    if text is None:
        return None
    return text[1:] if text.startswith("=") else text


def _exact_cell_equal(actual: object, expected: object) -> bool:
    if isinstance(actual, bool) or isinstance(expected, bool):
        return type(actual) is type(expected) and actual == expected
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        if (
            isinstance(actual, float) and not math.isfinite(actual)
        ) or (
            isinstance(expected, float) and not math.isfinite(expected)
        ):
            return False
        return actual == expected
    return type(actual) is type(expected) and actual == expected


@contextmanager
def _check_xlsx_path(path: str):
    descriptor = _open_stable_descriptor(path, "spreadsheet check input")
    try:
        stable_path = f"/proc/self/fd/{descriptor}"
        if (
            os.pread(descriptor, len(_OLE_SIGNATURE), 0)
            != _OLE_SIGNATURE
        ):
            yield stable_path
            return
        with tempfile.TemporaryDirectory(
            prefix="skillopt-xls-check-"
        ) as scratch_root:
            with scratch_transaction(
                scratch_root,
                max_bytes=DEFAULT_SCRATCH_BYTES,
                max_entries=DEFAULT_SCRATCH_ENTRIES,
                max_depth=DEFAULT_SCRATCH_DEPTH,
            ) as transaction:
                yield _convert_with_libreoffice(
                    stable_path,
                    transaction.proc_path,
                    "xlsx",
                )
    finally:
        os.close(descriptor)


def evaluate_xlsx_check(path: str, check: dict) -> dict:
    """Evaluate one deterministic XLS/XLSX cell or formula criterion."""
    if not isinstance(check, dict):
        raise InspectionError("spreadsheet check must be an object")
    check_id = check.get("id")
    check_path = check.get("path")
    check_type = check.get("type")
    spec = check.get("spec")
    if not isinstance(check_id, str) or not check_id:
        raise InspectionError("spreadsheet check id must be a non-empty string")
    if not isinstance(check_path, str) or not check_path:
        raise InspectionError("spreadsheet check path must be a non-empty string")
    if check_type not in {"xlsx_cell", "xlsx_formula"}:
        raise InspectionError(f"unsupported spreadsheet check type {check_type!r}")
    if not isinstance(spec, dict):
        raise InspectionError("spreadsheet check spec must be an object")
    sheet_name = spec.get("sheet")
    if not isinstance(sheet_name, str) or not sheet_name:
        raise InspectionError("spreadsheet check sheet must be a non-empty string")
    cell_ref, row, column = _canonical_cell_ref(spec.get("cell"))

    with _check_xlsx_path(path) as xlsx_path:
        with _opened_workbooks(xlsx_path) as (formula_book, cached_book):
            formula_sheets = {
                sheet.title: sheet for sheet in formula_book.worksheets
            }
            cached_sheets = {
                sheet.title: sheet for sheet in cached_book.worksheets
            }
            if sheet_name not in formula_sheets:
                reason = (
                    f"sheet {sheet_name!r} is not a worksheet"
                    if sheet_name in formula_book.sheetnames
                    else f"worksheet {sheet_name!r} is missing"
                )
                return _criterion(
                    check,
                    sheet_name,
                    cell_ref,
                    False,
                    reason,
                )
            formula_sheet = formula_sheets[sheet_name]
            cached_sheet = cached_sheets[sheet_name]
            formula_cell = formula_sheet._cells.get((row, column))
            cached_cell = cached_sheet._cells.get((row, column))

            if check_type == "xlsx_formula":
                expected = spec.get("formula")
                if not isinstance(expected, str):
                    raise InspectionError(
                        "xlsx_formula check formula must be a string"
                    )
                actual_value = (
                    None if formula_cell is None else formula_cell.value
                )
                actual, actual_metadata = _formula_text_and_metadata(
                    actual_value
                )
                if actual is None:
                    detail = (
                        "unsupported formula object"
                        if actual_metadata is not None
                        else "missing or has no formula"
                    )
                    return _criterion(
                        check,
                        sheet_name,
                        cell_ref,
                        False,
                        f"cell {cell_ref} {detail}",
                    )
                passed = (
                    formula_cell.data_type == "f"
                    and _formula_body(actual) == _formula_body(expected)
                )
                reason = (
                    "formula matches exactly"
                    if passed
                    else "formula mismatch: expected "
                    f"{bounded_diagnostic(expected, 200)!r}, found "
                    f"{bounded_diagnostic(actual, 200)!r}"
                )
            else:
                if "value" not in spec:
                    raise InspectionError(
                        "xlsx_cell check value is required"
                    )
                if formula_cell is None and cached_cell is None:
                    return _criterion(
                        check,
                        sheet_name,
                        cell_ref,
                        False,
                        f"cell {cell_ref} is missing",
                    )
                actual = (
                    cached_cell.value
                    if cached_cell is not None
                    else formula_cell.value
                )
                if actual is None:
                    return _criterion(
                        check,
                        sheet_name,
                        cell_ref,
                        False,
                        f"cell {cell_ref} is missing or blank",
                    )
                expected = spec["value"]
                passed = _exact_cell_equal(actual, expected)
                reason = (
                    "cell value matches exactly"
                    if passed
                    else "cell value mismatch: expected "
                    f"{bounded_diagnostic(expected, 200)!r}, found "
                    f"{bounded_diagnostic(actual, 200)!r}"
                )
            return _criterion(
                check,
                sheet_name,
                cell_ref,
                passed,
                reason,
            )


__all__ = [
    "SpreadsheetInspector",
    "evaluate_xlsx_check",
    "preflight_xlsx",
]
